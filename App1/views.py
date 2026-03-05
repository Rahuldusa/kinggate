from django.contrib import messages
from django.forms import BooleanField, CharField
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt


from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect
from pytz import timezone
from .models import *

from django.contrib.auth.hashers import check_password
from openai import OpenAI
import math
from django.db.models import Sum
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings as django_settings
import random

def login(request):
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        try:
            user = Custom_User.objects.get(username=username)
        except Custom_User.DoesNotExist:
            # Log failed login attempt - user not found
            user_logs.objects.create(
                user=username,
                action='Failed Login - User Not Found',
                details=f'IP: {get_client_ip(request)}, User Agent: {request.META.get("HTTP_USER_AGENT", "Unknown")}'
            )
            messages.error(request, "Invalid username or password")
            return render(request, 'login.html')

        
        if not check_password(password, user.password):
            # Log failed login attempt - wrong password
            user_logs.objects.create(
                user=user.username,
                action='Failed Login - Wrong Password',
                details=f'IP: {get_client_ip(request)}, User Agent: {request.META.get("HTTP_USER_AGENT", "Unknown")}'
            )
            messages.error(request, "Invalid username or password")
            return render(request, 'login.html')

        # Log successful login
        user_logs.objects.create(
            user=user.username,
            action='Successful Login',
            details=f'IP: {get_client_ip(request)}, User Agent: {request.META.get("HTTP_USER_AGENT", "Unknown")}'
        )
        
        request.session['user_id'] = user.id
        request.session['username'] = user.username

        messages.success(request, f"Welcome {user.username}!")
        return redirect('dashboard')

    return render(request, 'login.html')


# Helper function to get client IP
def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip
def logout_view(request):
    request.session.flush()
    messages.success(request, "You have been logged out successfully")
    return redirect('login')


def forgot_password(request):
    if request.method == 'POST':
        email = request.POST.get('email')

        try:
            user = Custom_User.objects.get(email=email)
        except Custom_User.DoesNotExist:
            messages.error(request, "Email not registered")
            return render(request, 'forgot_password.html')

        otp = random.randint(100000, 999999)

        request.session['reset_otp'] = str(otp)
        request.session['reset_user_id'] = user.id

        send_mail(
            subject='Password Reset OTP',
            message=f'Your OTP for password reset is {otp}',
            from_email=django_settings.EMAIL_HOST_USER,
            recipient_list=[email],
            fail_silently=False
        )

        messages.success(request, "OTP sent to your email")
        return redirect('verify_otp')

    return render(request, 'forgot_password.html')

def verify_otp(request):
    if request.method == 'POST':
        otp_entered = request.POST.get('otp')
        otp_session = request.session.get('reset_otp')

        if otp_entered == otp_session:
            return redirect('reset_password')
        else:
            messages.error(request, "Invalid OTP")

    return render(request, 'verify_otp.html')

def reset_password(request):
    user_id = request.session.get('reset_user_id')

    if not user_id:
        return redirect('login')

    if request.method == 'POST':
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')

        if password != confirm_password:
            messages.error(request, "Passwords do not match")
            return render(request, 'reset_password.html')

        user = Custom_User.objects.get(id=user_id)
        user.password = password
        user.save()

        request.session.flush()
        messages.success(request, "Password reset successful. Please login.")
        return redirect('login')

    return render(request, 'reset_password.html')


from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from django.core.mail import EmailMultiAlternatives
from django.utils.html import strip_tags
from django.conf import settings
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, KeepTogether
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_LEFT
from io import BytesIO
import os

from django.shortcuts import render, redirect
from django.contrib import messages
from .utils import get_filtered_companies, get_company_filter_context

from django.shortcuts import render, redirect
from django.contrib import messages
from .utils import get_filtered_companies, get_company_filter_context
from datetime import datetime, timedelta, date
from django.utils import timezone
from django.core.mail import EmailMultiAlternatives
from django.template.defaultfilters import strip_tags
from django.conf import settings
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from io import BytesIO
from decimal import Decimal  # ← ADD THIS IMPORT
import os
import calendar

def customer_management(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    
    if request.method == "POST":
        customer_name = request.POST.get("customer_name")
        email = request.POST.get("email")
        contacts = request.POST.getlist("contacts[]")
        state_id = request.POST.get("state") or None
        city_id = request.POST.get("city") or None
        area_id = request.POST.get("area") or None
        pincode = request.POST.get("pincode")
        address = request.POST.get("address")
        gst_registered = request.POST.get("gst_type") == "GST"
        start_date = request.POST.get("start_date")
        status = request.POST.get("status")
        # FIX: Convert empty string to None for subscription_plan_id
        subscription_plan_id = request.POST.get("subscription_plan") or None
        location_link = request.POST.get("location_link")
        remarks = request.POST.get("remarks")
        # FIX: Convert empty string to None for company_id
        company_id = request.POST.get("company") or None
        tax_percentage = request.POST.get("tax_percentage")
        advance = request.POST.get("advance_amount")
        custom_subscription_amount = request.POST.get("custom_subscription_amount")
        gst_number = request.POST.get("gst_number")
        outstanding = request.POST.get("outstanding_amount")
        customer = Customer.objects.create(
            name=customer_name,
            email=email,
            address=address,
            company_id=company_id,
            state_id=state_id,
            city_id=city_id,
            area_id=area_id,
            pincode=pincode,
            gst_registered=gst_registered,
            gst_number=gst_number if gst_number else None,
            start_date=start_date,
            status=status,
            subscription_plan_id=subscription_plan_id,
            location_link=location_link,
            remarks=remarks,
            tax_percentage=tax_percentage if tax_percentage else 0.0,
            advance_amount=advance if advance else 0.0,
            outstanding_amount=outstanding if outstanding else 0.0,
            custom_subscription_amount=custom_subscription_amount if custom_subscription_amount else None
        )

        # Bulk create contacts
        contact_objects = [
            CustomerContact(customer=customer, phone_number=number)
            for number in contacts if number.strip()
        ]
        if contact_objects:
            CustomerContact.objects.bulk_create(contact_objects)

        camera_serials = request.POST.getlist("camera_serials[]")

        if subscription_plan_id:
            plan = SubscriptionPlan.objects.get(id=subscription_plan_id)
            if plan.name.lower() == "live monitoring":
                camera_objects = [
                    customer_cameras(customer=customer, seriak_number=serial)
                    for serial in camera_serials if serial.strip()
                ]
                if camera_objects:
                    customer_cameras.objects.bulk_create(camera_objects)
        
        # Log customer creation
        user_logs.objects.create(
            user=user.username,
            action='Created Customer',
            details=f'Customer Name: {customer_name}'
        )
        
        try:
            generate_automatic_bill(
                entity=customer,
                entity_type="customer",
                user=user
            )
            messages.success(request, f'Customer "{customer_name}" added successfully and bills generated!')
        except Exception as e:
            messages.warning(request, f'Customer added but billing failed: {str(e)}')
            import traceback
            traceback.print_exc()

        return redirect("customer_management")

    else:
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        
        # Get filtered companies from the navbar filter
        filtered_companies = get_filtered_companies(request)
        company_ids = list(filtered_companies.values_list('id', flat=True))
        
        # Check if filters are actively being used
        selected_gst_type = request.session.get('selected_gst_type', 'gst')
        selected_company_ids = request.session.get('selected_company_ids', [])
        filters_active = selected_gst_type or selected_company_ids
        
        # Filter customers based on company filter
        if company_ids:
            customers = Customer.objects.filter(company_id__in=company_ids)
        elif filters_active:
            # ✅ Filters are active but returned no companies - show no customers
            customers = Customer.objects.none()
        else:
            # No filters applied - show all customers
            customers = Customer.objects.all()
            
        if start_date:
            customers = customers.filter(start_date__gte=start_date)
        if end_date:
            customers = customers.filter(start_date__lte=end_date)
            
        customers = customers.select_related(
            'state', 'city', 'area', 'company', 'subscription_plan'
        ).prefetch_related(
            'contacts', 'customer_cameras_set'
        ).all()
        
        states = State.objects.all()
        
        # Filter companies for the dropdown based on GST type
        companies = Company.objects.filter(status='Active')
        if selected_gst_type == 'gst':
            companies = companies.filter(gst_registered=True)
        elif selected_gst_type == 'non-gst':
            companies = companies.filter(gst_registered=False)
        
        companies = companies.order_by('name')
        
        citys = City.objects.all()
        areas = Area.objects.all()
        plans = SubscriptionPlan.objects.all()
        
        context = {
            'customers': customers, 
            'subscription_plans': plans, 
            'states': states, 
            'citys': citys, 
            'areas': areas,
            'companies': companies,
            'filtered_company_ids': company_ids,
            'user': user,
            'role_permissions': role_permissions,
            'start_date': start_date,
            'end_date': end_date,
            'selected_gst_type': selected_gst_type,
        }
        
        context.update(get_company_filter_context(request))
        
        return render(request, 'customermgmt.html', context)
    
def adjust_advance_amount(customer, bill_amount):
    customer.refresh_from_db()
    
    bill_amount = Decimal(str(bill_amount))
    advance_amount = Decimal(str(customer.advance_amount or 0))
    
    if advance_amount <= 0:
        return {
            'original_amount': bill_amount,
            'advance_used': Decimal('0.00'),
            'final_amount': bill_amount,
            'remaining_advance': Decimal('0.00'),
            'paid_amount': Decimal('0.00'),
            'balance_amount': bill_amount,
        }
    
    advance_used = min(advance_amount, bill_amount)
    final_amount = bill_amount - advance_used
    remaining_advance = advance_amount - advance_used
    
    customer.advance_amount = remaining_advance
    customer.save(update_fields=['advance_amount'])
    customer.refresh_from_db()
    
    return {
        'original_amount': bill_amount,
        'advance_used': advance_used,
        'final_amount': final_amount,
        'remaining_advance': remaining_advance,
        'paid_amount': advance_used,
        'balance_amount': final_amount,
    }


def generate_automatic_bill(entity, entity_type, user):
    
    if not entity.start_date or not entity.subscription_plan:
        return
    
    company = getattr(entity, 'company', None)
    if not company:
        return
    entity.refresh_from_db()
    
    if isinstance(entity.start_date, str):
        start_date = datetime.strptime(entity.start_date, '%Y-%m-%d').date()
    else:
        start_date = entity.start_date

    today = datetime.now().date()
    first_day_current_month = today.replace(day=1)
    last_day_last_month = first_day_current_month - timedelta(days=1)
    
    plan = entity.subscription_plan
    
    if start_date >= first_day_current_month:
        return

    if hasattr(entity, 'custom_subscription_amount') and entity.custom_subscription_amount:
        monthly_price = Decimal(str(entity.custom_subscription_amount))
    else:
        monthly_price = Decimal(str(plan.price))

    # -------------------------------------------------------
    # GST applicability check
    # Check if the company associated with the customer is GST registered.
    # -------------------------------------------------------
    is_gst_applicable = True
    if entity_type == "customer":
        is_gst_applicable = getattr(company, 'gst_registered', False)

    is_same_state = False
    if is_gst_applicable and company and entity.state:
        if hasattr(company, 'state') and company.state:
            is_same_state = (company.state.id == entity.state.id)
        elif hasattr(company, 'state_name'):
            is_same_state = (str(company.state_name).lower() == str(entity.state).lower())

    initial_outstanding = Decimal(str(getattr(entity, 'outstanding_amount', 0) or 0))
    
    current = start_date
    bills_generated = 0
    generated_bills = []

    while current <= last_day_last_month:
        entity.refresh_from_db()
        
        year = current.year
        month = current.month
        days_in_month = calendar.monthrange(year, month)[1]

        if year == start_date.year and month == start_date.month:
            billing_start = start_date
            billing_end = date(year, month, days_in_month)
            remaining_days = days_in_month - start_date.day + 1
            base_amount = (monthly_price / Decimal(str(days_in_month))) * Decimal(str(remaining_days))
        else:
            billing_start = date(year, month, 1)
            billing_end = date(year, month, days_in_month)
            base_amount = monthly_price

        base_amount = base_amount.quantize(Decimal('0.01'))

        # -------------------------------------------------------
        # GST Calculation — skipped entirely if company is not GST registered
        # -------------------------------------------------------
        cgst = sgst = igst = Decimal('0.00')

        if is_gst_applicable:
            if is_same_state:
                cgst = (base_amount * Decimal('0.09')).quantize(Decimal('0.01'))
                sgst = (base_amount * Decimal('0.09')).quantize(Decimal('0.01'))
                gst_type = 'INTRA-STATE'
            else:
                igst = (base_amount * Decimal('0.18')).quantize(Decimal('0.01'))
                gst_type = 'INTER-STATE'
        else:
            gst_type = 'NON-GST'

        total_gst = cgst + sgst + igst
        subtotal_before_outstanding = base_amount + total_gst

        outstanding_charge = Decimal('0.00')
        if bills_generated == 0 and entity_type == "customer":
            outstanding_charge = Decimal(str(getattr(entity, 'outstanding_amount', 0) or 0))
            if outstanding_charge > 0:
                entity.outstanding_amount = Decimal('0.00')
                entity.save(update_fields=['outstanding_amount'])
                entity.refresh_from_db()

        total_bill_amount = subtotal_before_outstanding + outstanding_charge

        if entity_type == "customer" and hasattr(entity, 'advance_amount'):
            advance_adjustment = adjust_advance_amount(entity, total_bill_amount)
            paid_amount = advance_adjustment['paid_amount']
            balance_amount = advance_adjustment['balance_amount']
            discount_amount = advance_adjustment['advance_used']
        else:
            paid_amount = Decimal('0.00')
            balance_amount = total_bill_amount
            discount_amount = Decimal('0.00')
            advance_adjustment = {
                'original_amount': total_bill_amount,
                'advance_used': Decimal('0.00'),
                'final_amount': total_bill_amount,
                'remaining_advance': Decimal('0.00'),
                'paid_amount': Decimal('0.00'),
                'balance_amount': total_bill_amount,
            }

        # -------------------------------------------------------
        # If advance was used, set payment_date and payment_mode
        # on the BillingRecord so it reflects the adjustment date.
        # -------------------------------------------------------
        billing_kwargs = {
            "amount": total_bill_amount,
            "discount_amount": discount_amount,
            "gst_amount": total_gst,
            "gst_type": gst_type,
            "billing_date": today,
            "due_date": today + timedelta(days=15),
            "paid": balance_amount <= 0,
            "paid_amount": paid_amount,
            "balance_amount": balance_amount,
            "billing_start_date": billing_start,
            "billing_end_date": billing_end,
            "payment_date": today if paid_amount > 0 else None,          # ← SET today if advance used
            "payment_mode": 'Advance Adjustment' if paid_amount > 0 else None,  # ← SET mode if advance used
            "notes": (
                f"{entity_type.title()} subscription bill "
                f"{billing_start.strftime('%d-%b-%Y')} "
                f"to {billing_end.strftime('%d-%b-%Y')}"
            )
        }

        if entity_type == "customer":
            billing_kwargs["customer"] = entity
        else:
            billing_kwargs["company"] = entity

        billing_record = BillingRecord.objects.create(**billing_kwargs)

        amounts = {
            'base_amount': f"{base_amount:.2f}",
            'cgst': f"{cgst:.2f}",
            'sgst': f"{sgst:.2f}",
            'igst': f"{igst:.2f}",
            'total_gst': f"{total_gst:.2f}",
            'subtotal': f"{subtotal_before_outstanding:.2f}",
            'outstanding_charge': f"{outstanding_charge:.2f}",
            'advance_used': f"{discount_amount:.2f}",
            'total_amount': f"{balance_amount:.2f}",
            'is_gst_applicable': is_gst_applicable,
        }

        gst_details = {
            'is_same_state': is_same_state,
            'is_gst_applicable': is_gst_applicable,
        }

        if balance_amount <= 0:
            bill_status = 'Paid'
            bill_payment_mode = 'Advance Adjustment'
            bill_payment_date = today
        else:
            bill_status = 'Pending'
            bill_payment_mode = None
            bill_payment_date = None

        # ── LOG 1: Bill generation entry (always created) ──────────────────
        data_logs.objects.create(
            user=user,
            timestamp=timezone.now(),
            customer=entity if entity_type == "customer" else None,
            location=None,
            billing_record=billing_record,
            payment_amount=Decimal('0.00'),
            billing_period_start=billing_start,
            billing_period_end=billing_end,
            balance_amount=balance_amount,
            is_payment=False,
            is_approved=False,
            status=bill_status,
            payment_mode=None,
            payment_date=None,
            total_paid=Decimal('0.00'),
            action=f"Auto-generated bill for {billing_start.strftime('%B %Y')}",
            details=(
                f"Subscription: {plan.name}, "
                f"Base: Rs.{base_amount:.2f}, "
                f"GST: Rs.{total_gst:.2f}, "
                f"Outstanding: Rs.{outstanding_charge:.2f}, "
                f"Total: Rs.{total_bill_amount:.2f}, "
                f"Balance: Rs.{balance_amount:.2f}"
            )
        )

        # ── LOG 2: Advance adjustment log (only if advance was used) ────────
        if paid_amount > 0:
            data_logs.objects.create(
                user=user,
                timestamp=timezone.now(),
                customer=entity if entity_type == "customer" else None,
                location=None,
                billing_record=billing_record,
                payment_amount=paid_amount,
                billing_period_start=billing_start,
                billing_period_end=billing_end,
                balance_amount=balance_amount,
                is_payment=True,                   
                status=bill_status,
                payment_mode='Advance Adjustment',
                payment_date=today,                   # ← Payment date is today
                total_paid=paid_amount,
                action=f"Advance adjusted for {billing_start.strftime('%B %Y')}",
                details=(
                    f"Advance Adjustment for Invoice: {billing_record.invoice_number or billing_record.id}, "
                    f"Subscription: {plan.name}, "
                    f"Bill Total: Rs.{total_bill_amount:.2f}, "
                    f"Advance Used: Rs.{paid_amount:.2f}, "
                    f"Remaining Advance: Rs.{advance_adjustment['remaining_advance']:.2f}, "
                    f"Balance After Adjustment: Rs.{balance_amount:.2f}"
                )
            )

        days_diff = (billing_end - billing_start).days + 1

        pdf_buffer = generate_invoice_pdf(
            billing_record=billing_record,
            subscription_plan=plan,
            entity=entity,
            gst_details=gst_details,
            amounts=amounts,
            company=company,
            billing_period=f"{billing_start.strftime('%d-%b-%Y')} to {billing_end.strftime('%d-%b-%Y')}",
            months_count=1,
            days_count=days_diff,
            advance_adjustment=advance_adjustment,
            monthly_price=monthly_price
        )

        generated_bills.append({
            'billing_record': billing_record,
            'pdf_buffer': pdf_buffer,
            'billing_start': billing_start,
            'billing_end': billing_end,
            'amounts': amounts,
            'gst_details': gst_details,
            'days_count': days_diff,
            'advance_adjustment': advance_adjustment,
            'monthly_price': monthly_price
        })

        bills_generated += 1

        if month == 12:
            current = date(year + 1, 1, 1)
        else:
            current = date(year, month + 1, 1)

    if generated_bills and entity.email and company:
        try:
            send_consolidated_billing_email(
                recipient_email=entity.email,
                recipient_name=entity.name,
                generated_bills=generated_bills,
                subscription_plan=plan,
                company=company,
                entity=entity
            )
        except Exception as e:
            import traceback
            traceback.print_exc()
    else:
        print(f"[BILLING] Email skipped: bills={len(generated_bills)}, email={entity.email}, company={company}")

def send_consolidated_billing_email(recipient_email, recipient_name, generated_bills,
                                    subscription_plan, company, entity=None):

    if len(generated_bills) == 1:
        bill = generated_bills[0]
        subject = f'Invoice IN-{bill["billing_record"].id} - {subscription_plan.name} Plan'
    else:
        invoice_numbers = [f"IN-{bill['billing_record'].id}" for bill in generated_bills]
        subject = f'Invoices {", ".join(invoice_numbers)} - {subscription_plan.name} Plan'

    total_billed = sum(Decimal(bill['amounts']['subtotal']) for bill in generated_bills)
    total_paid = sum(Decimal(bill['amounts']['advance_used']) for bill in generated_bills)
    total_balance = sum(Decimal(bill['amounts']['total_amount']) for bill in generated_bills)

    invoices_html = ""
    for idx, bill in enumerate(generated_bills, 1):
        br = bill['billing_record']
        amounts = bill['amounts']
        gst_details = bill['gst_details']
        is_gst_applicable = gst_details.get('is_gst_applicable', True)

        # -------------------------------------------------------
        # GST breakdown section — hidden for non-GST customers
        # -------------------------------------------------------
        gst_breakdown = ''
        if is_gst_applicable:
            if gst_details['is_same_state']:
                gst_breakdown = f"""
                <div class="detail-row">
                    <span class="detail-label">CGST (9%):</span>
                    <span>₹{amounts['cgst']}</span>
                </div>
                <div class="detail-row">
                    <span class="detail-label">SGST (9%):</span>
                    <span>₹{amounts['sgst']}</span>
                </div>
                """
            else:
                gst_breakdown = f"""
                <div class="detail-row">
                    <span class="detail-label">IGST (18%):</span>
                    <span>₹{amounts['igst']}</span>
                </div>
                """
        # No gst_breakdown block at all for non-GST customers

        advance_section = ""
        if Decimal(amounts['advance_used']) > 0:
            advance_section = f"""
            <div class="detail-row" style="border-top: 2px solid #dee2e6; margin-top: 10px; padding-top: 10px;">
                <span class="detail-label">Subtotal:</span>
                <span>₹{amounts['subtotal']}</span>
            </div>
            <div class="detail-row">
                <span class="detail-label" style="color: #28a745;">Advance Adjusted:</span>
                <span style="color: #28a745;">- ₹{amounts['advance_used']}</span>
            </div>
            """

        period_desc = f"{bill['billing_start'].strftime('%d-%b-%Y')} to {bill['billing_end'].strftime('%d-%b-%Y')}"
        if bill.get('days_count'):
            period_desc += f" ({bill['days_count']} days)"

        # Use "Tax Invoice" label only when GST is applicable
        invoice_label = "Tax Invoice" if is_gst_applicable else "Invoice"

        invoices_html += f"""
        <div class="invoice-details">
            <h3 style="color: #3d6a8e; margin-top: 0;">{invoice_label} #{idx}: IN-{br.id}</h3>
            <div class="detail-row">
                <span class="detail-label">Billing Period:</span>
                <span>{period_desc}</span>
            </div>
            <div class="detail-row">
                <span class="detail-label">Invoice Date:</span>
                <span>{br.billing_date.strftime('%d-%b-%Y')}</span>
            </div>
            <div class="detail-row">
                <span class="detail-label">Due Date:</span>
                <span>{br.due_date.strftime('%d-%b-%Y')}</span>
            </div>
            <div class="detail-row">
                <span class="detail-label">Base Amount:</span>
                <span>₹{amounts['base_amount']}</span>
            </div>
            {gst_breakdown}
            {advance_section}
            <div class="detail-row" style="border-top: 2px solid #3d6a8e; margin-top: 10px; padding-top: 10px;">
                <span class="detail-label"><strong>Amount Due:</strong></span>
                <span><strong>₹{amounts['total_amount']}</strong></span>
            </div>
        </div>
        """

    # Show GST info block only for GST-registered customers
    first_gst = generated_bills[0]['gst_details']
    is_gst_applicable = first_gst.get('is_gst_applicable', True)

    if is_gst_applicable:
        gst_info_block = f"""
        <div class="gst-info">
            <strong>GST Information:</strong><br>
            {'Same State Transaction - CGST (9%) + SGST (9%) = 18%' if first_gst['is_same_state'] else 'Interstate Transaction - IGST (18%)'}
        </div>
        """
    else:
        gst_info_block = """
        <div class="gst-info" style="background-color:#f0f0f0; border-left-color:#999;">
            <strong>Note:</strong> This customer is not GST registered. No GST has been applied.
        </div>
        """

    # Invoice label for the header (Tax Invoice vs Invoice)
    invoice_header_label = "Tax Invoice" if is_gst_applicable else "Invoice"

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{
                font-family: Arial, sans-serif;
                line-height: 1.6;
                color: #333;
                max-width: 700px;
                margin: 0 auto;
                padding: 20px;
            }}
            .header {{
                background-color: #3d6a8e;
                color: white;
                padding: 20px;
                text-align: center;
                border-radius: 5px 5px 0 0;
            }}
            .content {{
                background-color: #f8f9fa;
                padding: 30px;
                border: 1px solid #dee2e6;
            }}
            .invoice-details {{
                background-color: white;
                padding: 20px;
                border-radius: 5px;
                margin: 20px 0;
                border-left: 4px solid #3d6a8e;
            }}
            .detail-row {{
                display: flex;
                justify-content: space-between;
                padding: 8px 0;
                border-bottom: 1px solid #e9ecef;
            }}
            .detail-row:last-child {{ border-bottom: none; }}
            .detail-label {{ font-weight: bold; color: #6c757d; }}
            .summary-box {{
                background-color: #e7f1ff;
                padding: 20px;
                border-radius: 5px;
                margin: 20px 0;
            }}
            .summary-row {{
                display: flex;
                justify-content: space-between;
                padding: 8px 0;
            }}
            .total-amount {{ font-size: 24px; font-weight: bold; color: #3d6a8e; }}
            .footer {{
                text-align: center;
                color: #6c757d;
                font-size: 12px;
                margin-top: 30px;
                padding-top: 20px;
                border-top: 1px solid #dee2e6;
            }}
            .attachment-note {{
                background-color: #fff3cd;
                border-left: 4px solid #ffc107;
                padding: 15px;
                margin: 20px 0;
            }}
            .gst-info {{
                background-color: #d1ecf1;
                border-left: 4px solid #0c5460;
                padding: 15px;
                margin: 20px 0;
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>{company.name}</h1>
            <p>{invoice_header_label}{'s' if len(generated_bills) > 1 else ''} Notification</p>
        </div>

        <div class="content">
            <p>Dear {recipient_name},</p>
            <p>{'Invoices have' if len(generated_bills) > 1 else 'An invoice has'} been generated for your subscription with {company.name}.</p>

            {gst_info_block}
            {invoices_html}

            <div class="summary-box">
                <div class="summary-row">
                    <span><strong>Total Billed:</strong></span>
                    <span>₹{total_billed:.2f}</span>
                </div>
                {f'''<div class="summary-row">
                    <span><strong>Paid (Advance):</strong></span>
                    <span style="color: #28a745;">₹{total_paid:.2f}</span>
                </div>''' if total_paid > 0 else ''}
                <div class="summary-row" style="border-top: 2px solid #3d6a8e; padding-top: 12px; margin-top: 12px;">
                    <span class="total-amount">Total Amount Due:</span>
                    <span class="total-amount">₹{total_balance:.2f}</span>
                </div>
                {f'''<div class="summary-row" style="margin-top: 10px; font-size: 14px;">
                    <span>Remaining Advance Balance:</span>
                    <span style="color: #28a745;">₹{entity.advance_amount:.2f}</span>
                </div>''' if entity and hasattr(entity, 'advance_amount') and entity.advance_amount and entity.advance_amount > 0 else ''}
            </div>

            <div class="attachment-note">
                <strong>📎 {len(generated_bills)} PDF Invoice{'s' if len(generated_bills) > 1 else ''} Attached</strong><br>
                Please find your detailed invoice{'s' if len(generated_bills) > 1 else ''} attached to this email.
            </div>

            <p><strong>Payment Instructions:</strong></p>
            <ul>
                <li>Payment is due by {generated_bills[-1]['billing_record'].due_date.strftime('%d-%b-%Y')}</li>
                <li>Please include invoice reference{'s' if len(generated_bills) > 1 else ''} in the payment description</li>
                {f"<li>Scan the QR code in the attached PDF for quick payment</li>" if hasattr(company, 'qr_code') and company.qr_code else ""}
                <li>Contact us at {company.email} or {company.contact} if you have any questions</li>
            </ul>

            <p><strong>Subscription Plan:</strong> {subscription_plan.name}</p>
            <p>Thank you for choosing {company.name}!</p>
            <p>Best regards,<br><strong>{company.name} Team</strong></p>
        </div>

        <div class="footer">
            <p>This is an automated message. Please do not reply to this email.</p>
            <p>{company.name} | {company.contact} | {company.email}</p>
            <p>&copy; {datetime.now().year} {company.name}. All rights reserved.</p>
        </div>
    </body>
    </html>
    """

    text_content = strip_tags(html_content)

    email = EmailMultiAlternatives(
        subject=subject,
        body=text_content,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[recipient_email]
    )
    email.attach_alternative(html_content, "text/html")

    for bill in generated_bills:
        invoice_filename = f"Invoice_IN-{bill['billing_record'].id}.pdf"
        email.attach(invoice_filename, bill['pdf_buffer'].getvalue(), 'application/pdf')

    email.send()


def generate_invoice_pdf(billing_record, subscription_plan, entity, gst_details, amounts,
                         company, billing_period, months_count, days_count=None,
                         advance_adjustment=None, monthly_price=None):
    """Generate PDF invoice. GST columns are omitted for non-GST registered customers."""
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    PAGE_W = A4[0] - doc.leftMargin - doc.rightMargin
    styles = getSampleStyleSheet()

    small_left = ParagraphStyle(
        "small_left", parent=styles["Normal"],
        fontSize=9, leading=12, alignment=TA_LEFT
    )
    small_right = ParagraphStyle(
        "small_right", parent=styles["Normal"],
        fontSize=9, leading=12, alignment=TA_RIGHT
    )
    desc_style = ParagraphStyle(
        "desc", parent=styles["Normal"],
        fontSize=8, leading=10, alignment=TA_LEFT
    )

    elements = []

    is_gst_applicable = gst_details.get('is_gst_applicable', True)

    def safe_str(value):
        if not value:
            return ""
        if hasattr(value, "name"):
            return str(value.name)
        if hasattr(value, "city_name"):
            return str(value.city_name)
        if hasattr(value, "state_name"):
            return str(value.state_name)
        return str(value)

    # -------------------------------------------------------
    # HEADER
    # -------------------------------------------------------
    logo = ""
    if getattr(company, "logo", None):
        try:
            if hasattr(company.logo, 'path') and os.path.exists(company.logo.path):
                logo = Image(company.logo.path, 1.2 * inch, 1.2 * inch, kind='proportional')
        except Exception as e:
            print(f"Logo error: {e}")

    company_name_block = Paragraph(
        f"""
        <font size="16"><b>{company.name}</b></font><br/>
        <font size="9">
            {safe_str(getattr(company, "address", ""))}<br/>
            Contact: {safe_str(getattr(company, "contact", ""))}<br/>
            GSTIN: {safe_str(getattr(company, "gst_number", ""))}
        </font>
        """,
        small_left
    )

    # Use "TAX INVOICE" only when GST applies, otherwise plain "INVOICE"
    invoice_title = "TAX INVOICE" if is_gst_applicable else "INVOICE"

    invoice_block = Paragraph(
        f"""
        <b>{invoice_title}</b><br/>
        <font size="9">Original for Recipient</font><br/><br/>
        <b>IN-{billing_record.id}</b>
        """,
        ParagraphStyle(
            "hdr",
            fontSize=14,
            alignment=TA_RIGHT,
            fontName="Helvetica-Bold"
        )
    )

    header = Table(
        [[logo, company_name_block, invoice_block]],
        colWidths=[1.5 * inch, PAGE_W * 0.45, PAGE_W * 0.30]
    )
    header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(header)
    elements.append(Spacer(1, 10))

    # -------------------------------------------------------
    # Amount Due Bar
    # -------------------------------------------------------
    final_amount = amounts.get('total_amount', '0.00')
    amount_due = Table(
        [[
            Paragraph("<b>Amount Due</b>", small_left),
            Paragraph(
                f"<b>₹ {final_amount}</b>",
                ParagraphStyle(
                    "amt", fontSize=14,
                    alignment=TA_RIGHT,
                    textColor=colors.white,
                    fontName="Helvetica-Bold"
                )
            )
        ]],
        colWidths=[PAGE_W * 0.55, PAGE_W * 0.45]
    )
    amount_due.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#2f5f7a")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(amount_due)
    elements.append(Spacer(1, 12))

    # -------------------------------------------------------
    # Company Info + Invoice Dates
    # -------------------------------------------------------
    company_block = Paragraph(
        f"""
        <b>{company.name}</b><br/>
        {safe_str(getattr(company, "address", ""))}<br/>
        GSTIN: {safe_str(getattr(company, "gst_number", ""))}
        """,
        small_left
    )
    date_block = Paragraph(
        f"""
        Issue Date: {billing_record.billing_date.strftime('%d-%b-%Y')}<br/>
        Due Date: {billing_record.due_date.strftime('%d-%b-%Y')}
        """,
        small_right
    )
    company_info = Table([[company_block, date_block]], colWidths=[PAGE_W * 0.62, PAGE_W * 0.38])
    company_info.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(company_info)
    elements.append(Spacer(1, 10))

    # -------------------------------------------------------
    # Client Details
    # -------------------------------------------------------
    client_name_block = Paragraph(
        f"""
        <b>Client Details</b><br/>
        <b>{entity.name}</b><br/>
        Email: {safe_str(getattr(entity, 'email', ''))}
        """,
        small_left
    )
    billing_address_block = Paragraph(
        f"""
        <b>Billing Address</b><br/>
        {safe_str(entity.address)}<br/>
        {safe_str(entity.city)}, {safe_str(entity.state)}<br/>
        {safe_str(entity.pincode)}
        """,
        small_right
    )
    client_billing_row = Table(
        [[client_name_block, billing_address_block]],
        colWidths=[PAGE_W * 0.5, PAGE_W * 0.5]
    )
    client_billing_row.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))
    elements.append(client_billing_row)
    elements.append(Spacer(1, 12))

    # -------------------------------------------------------
    # Main Billing Table
    # GST columns are included only when is_gst_applicable is True
    # -------------------------------------------------------
    headers = ["S.No", "Item Description", "Price", "Taxable"]

    if is_gst_applicable:
        if gst_details["is_same_state"]:
            headers += ["CGST @ 9%", "SGST @ 9%"]
        else:
            headers += ["IGST @ 18%"]

    headers.append("Amount")

    period_description = f"{subscription_plan.name}"
    if monthly_price and hasattr(entity, 'custom_subscription_amount') and entity.custom_subscription_amount:
        period_description += f" (Custom: ₹{monthly_price}/month)"
    if days_count:
        period_description += f" ({days_count} days)"
    elif months_count > 1:
        period_description += f" ({months_count} months)"

    item_description = f"{period_description}<br/>Period: {billing_period}"

    row = [
        "1",
        Paragraph(item_description, desc_style),
        amounts['base_amount'],
        amounts['base_amount'],
    ]

    if is_gst_applicable:
        if gst_details["is_same_state"]:
            row += [amounts['cgst'], amounts['sgst']]
        else:
            row += [amounts['igst']]

    # For non-GST customers, subtotal == base_amount (no GST added)
    row.append(amounts['subtotal'])

    num_cols = len(headers)

    # Adjust column widths based on GST applicability and state
    if is_gst_applicable:
        if gst_details["is_same_state"]:
            # 7 columns: S.No, Desc, Price, Taxable, CGST, SGST, Amount
            col_ratios = [0.07, 0.33, 0.12, 0.13, 0.12, 0.12, 0.11]
        else:
            # 6 columns: S.No, Desc, Price, Taxable, IGST, Amount
            col_ratios = [0.07, 0.38, 0.13, 0.14, 0.17, 0.11]
    else:
        # 5 columns: S.No, Desc, Price, Taxable, Amount (no GST columns)
        col_ratios = [0.07, 0.43, 0.17, 0.18, 0.15]

    col_widths = [PAGE_W * r for r in col_ratios[:num_cols]]

    items_table = Table([headers, row], colWidths=col_widths, repeatRows=1)
    items_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.7, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2f5f7a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 4))

    # -------------------------------------------------------
    # Summary rows
    # -------------------------------------------------------
    outstanding_val = Decimal(str(amounts.get('outstanding_charge', '0.00')))
    advance_val = Decimal(str(amounts.get('advance_used', '0.00')))

    summary_rows = []

    # Show subtotal with GST label only when GST applies
    if is_gst_applicable:
        summary_rows.append(["Subtotal (incl. GST):", f"₹ {amounts['subtotal']}"])
    else:
        summary_rows.append(["Subtotal:", f"₹ {amounts['subtotal']}"])

    if outstanding_val > 0:
        summary_rows.append(["Outstanding Balance (no GST):", f"₹ {amounts['outstanding_charge']}"])

    if advance_val > 0:
        summary_rows.append(["Advance Adjusted:", f"- ₹ {amounts['advance_used']}"])

    summary_rows.append(["Balance Due:", f"₹ {amounts['total_amount']}"])

    padded_summary = []
    for label, value in summary_rows:
        padded_summary.append([""] * (num_cols - 2) + [label, value])

    summary_table = Table(padded_summary, colWidths=col_widths)

    summary_style = [
        ("ALIGN",    (-2, 0), (-1, -1), "RIGHT"),
        ("FONTSIZE", (0,  0), (-1, -1), 8),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]

    if outstanding_val > 0:
        outstanding_row_idx = 1
        summary_style += [
            ("TEXTCOLOR", (-2, outstanding_row_idx), (-1, outstanding_row_idx), colors.HexColor("#c0392b")),
            ("FONTNAME",  (-2, outstanding_row_idx), (-1, outstanding_row_idx), "Helvetica-Bold"),
        ]

    if advance_val > 0:
        advance_row_idx = 2 if outstanding_val > 0 else 1
        summary_style += [
            ("TEXTCOLOR", (-2, advance_row_idx), (-1, advance_row_idx), colors.HexColor("#27ae60")),
        ]

    last_row = len(summary_rows) - 1
    summary_style += [
        ("FONTNAME",  (-2, last_row), (-1, last_row), "Helvetica-Bold"),
        ("LINEABOVE", (-2, last_row), (-1, last_row), 1.5, colors.HexColor("#2f5f7a")),
        ("FONTSIZE",  (-2, last_row), (-1, last_row), 9),
    ]

    summary_table.setStyle(TableStyle(summary_style))
    elements.append(summary_table)
    elements.append(Spacer(1, 12))

    # -------------------------------------------------------
    # Bank Details + Amount in Words
    # -------------------------------------------------------
    bank_info_lines = []
    if company.bank_name:
        bank_info_lines.append(f"<b>Bank:</b> {company.bank_name}")
    if company.account_number:
        bank_info_lines.append(f"<b>Account Holder:</b> {company.name}")
        bank_info_lines.append(f"<b>A/C:</b> {company.account_number}")
    if company.ifsc_code:
        bank_info_lines.append(f"<b>IFSC:</b> {company.ifsc_code}")
    if company.branch_name:
        bank_info_lines.append(f"<b>Branch:</b> {company.branch_name}")

    bank_details_text = "<br/>".join(bank_info_lines) if bank_info_lines else "Bank details not available"
    bank_block = Paragraph(bank_details_text, small_left)

    amount_words = Paragraph(
        f"<b>Amount in Words</b><br/>{number_to_words(float(final_amount))}",
        small_left
    )

    elements.append(Table(
        [[bank_block, amount_words]],
        colWidths=[PAGE_W * 0.55, PAGE_W * 0.45]
    ))
    elements.append(Spacer(1, 10))

    # -------------------------------------------------------
    # QR Code + Payment Terms
    # -------------------------------------------------------
    qr_img = ""
    if company.qr_code:
        try:
            if hasattr(company.qr_code, 'path') and os.path.exists(company.qr_code.path):
                qr_img = Image(company.qr_code.path, 1.6 * inch, 1.6 * inch, kind='proportional')
        except Exception as e:
            print(f"QR Code error: {e}")

    payment_terms = Paragraph(
        """
        <b>Payment Terms</b><br/>
        • Payment due within 15 days<br/>
        • Late payment charges applicable<br/>
        • Mention Invoice Number during payment
        """,
        ParagraphStyle("payment_right", parent=small_left, alignment=TA_RIGHT)
    )

    if qr_img:
        elements.append(KeepTogether([
            Table([[qr_img, payment_terms]], colWidths=[PAGE_W * 0.35, PAGE_W * 0.65])
        ]))
    else:
        elements.append(payment_terms)

    elements.append(Spacer(1, 16))

    # -------------------------------------------------------
    # Signature
    # -------------------------------------------------------
    elements.append(Paragraph(
        f"<b>For {company.name}</b><br/><br/>Authorized Signatory",
        ParagraphStyle("sign", alignment=TA_RIGHT, fontSize=10)
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def number_to_words(number):
    """Convert number to words (Indian numbering system)"""
    ones  = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine']
    tens  = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']
    teens = ['Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen',
             'Sixteen', 'Seventeen', 'Eighteen', 'Nineteen']

    def convert_below_hundred(n):
        if n < 10:   return ones[n]
        elif n < 20: return teens[n - 10]
        else:        return tens[n // 10] + (' ' + ones[n % 10] if n % 10 != 0 else '')

    def convert_below_thousand(n):
        if n < 100:
            return convert_below_hundred(n)
        return ones[n // 100] + ' Hundred' + (' ' + convert_below_hundred(n % 100) if n % 100 != 0 else '')

    if number == 0:
        return 'Zero Only'

    integer_part = int(number)
    decimal_part = int(round((number - integer_part) * 100))
    result = []

    if integer_part >= 10000000:
        result.append(convert_below_thousand(integer_part // 10000000) + ' Crore')
        integer_part %= 10000000
    if integer_part >= 100000:
        result.append(convert_below_thousand(integer_part // 100000) + ' Lakh')
        integer_part %= 100000
    if integer_part >= 1000:
        result.append(convert_below_thousand(integer_part // 1000) + ' Thousand')
        integer_part %= 1000
    if integer_part > 0:
        result.append(convert_below_thousand(integer_part))

    words = ' '.join(result)
    if decimal_part > 0:
        words += f' and {convert_below_hundred(decimal_part)} Paise'
    return words + ' Only'




def active_customers_list(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    customers = Customer.objects.all()

    active_customers_qs = Customer.objects.filter(status__iexact="Active")
    active_customers_count = active_customers_qs.count()

    context = {
        "customers": customers,
        "active_customers": active_customers_count,
        "active_customers_list": active_customers_qs,
        'user': user,
        'role_permissions': role_permissions,
    }
    return render(request, "active_customers_list.html", context)    

def inactive_customers_list(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    inactive_customers_qs = Customer.objects.filter(status__iexact="Inactive")
    inactive_customers_count = inactive_customers_qs.count()

    context = {
        'user': user,
        'role_permissions': role_permissions,
        "inactive_customers": inactive_customers_count,
        "inactive_customers_list": inactive_customers_qs,
    }
    return render(request, "inactive_customers_list.html", context)
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Company, Custom_User, SubscriptionPlan, State, City, Area
from .utils import get_filtered_companies, get_company_filter_context
def company_management(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    
    if request.method == 'POST':
        try:
            # Get location IDs
            state_id = request.POST.get("state")
            city_id = request.POST.get("city")
            area_id = request.POST.get("area")
            
            # Get file uploads
            logo = request.FILES.get("logo")
            qr_code = request.FILES.get("qr_code")
            
            # Get date fields
            start_date = request.POST.get('start_date', None)
            if start_date == '':
                start_date = None
            
            # Get decimal fields
            processing_fee = request.POST.get('processing_fee', None)
            if processing_fee == '':
                processing_fee = None
            
            company_name = request.POST.get('name')
            
            # Create company with all fields
            company = Company.objects.create(
                name=company_name,
                email=request.POST.get('email'),
                contact=request.POST.get('contact'),
                contact_2=request.POST.get('contact_2', ''),
                address=request.POST.get('address'),
                state_id=state_id if state_id else None,
                city_id=city_id if city_id else None,
                area_id=area_id if area_id else None,
                pincode=request.POST.get('pincode', ''),
                gst_registered=request.POST.get('gst_registered') == 'yes',
                gst_number=request.POST.get('gst_number', ''),
                Contact_person=request.POST.get('contact_person', ''),
                contact_person_email=request.POST.get('contact_person_email', ''),
                contact_person_phone=request.POST.get('contact_person_phone', ''),
                status=request.POST.get('status', ''),
                start_date=start_date,
                location_link=request.POST.get('location_link', ''),
                processing_fee=processing_fee,
                # Bank Details
                bank_name=request.POST.get('bank_name', ''),
                account_number=request.POST.get('account_number', ''),
                ifsc_code=request.POST.get('ifsc_code', ''),
                branch_name=request.POST.get('branch_name', ''),
            )
            
            # Handle file uploads separately
            if logo:
                company.logo = logo
            
            if qr_code:
                company.qr_code = qr_code
            
            # Save if files were uploaded
            if logo or qr_code:
                company.save()
            
            # Log company creation
            user_logs.objects.create(
                user=user.username,
                action='Created Company',
                details=f'Company Name: {company_name}'
            )
            
            messages.success(request, f'{company.name} added successfully!')
            return redirect('company_management')
            
        except Exception as e:
            messages.error(request, f'Error adding company: {str(e)}')
            return redirect('company_management')
    
    companies = get_filtered_companies(request)
    
    subscription_plans = SubscriptionPlan.objects.all()
    state_list = State.objects.all()
    city_list = City.objects.all()
    area_list = Area.objects.all()
    
    context = {
        'companies': companies,
        'subscription_plans': subscription_plans,
        'state_list': state_list,
        'city_list': city_list,
        'area_list': area_list,
        'user': user,
        'role_permissions': role_permissions,
        **get_company_filter_context(request)
    }
    
    return render(request, 'companymaster.html', context)

def company_gst(request):
    username=request.session.get('username')
    user=Custom_User.objects.get(username=username)
    if not user:
        return redirect('login')
    companies = Company.objects.filter(gst_registered=False)
    subscription_plans = SubscriptionPlan.objects.all()
    state_list = State.objects.all()
    city_list = City.objects.all()
    area_list = Area.objects.all()
    return render(request, 'companymaster_gst.html', {'companies': companies, 'subscription_plans': subscription_plans, 'state_list': state_list, 'city_list': city_list, 'area_list': area_list})

def update_company(request, company_id):
    username = request.session.get('username')
    user = Custom_User.objects.get(username=username)
  
    if not user:
        return redirect('login')
    
    if request.method == 'POST':
        try:
            company = get_object_or_404(Company, id=company_id)
            
            # Store old values for comparison
            old_name = company.name
            old_status = company.status
            old_gst_registered = company.gst_registered
            
            state_id = request.POST.get("state")
            city_id = request.POST.get("city")
            area_id = request.POST.get("area")
            state = State.objects.get(id=state_id) if state_id else None
            city = City.objects.get(id=city_id) if city_id else None
            area = Area.objects.get(id=area_id) if area_id else None
            
            # Handle logo upload
            logo = request.FILES.get("editlogo")
            if logo:
                company.logo = logo
            
            # Handle QR code upload
            qr_code = request.FILES.get("editqr_code")
            if qr_code:
                company.qr_code = qr_code
            
            # Get new values
            new_name = request.POST.get('name', company.name)
            new_status = request.POST.get('status', company.status)
            new_gst_registered = request.POST.get('gst_registered') == 'yes'
            
            # Update basic fields
            company.name = new_name
            company.email = request.POST.get('email', company.email)
            company.contact = request.POST.get('contact', company.contact)
            company.contact_2 = request.POST.get('contact_2', company.contact_2)
            company.address = request.POST.get('address', company.address)
            company.area = area if area else None  
            company.city = city if city else None
            company.state = state if state else None
            company.pincode = request.POST.get('pincode', company.pincode)
            company.gst_registered = new_gst_registered
            company.gst_number = request.POST.get('gst_number', company.gst_number)
            company.Contact_person = request.POST.get('contact_person', company.Contact_person)
            company.contact_person_email = request.POST.get('contact_person_email', company.contact_person_email)
            company.contact_person_phone = request.POST.get('contact_person_phone', company.contact_person_phone)
            company.start_date = request.POST.get('start_date', company.start_date)
            company.status = new_status
            company.location_link = request.POST.get('location_link', company.location_link)
            company.processing_fee = request.POST.get('processing_fee', company.processing_fee)
            
            # Update bank details
            company.bank_name = request.POST.get('bank_name', company.bank_name)
            company.account_number = request.POST.get('account_number', company.account_number)
            company.ifsc_code = request.POST.get('ifsc_code', company.ifsc_code)
            company.branch_name = request.POST.get('branch_name', company.branch_name)
            
            company.save()
            
            # Build change summary for user_logs
            changes = []
            if old_name != new_name:
                changes.append(f"Name: {old_name} → {new_name}")
            if old_status != new_status:
                changes.append(f"Status: {old_status} → {new_status}")
            if old_gst_registered != new_gst_registered:
                gst_old = "Yes" if old_gst_registered else "No"
                gst_new = "Yes" if new_gst_registered else "No"
                changes.append(f"GST Registered: {gst_old} → {gst_new}")
            
            # Create user log entry
            if changes:
                change_details = ", ".join(changes)
                user_logs.objects.create(
                    user=user.username,
                    action='Updated Company',
                    details=f'Company: {old_name} | Changes: {change_details}'
                )
            else:
                user_logs.objects.create(
                    user=user.username,
                    action='Updated Company',
                    details=f'Company: {old_name}'
                )

            messages.success(request, f'{company.name} updated successfully!')
            return redirect('company_management')
        
        except Exception as e:
            messages.error(request, f'Error updating company: {str(e)}')
            return redirect('company_management')
    
    return redirect('company_management')


def delete_company(request, company_id):
    username=request.session.get('username')
    user=Custom_User.objects.get(username=username)
    if not user:
        return redirect('login')
    if request.method == 'POST':
        try:
            company = get_object_or_404(Company, id=company_id)
            company_name = company.name
            
            # Log before deletion
            user_logs.objects.create(
                user=user.username,
                action='Deleted Company',
                details=f'Company Name: {company_name}'
            )
            
            company.delete()
            
            messages.success(request, f'{company_name} deleted successfully!')
            return redirect('company_management')
            
        except Exception as e:
            messages.error(request, f'Error deleting company: {str(e)}')
            return redirect('company_management')
    
    return redirect('company_management')

def update_customer(request, customer_id):
    """
    Update customer with automatic subscription change tracking
    """
    username = request.session.get('username')
    user = Custom_User.objects.get(username=username)
    if not user:
        return redirect('login')

    if request.method == 'POST':
        try:
            customer = get_object_or_404(Customer, id=customer_id)
            
            # Store old values for comparison and logging
            old_name = customer.name
            old_subscription_plan = customer.subscription_plan
            old_custom_amount = customer.custom_subscription_amount
            old_status = customer.status
            
            # Get new subscription details from form
            subscription_plan_id = request.POST.get("subscription_plan")
            new_subscription_plan = None
            if subscription_plan_id:
                new_subscription_plan = SubscriptionPlan.objects.get(id=subscription_plan_id)
            
            # Get custom subscription amount
            new_custom_amount = request.POST.get('custom_subscription_amount', '').strip()
            if new_custom_amount:
                try:
                    new_custom_amount = Decimal(new_custom_amount)
                except (ValueError, InvalidOperation):
                    new_custom_amount = None
            else:
                new_custom_amount = None
            
            # Calculate old and new amounts for comparison
            old_amount = Decimal('0.00')
            if old_custom_amount:
                old_amount = old_custom_amount
            elif old_subscription_plan:
                old_amount = old_subscription_plan.price
            
            new_amount = Decimal('0.00')
            if new_custom_amount:
                new_amount = new_custom_amount
            elif new_subscription_plan:
                new_amount = new_subscription_plan.price
            
            # Detect if this is a subscription change
            plan_changed = False
            amount_changed = False
            is_upgrade = False
            upgrade_amount = Decimal('0.00')
            
            if old_subscription_plan and new_subscription_plan:
                # Check if plan itself changed
                if old_subscription_plan.id != new_subscription_plan.id:
                    plan_changed = True
                
                # Check if effective amount changed
                if new_amount != old_amount:
                    amount_changed = True
                
                # Determine if it's an upgrade (amount increased)
                if new_amount > old_amount:
                    is_upgrade = True
                    upgrade_amount = new_amount - old_amount
                    customer.previous_subscription_plan = old_subscription_plan
                    customer.upgrade_date = timezone.now().date()
            
            # Handle contact numbers: clear and reassign
            contacts = request.POST.getlist("contacts[]")
            customer.contacts.all().delete()
            for number in contacts:
                if number.strip():
                    CustomerContact.objects.create(customer=customer, phone_number=number)
            
            # Handle tax percentage
            tax_percentage = request.POST.get('tax_percentage')
            if tax_percentage and tax_percentage.strip():
                customer.tax_percentage = tax_percentage
            
            # Get new values
            new_name = request.POST.get('customer_name', customer.name)
            new_status = request.POST.get('status', customer.status)
            
            # Update basic fields
            customer.name = new_name
            customer.email = request.POST.get('email', customer.email)
            customer.state_id = request.POST.get('state', customer.state_id)
            customer.city_id = request.POST.get('city', customer.city_id)
            customer.area_id = request.POST.get('area', customer.area_id)
            customer.pincode = request.POST.get('pincode', customer.pincode)
            customer.address = request.POST.get('address', customer.address)
            customer.gst_registered = request.POST.get('gst_type') == 'GST'
            customer.location_link = request.POST.get('location_link', customer.location_link)
            customer.remarks = request.POST.get('remarks', customer.remarks)
            customer.start_date = request.POST.get('start_date', customer.start_date)
            customer.status = new_status
            
            # Update subscription fields
            customer.subscription_plan = new_subscription_plan
            customer.custom_subscription_amount = new_custom_amount
            
            # Handle advance_amount
            advance_amount = request.POST.get('advance_amount', '').strip()
            if advance_amount:
                try:
                    customer.advance_amount = Decimal(advance_amount)
                except (ValueError, InvalidOperation):
                    customer.advance_amount = Decimal('0.00')
            else:
                # Don't reset to 0 if field is empty, keep existing value
                pass
            
            # Handle GST number
            if customer.gst_registered:
                customer.gst_number = request.POST.get('gst_number', customer.gst_number)
            else:
                customer.gst_number = None
            
            # Save the customer (this will trigger the save() method which creates SubscriptionChange)
            customer.save()
            
            # Build change summary for user_logs
            changes = []
            if old_name != new_name:
                changes.append(f"Name: {old_name} → {new_name}")
            if plan_changed:
                old_plan_name = old_subscription_plan.name if old_subscription_plan else "None"
                new_plan_name = new_subscription_plan.name if new_subscription_plan else "None"
                changes.append(f"Plan: {old_plan_name} → {new_plan_name}")
            if amount_changed and not plan_changed:
                changes.append(f"Amount: ₹{old_amount} → ₹{new_amount}")
            if old_status != new_status:
                changes.append(f"Status: {old_status} → {new_status}")
            
            # Create user log entry
            if changes:
                change_details = ", ".join(changes)
                user_logs.objects.create(
                    user=user.username,
                    action='Updated Customer',
                    details=f'Customer: {old_name} | Changes: {change_details}'
                )
            else:
                user_logs.objects.create(
                    user=user.username,
                    action='Updated Customer',
                    details=f'Customer: {old_name}'
                )
            
            # Log the change in data_logs (existing logic)
            if plan_changed or amount_changed:
                change_type = "Upgrade" if is_upgrade else "Change"
                old_plan_name = old_subscription_plan.name if old_subscription_plan else "None"
                new_plan_name = new_subscription_plan.name if new_subscription_plan else "None"
                
                details = f"""
=== SUBSCRIPTION {change_type.upper()} ===

Customer: {customer.name}
Date: {timezone.now().strftime('%d-%m-%Y %H:%M:%S')}

Old Subscription:
  Plan: {old_plan_name}
  Amount: ₹{old_amount:.2f}/month

New Subscription:
  Plan: {new_plan_name}
  Amount: ₹{new_amount:.2f}/month

Change Details:
  {"Amount Increase" if is_upgrade else "Amount Change"}: ₹{abs(new_amount - old_amount):.2f}/month
  Effective From: {timezone.now().date().strftime('%d-%m-%Y')}

Notes:
- This change will be reflected in the next billing cycle
- Pro-rated billing will be applied for the current month
- Previous unbilled days will use old plan rate
- Remaining days in month will use new plan rate
                """.strip()
                
                data_logs.objects.create(
                    user=user,
                    timestamp=timezone.now(),
                    customer=customer,
                    action=f"Service {change_type}: {old_plan_name} → {new_plan_name}",
                    details=details,
                    is_upgrade=is_upgrade,
                    old_subscription_plan=old_subscription_plan,
                    new_subscription_plan=new_subscription_plan,
                    upgrade_amount=upgrade_amount if is_upgrade else (new_amount - old_amount),
                    status=change_type,
                    billing_period_start=timezone.now().date(),
                    billing_period_end=None
                )
                
                if is_upgrade:
                    messages.success(
                        request, 
                        f'Customer "{old_name}" upgraded successfully from {old_plan_name} '
                        f'(₹{old_amount}/month) to {new_plan_name} (₹{new_amount}/month)! '
                        f'Next bill will be pro-rated for the change.'
                    )
                else:
                    messages.success(
                        request, 
                        f'Customer "{old_name}" subscription changed from {old_plan_name} '
                        f'to {new_plan_name}. Next bill will reflect the pro-rated change.'
                    )
            else:
                # Regular update without subscription change
                data_logs.objects.create(
                    user=user,
                    timestamp=timezone.now(),
                    customer=customer,
                    action=f"Customer Updated: {old_name}",
                    details=f"Customer details updated by {user.username}",
                    status='Updated'
                )
                messages.success(request, f'Customer "{old_name}" updated successfully!')
            
            # Handle camera serials (for Live Monitoring plan)
            camera_serials = request.POST.getlist("camera_serials[]")
            customer_cameras.objects.filter(customer=customer).delete()

            if customer.subscription_plan and customer.subscription_plan.name.lower() == "live monitoring":
                for serial in camera_serials:
                    if serial.strip():
                        customer_cameras.objects.create(
                            customer=customer,
                            seriak_number=serial
                        )

            return redirect('customer_management')

        except Exception as e:
            messages.error(request, f'Error updating customer: {str(e)}')
            import traceback
            traceback.print_exc()
            return redirect('customer_management')
    
    # GET request - render form
    else:
        customer = get_object_or_404(Customer, id=customer_id)
        subscription_plans = SubscriptionPlan.objects.all()
        states = State.objects.all()
        
        context = {
            'customer': customer,
            'subscription_plans': subscription_plans,
            'states': states,
            'contacts': customer.contacts.all(),
        }
        return render(request, 'update_customer.html', context)


def delete_customer(request, customer_id):
    username=request.session.get('username')
    user=Custom_User.objects.get(username=username)
    if not user:
        return redirect('login')
    if request.method == 'POST':
        try:
            customer = get_object_or_404(Customer, id=customer_id)
            customer_name = customer.name
            
            # Log before deletion
            user_logs.objects.create(
                user=user.username,
                action='Deleted Customer',
                details=f'Customer Name: {customer_name}'
            )
            
            customer.delete()
            
            messages.success(request, f'Customer "{customer_name}" deleted successfully!')
            return redirect('customer_management')
            
        except Exception as e:
            messages.error(request, f'Error deleting customer: {str(e)}')
            return redirect('customer_management')
    
    return redirect('customer_management')
# views.py - Updated dashboard view

from django.shortcuts import render, redirect
from django.db.models import Sum, Count, Q, F
from django.db.models.functions import TruncMonth
from datetime import datetime, timedelta
import math
import json
from .utils import get_filtered_companies, get_company_filter_context
from django.shortcuts import render, redirect
from django.db.models import Q, Sum, Count
from django.db.models.functions import TruncMonth
from datetime import datetime, timedelta
import json
import math
from .models import Customer, Company, BillingRecord, Custom_User, CustomerLocation
from .utils import get_filtered_companies, get_company_filter_context
from django.db.models import Q, Sum, Count
from django.db.models.functions import TruncMonth
from datetime import datetime, timedelta
import json
import math
def dashboard(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    
    # Get filtered companies based on selection
    filtered_companies = get_filtered_companies(request)
    company_ids = list(filtered_companies.values_list('id', flat=True))
    
    # Check if filters are actively being used
    selected_gst_type = request.session.get('selected_gst_type', 'gst')
    selected_company_ids = request.session.get('selected_company_ids', [])
    filters_active = selected_gst_type or selected_company_ids
    
    if company_ids:
        # Filter billing records where:
        # 1. Customer belongs to filtered companies, OR
        # 2. Customer location's customer belongs to filtered companies
        billing_filter = Q(customer__company_id__in=company_ids) | Q(customer_location__customer__company_id__in=company_ids)
        customer_filter = Q(company_id__in=company_ids)
        
        # Filter for customer locations whose parent customer belongs to filtered companies
        customer_location_filter = Q(customer__company_id__in=company_ids)
    elif filters_active:
        # ✅ FIXED: Filters are active but returned no companies
        # Show NO data instead of ALL data
        billing_filter = Q(pk__in=[])  # Empty queryset
        customer_filter = Q(pk__in=[])
        customer_location_filter = Q(pk__in=[])
    else:
        # No filter selected - show all billing records
        billing_filter = Q()
        customer_filter = Q()
        customer_location_filter = Q()
    
    # Debug: Check how many billing records match
    total_billing_records = BillingRecord.objects.filter(billing_filter).count()
    total_bills = BillingRecord.objects.filter(billing_filter).count()
    
    # Count unique customers with billing records in filtered companies
    # This includes both direct customers and customer locations (counted separately)
    if company_ids:
        # Count customers directly linked to these companies
        direct_customers = Customer.objects.filter(
            company_id__in=company_ids
        ).distinct().count()
        
        # Count customer locations as separate customers
        customer_locations_count = CustomerLocation.objects.filter(
            customer_location_filter
        ).distinct().count()
        
        # Total customers = direct customers + locations
        total_customers = direct_customers + customer_locations_count
        
        # Active customers (only from Customer model, locations don't have status)
        active_customers = Customer.objects.filter(
            company_id__in=company_ids,
            status='Active'
        ).count()
        
        # Add active locations
        active_locations = CustomerLocation.objects.filter(
            customer_location_filter,
            is_active=True
        ).count()
        active_customers += active_locations
        
        # Inactive customers
        inactive_customers = Customer.objects.filter(
            company_id__in=company_ids,
            status='Inactive'
        ).count()
        
        # Add inactive locations
        inactive_locations = CustomerLocation.objects.filter(
            customer_location_filter,
            is_active=False
        ).count()
        inactive_customers += inactive_locations
        
    elif filters_active:
        # ✅ FIXED: Filters active but no matching companies - show zero counts
        direct_customers = 0
        customer_locations_count = 0
        total_customers = 0
        active_customers = 0
        inactive_customers = 0
    else:
        # No filter - count all customers and locations
        direct_customers = Customer.objects.count()
        customer_locations_count = CustomerLocation.objects.count()
        total_customers = direct_customers + customer_locations_count
        
        active_customers = Customer.objects.filter(status='Active').count()
        active_locations = CustomerLocation.objects.filter(is_active=True).count()
        active_customers += active_locations
        
        inactive_customers = Customer.objects.filter(status='Inactive').count()
        inactive_locations = CustomerLocation.objects.filter(is_active=False).count()
        inactive_customers += inactive_locations
    
    total_companies = filtered_companies.count()
    
    # Financial Statistics - considers both customer and customer_location
    total_amount = BillingRecord.objects.filter(billing_filter).aggregate(
        total=Sum('amount')
    )['total'] or 0.00
    total_amount_rounded = math.ceil(total_amount)
    
    total_collected = math.ceil(
        BillingRecord.objects.filter(billing_filter, paid=True).aggregate(
            total=Sum('amount')
        )['total'] or 0.00
    )
    
    total_outstanding = math.ceil(
        BillingRecord.objects.filter(billing_filter, paid=False).aggregate(
            total=Sum('amount')
        )['total'] or 0.00
    )
    
    # Calculate total advance received from filtered companies' customers
    # This only applies to Customer model, not CustomerLocation
    total_advance = Customer.objects.filter(
        customer_filter,
        advance_amount__isnull=False
    ).aggregate(
        total=Sum('advance_amount')
    )['total'] or 0.00
    total_advance_rounded = math.ceil(total_advance)
    if company_ids:
        print(f"  From companies: {company_ids}")
    elif filters_active:
        print(f"  Filters active but no matching companies")
    else:
        print(f"  From all companies")
    
    # Invoice Status Counts - considers both customer and customer_location
    today = datetime.now().date()
    paid_count = BillingRecord.objects.filter(billing_filter, paid=True).count()
    unpaid_count = BillingRecord.objects.filter(
        billing_filter, paid=False, due_date__gte=today
    ).count()
    overdue_count = BillingRecord.objects.filter(
        billing_filter, paid=False, due_date__lt=today
    ).count()
    pending_count = BillingRecord.objects.filter(
        billing_filter, paid=False, billing_date__gte=today
    ).count()
    
    # Monthly Revenue Trends (Last 6 months) - considers both customer and customer_location
    six_months_ago = today - timedelta(days=180)
    monthly_revenue = BillingRecord.objects.filter(
        billing_filter,
        billing_date__gte=six_months_ago
    ).annotate(
        month=TruncMonth('billing_date')
    ).values('month').annotate(
        total=Sum('amount')
    ).order_by('month')
    
    # Monthly Collection Trends - considers both customer and customer_location
    monthly_collection = BillingRecord.objects.filter(
        billing_filter,
        paid=True,
        payment_date__isnull=False,
        payment_date__gte=six_months_ago
    ).annotate(
        month=TruncMonth('payment_date')
    ).values('month').annotate(
        total=Sum('amount')
    ).order_by('month')
    
    # Format monthly data for charts
    revenue_labels = []
    revenue_data = []
    collection_data = []
    
    for entry in monthly_revenue:
        revenue_labels.append(entry['month'].strftime('%b'))
        revenue_data.append(float(entry['total']))
    
    # Create a dictionary for quick lookup
    collection_dict = {
        entry['month'].strftime('%b'): float(entry['total']) 
        for entry in monthly_collection
    }
    
    # Fill collection data matching revenue labels
    for label in revenue_labels:
        collection_data.append(collection_dict.get(label, 0))
    
    # Top Customers by Revenue - considers both direct billing and location-based billing
    if company_ids:
        # Get customers with direct billing records
        customers_with_direct_billing = Customer.objects.filter(
            name__isnull=False,
            company_id__in=company_ids,
            billing_records__isnull=False
        ).exclude(
            name=''
        ).annotate(
            direct_revenue=Sum('billing_records__amount'),
            direct_count=Count('billing_records')
        ).values('id', 'name', 'direct_revenue', 'direct_count')
        
        # Get customers with location-based billing records
        customers_with_location_billing = Customer.objects.filter(
            name__isnull=False,
            company_id__in=company_ids,
            locations__billing_records__isnull=False
        ).exclude(
            name=''
        ).annotate(
            location_revenue=Sum('locations__billing_records__amount'),
            location_count=Count('locations__billing_records')
        ).values('id', 'name', 'location_revenue', 'location_count')
        
        # Combine and aggregate
        customer_revenue_dict = {}
        
        for c in customers_with_direct_billing:
            customer_revenue_dict[c['id']] = {
                'name': c['name'],
                'total_revenue': c['direct_revenue'] or 0,
                'invoice_count': c['direct_count'] or 0
            }
        
        for c in customers_with_location_billing:
            if c['id'] in customer_revenue_dict:
                customer_revenue_dict[c['id']]['total_revenue'] += (c['location_revenue'] or 0)
                customer_revenue_dict[c['id']]['invoice_count'] += (c['location_count'] or 0)
            else:
                customer_revenue_dict[c['id']] = {
                    'name': c['name'],
                    'total_revenue': c['location_revenue'] or 0,
                    'invoice_count': c['location_count'] or 0
                }
        
        # Sort and get top 5
        top_customers_list = sorted(
            customer_revenue_dict.values(), 
            key=lambda x: x['total_revenue'], 
            reverse=True
        )[:5]
        
    elif filters_active:
        # ✅ FIXED: Filters active but no matching companies - empty list
        top_customers_list = []
    else:
        # No filter - get all customers
        customers_with_direct_billing = Customer.objects.filter(
            name__isnull=False,
            billing_records__isnull=False
        ).exclude(
            name=''
        ).annotate(
            direct_revenue=Sum('billing_records__amount'),
            direct_count=Count('billing_records')
        ).values('id', 'name', 'direct_revenue', 'direct_count')
        
        customers_with_location_billing = Customer.objects.filter(
            name__isnull=False,
            locations__billing_records__isnull=False
        ).exclude(
            name=''
        ).annotate(
            location_revenue=Sum('locations__billing_records__amount'),
            location_count=Count('locations__billing_records')
        ).values('id', 'name', 'location_revenue', 'location_count')
        
        # Combine and aggregate
        customer_revenue_dict = {}
        
        for c in customers_with_direct_billing:
            customer_revenue_dict[c['id']] = {
                'name': c['name'],
                'total_revenue': c['direct_revenue'] or 0,
                'invoice_count': c['direct_count'] or 0
            }
        
        for c in customers_with_location_billing:
            if c['id'] in customer_revenue_dict:
                customer_revenue_dict[c['id']]['total_revenue'] += (c['location_revenue'] or 0)
                customer_revenue_dict[c['id']]['invoice_count'] += (c['location_count'] or 0)
            else:
                customer_revenue_dict[c['id']] = {
                    'name': c['name'],
                    'total_revenue': c['location_revenue'] or 0,
                    'invoice_count': c['location_count'] or 0
                }
        
        # Sort and get top 5
        top_customers_list = sorted(
            customer_revenue_dict.values(), 
            key=lambda x: x['total_revenue'], 
            reverse=True
        )[:5]
    
    # Format top customers data
    top_customers_data = []
    for customer in top_customers_list:
        name_parts = customer['name'].strip().split()
        initials = ''.join([word[0].upper() for word in name_parts[:2]]) if name_parts else 'NA'
        
        top_customers_data.append({
            'name': customer['name'],
            'initials': initials,
            'total': float(customer['total_revenue']),
            'count': customer['invoice_count']
        })
    
    # Company Revenue - considers both direct and location-based billing
    if company_ids or not filters_active:
        # Only calculate if we have companies or no filter is active
        company_revenue = filtered_companies.filter(
            name__isnull=False
        ).exclude(
            name=''
        ).annotate(
            # Direct billing through customers
            direct_revenue=Sum('customer__billing_records__amount'),
            # Location-based billing
            location_revenue=Sum('customer__locations__billing_records__amount')
        ).values('name', 'direct_revenue', 'location_revenue')
        
        # Combine revenues for each company
        company_revenue_list = []
        for comp in company_revenue:
            total_rev = (comp['direct_revenue'] or 0) + (comp['location_revenue'] or 0)
            if total_rev > 0:
                company_revenue_list.append({
                    'name': comp['name'],
                    'total_revenue': total_rev
                })
        
        # Sort and get top 4
        company_revenue_list.sort(key=lambda x: x['total_revenue'], reverse=True)
        company_revenue_list = company_revenue_list[:4]
    else:
        # ✅ FIXED: Filters active but no matching companies - empty list
        company_revenue_list = []
    
    company_labels = []
    company_data = []
    
    for comp in company_revenue_list:
        comp_name = comp['name'][:15] if comp['name'] else 'Unknown'
        company_labels.append(comp_name)
        company_data.append(float(comp['total_revenue']))
    
    # Payment Method Distribution - considers both customer and customer_location
    payment_methods = BillingRecord.objects.filter(
        billing_filter,
        paid=True,
        payment_mode__isnull=False
    ).exclude(
        payment_mode=''
    ).values('payment_mode').annotate(
        count=Count('id')
    ).order_by('-count')
    
    payment_method_labels = []
    payment_method_data = []
    
    for pm in payment_methods:
        mode = pm['payment_mode'].title() if pm['payment_mode'] else 'Other'
        payment_method_labels.append(mode)
        payment_method_data.append(pm['count'])
    
    if not payment_method_labels:
        payment_method_labels = ['No Data']
        payment_method_data = [0]
    
    # Recent Transactions - considers both customer and customer_location
    recent_transactions_query = BillingRecord.objects.filter(
        billing_filter,
        paid=True,
        payment_date__isnull=False
    ).select_related('customer', 'customer__company', 'customer_location', 'customer_location__customer').order_by('-payment_date')[:5]
    
    transactions_data = []
    for transaction in recent_transactions_query:
        # Get client name from customer or customer_location
        if transaction.customer_location:
            client_name = f"{transaction.customer_location.customer.name} - {transaction.customer_location.location_name}"
        elif transaction.customer and transaction.customer.name:
            client_name = transaction.customer.name
        else:
            client_name = 'Unknown Client'
        
        transactions_data.append({
            'client': client_name,
            'amount': float(transaction.amount),
            'date': transaction.payment_date.strftime('%d %b'),
            'mode': transaction.payment_mode.upper() if transaction.payment_mode else 'N/A'
        })
    
    # Pending Payments - considers both customer and customer_location
    pending_payments_query = BillingRecord.objects.filter(
        billing_filter,
        paid=False
    ).select_related('customer', 'customer__company', 'customer_location', 'customer_location__customer').order_by('due_date')[:5]
    
    pending_data = []
    for payment in pending_payments_query:
        # Get client name from customer or customer_location
        if payment.customer_location:
            client_name = f"{payment.customer_location.customer.name} - {payment.customer_location.location_name}"
        elif payment.customer and payment.customer.name:
            client_name = payment.customer.name
        else:
            client_name = 'Unknown Client'
        
        pending_data.append({
            'client': client_name,
            'invoice': payment.invoice_number if payment.invoice_number else 'N/A',
            'due_date': payment.due_date.strftime('%d %b') if payment.due_date else 'N/A',
            'amount': float(payment.amount)
        })
    
    # Monthly Invoice Breakdown - considers both customer and customer_location
    monthly_breakdown = []
    for i in range(6):
        month_date = today - timedelta(days=30*i)
        month_start = month_date.replace(day=1)
        
        if i == 0:
            month_end = today
        else:
            next_month = month_start.replace(day=28) + timedelta(days=4)
            month_end = next_month - timedelta(days=next_month.day)
        
        paid = BillingRecord.objects.filter(
            billing_filter,
            billing_date__gte=month_start,
            billing_date__lte=month_end,
            paid=True
        ).count()
        
        unpaid = BillingRecord.objects.filter(
            billing_filter,
            billing_date__gte=month_start,
            billing_date__lte=month_end,
            paid=False,
            due_date__gte=today
        ).count()
        
        pending = BillingRecord.objects.filter(
            billing_filter,
            billing_date__gte=month_start,
            billing_date__lte=month_end,
            paid=False,
        ).count()
        
        monthly_breakdown.insert(0, {
            'month': month_start.strftime('%b'),
            'paid': paid,
            'unpaid': unpaid,
            'pending': pending
        })
    
    context = {
        # Basic Stats
        'customers': total_customers,
        'companies': total_companies,
        'bills': total_bills,
        'bill_total': total_amount_rounded,
        'active_customers': active_customers,
        'inacive_customers': inactive_customers,
        'total_collected': total_collected,
        'total_outstanding': total_outstanding,
        'total_advance': total_advance_rounded,
        
        # Invoice Status
        'paid_count': paid_count,
        'unpaid_count': unpaid_count,
        'pending_count': pending_count,
        'overdue_count': overdue_count,
        
        # Chart Data (as JSON)
        'revenue_labels': json.dumps(revenue_labels if revenue_labels else ['No Data']),
        'revenue_data': json.dumps(revenue_data if revenue_data else [0]),
        'collection_data': json.dumps(collection_data if collection_data else [0]),
        'company_labels': json.dumps(company_labels if company_labels else ['No Data']),
        'company_data': json.dumps(company_data if company_data else [0]),
        'payment_method_labels': json.dumps(payment_method_labels),
        'payment_method_data': json.dumps(payment_method_data),
        
        # Lists
        'top_customers': top_customers_data,
        'recent_transactions': transactions_data,
        'pending_payments': pending_data,
        'monthly_breakdown': json.dumps(monthly_breakdown),
        
        'user': user,
        'role_permissions': role_permissions,
        **get_company_filter_context(request)
    }
    
    return render(request, 'dashboard.html', context)

from django.http import JsonResponse
from django.views.decorators.http import require_http_methods

@require_http_methods(["GET"])
def session_diagnostic(request):
    session_data = {
        'session_key': request.session.session_key,
        'gst_type': request.session.get('gst_type', 'NOT SET'),
        'selected_companies': request.session.get('selected_companies', []),
        'username': request.session.get('username', 'NOT SET'),
        'all_session_keys': list(request.session.keys()),
        'session_is_empty': request.session.is_empty(),
    }
    for key, value in session_data.items():
        print(f"{key}: {value}")
    print(f"{'='*70}\n")
    
    return JsonResponse(session_data)




def subscription_plan_management(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    if request.method == 'POST':
        plan_name = request.POST.get('plan_name')
        price = request.POST.get('price')
        user = Custom_User.objects.get(pk=request.session['user_id'])
        
        SubscriptionPlan.objects.create(
            name=plan_name,
            price=price,
            created_by=user
        )
        user_logs.objects.create(
            user=user.username,
            action='Created Subscription Plan',
            details=f'Plan Name: {plan_name}, Price: {price}'
        )
        messages.success(request, f'Subscription plan "{plan_name}" created successfully!')
        return redirect('subscription_plan_management')
    
    subscription_plans = SubscriptionPlan.objects.all().order_by('-created_at')
    
    context = {
        'user': user,
        'role_permissions': role_permissions,
        'subscription_plans': subscription_plans,
        **get_company_filter_context(request) 
    }
    return render(request, 'subscriptionplans.html', context)


def update_subscription_plan(request, plan_id):
    username=request.session.get('username')
    user=Custom_User.objects.get(username=username)
    if not user:
        return redirect('login')
    plan = get_object_or_404(SubscriptionPlan, id=plan_id)
    user = Custom_User.objects.get(pk=request.session['user_id'])
    
    if request.method == 'POST':
        old_name = plan.name
        old_price = plan.price
        
        new_name = request.POST.get('plan_name')
        new_price = request.POST.get('price')
        
        plan.name = new_name
        plan.price = new_price
        plan.updated_by = user
        plan.save()
        changes = []
        if old_name != new_name:
            changes.append(f"Name: {old_name} -> {new_name}")
        if str(old_price) != str(new_price):
            changes.append(f"Price: {old_price} -> {new_price}")
        
        if changes:
            change_details = ", ".join(changes)
            user_logs.objects.create(
                user=user.username,
                action='Updated Subscription Plan',
                details=f'Plan: {old_name} | Changes: {change_details}'
            )
        else:
            user_logs.objects.create(
                user=user.username,
                action='Updated Subscription Plan',
                details=f'Plan: {old_name}'
            )
        
        messages.success(request, f'Subscription plan "{old_name}" updated successfully!')
        
    return redirect('subscription_plan_management')


def delete_subscription_plan(request, plan_id):
    username=request.session.get('username')
    user=Custom_User.objects.get(username=username)
    if not user:
        return redirect('login')
    plan = get_object_or_404(SubscriptionPlan, id=plan_id)
    
    if request.method == 'POST':
        plan_name = plan.name
        plan_price = plan.price
        
        # Log before deletion
        user_logs.objects.create(
            user=user.username,
            action='Deleted Subscription Plan',
            details=f'Plan Name: {plan_name}, Price: {plan_price}'
        )
        
        plan.delete()
        messages.success(request, f'Subscription plan "{plan_name}" deleted successfully!')
        
    return redirect('subscription_plan_management')

def get_logged_in_user(request):

    user_session = request.session.get('username')
    if not user_session:
        return None, None

    try:
        user = Custom_User.objects.get(username=user_session)
        try:
            role_permission = RolePermissions.objects.get(role=user.role)
        except RolePermissions.DoesNotExist:
            role_permission = None   
        return user, role_permission
    except Custom_User.DoesNotExist:
        return None, None


def role_list(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    roles = Roles.objects.all()
    context = {
        'roles': roles,
        **get_company_filter_context(request),
        'user': user,
        'role_permissions': role_permissions 
    }   
    return render(request, 'roles.html', context)
def add_role(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    
    if request.method == 'POST':
        role_name = request.POST.get('role_name')
        description = request.POST.get('description')
        
        if Roles.objects.filter(role_name__iexact=role_name).exists():
            messages.error(request, "This Role already exists.")
            return render(request, 'roles.html')
        
        role = Roles.objects.create(
            role_name=role_name,
            description=description
        )
        
        # Log role creation
        user_logs.objects.create(
            user=user.username,
            action='Created Role',
            details=f'Role Name: {role_name}'
        )
        
        messages.success(request, f'Role added successfully!')
        return redirect('role_permissions', role_id=role.id)
    
    return redirect('role_list')


def update_role(request, role_id):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    
    role = get_object_or_404(Roles, id=role_id)

    if request.method == 'POST':
        old_name = role.role_name
        old_description = role.description
        
        new_name = request.POST.get('role_name')
        new_description = request.POST.get('description')
        
        role.role_name = new_name
        role.description = new_description
        role.save()
        
        # Log role update
        changes = []
        if old_name != new_name:
            changes.append(f"Name: {old_name} → {new_name}")
        if old_description != new_description:
            changes.append(f"Description changed")
        
        if changes:
            change_details = ", ".join(changes)
            user_logs.objects.create(
                user=user.username,
                action='Updated Role',
                details=f'Role: {old_name} | Changes: {change_details}'
            )
        else:
            user_logs.objects.create(
                user=user.username,
                action='Updated Role',
                details=f'Role: {old_name}'
            )

        messages.success(request, f'Role "{role.role_name}" updated successfully!')
    return redirect('role_list')


def role_permissions(request, role_id):
    user, user_role_perms = get_logged_in_user(request)  # renamed to avoid collision with view name
    if not user:
        return redirect('login')

    role = get_object_or_404(Roles, id=role_id)
    permissions, created = RolePermissions.objects.get_or_create(role=role)

    if request.method == 'POST':
        changed_permissions = []

        # Only iterate over actual boolean permission fields — skip id, role, and relations
        PERMISSION_FIELDS = [
            'dashboard_v',
            'customer_management_v', 'customer_management_a', 'customer_management_e', 'customer_management_d',
            'company_master_v', 'company_master_a', 'company_master_e', 'company_master_d',
            'subscription_master_v', 'subscription_master_a', 'subscription_master_e', 'subscription_master_d',
            'billing_invoices_v', 'billing_invoices_a', 'billing_invoices_e', 'billing_invoices_d',
            'payment_approval_v', 'payment_approval_a', 'payment_approval_e', 'payment_approval_d',
            'roles_v', 'roles_a', 'roles_e', 'roles_d',
            'users_v', 'users_a', 'users_e', 'users_d',
            'manual_receipts_v', 'manual_receipts_a',
            'reports_v',
            'data_logs_v',
            'user_logs_v',
            'settings_v',
        ]

        for field_name in PERMISSION_FIELDS:
            old_value = getattr(permissions, field_name)
            new_value = field_name in request.POST  # checkbox: present = True, absent = False

            if old_value != new_value:
                status = "Enabled" if new_value else "Disabled"
                changed_permissions.append(f"{field_name}: {status}")

            setattr(permissions, field_name, new_value)

        permissions.save()

        # Audit log
        if changed_permissions:
            change_details = ", ".join(changed_permissions[:5])
            if len(changed_permissions) > 5:
                change_details += f" and {len(changed_permissions) - 5} more..."
        else:
            change_details = "No changes"

        user_logs.objects.create(
            user=user.username,
            action='Updated Role Permissions',
            details=f'Role: {role.role_name} | Changes: {change_details}'
        )

        messages.success(request, "Permissions updated successfully.")
        return redirect('role_list')

    return render(request, 'role_permissions.html', {
        'role': role,
        'permissions': permissions,
        **get_company_filter_context(request),
        'user': user,
        'role_permissions': user_role_perms,  # pass the logged-in user's perms for sidebar etc.
    })


def delete_role(request, role_id):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    role = get_object_or_404(Roles, id=role_id)

    if role.custom_user_set.exists():
        messages.error(request, 'Cannot delete role assigned to users!')
        return redirect('role_list')
    
    role_name = role.role_name
    
    # Log before deletion
    user_logs.objects.create(
        user=user.username,
        action='Deleted Role',
        details=f'Role Name: {role_name}'
    )
    
    role.delete()
    messages.success(request, 'Role deleted successfully!')
    return redirect('role_list')


def user_list(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    
    users = Custom_User.objects.all()
    roles = Roles.objects.all()

    context = {
        'users': users,
        'roles': roles,
        'user': user,
        'role_permissions': role_permissions,
        **get_company_filter_context(request)
    }

    return render(request, 'users.html', context)



def add_user(request):
    logged_in_user, role_permissions = get_logged_in_user(request)
    if not logged_in_user:
        return redirect('login')
    
    if request.method == 'POST':
        try:
            password = request.POST.get('password')
            confirm_password = request.POST.get('confirm_password')
            
            if password != confirm_password:
                messages.error(request, 'Passwords do not match!')
                return redirect('user_list')
            
            role_id = request.POST.get('role')
            role = Roles.objects.get(id=role_id) if role_id else None
            
            username = request.POST.get('username')
            
            new_user = Custom_User.objects.create(
                username=username,
                email=request.POST.get('email'),
                name=request.POST.get('name', ''),
                role=role,
                phone_number=request.POST.get('phone_number', ''),
                password=password  
            )
            
            # Log user creation
            user_logs.objects.create(
                user=logged_in_user.username,
                action='Created User',
                details=f'Username: {username}, Role: {role.role_name if role else "None"}'
            )
            
            messages.success(request, f'User "{new_user.username}" added successfully!')
            return redirect('user_list')
            
        except Exception as e:
            messages.error(request, f'Error adding user: {str(e)}')
            return redirect('user_list')
    
    return redirect('user_list')


def update_user(request, user_id):
    logged_in_user, role_permissions = get_logged_in_user(request)
    if not logged_in_user:
        return redirect('login')
    
    if request.method == 'POST':
        try:
            target_user = get_object_or_404(Custom_User, id=user_id)
            
            # Store old values
            old_username = target_user.username
            old_email = target_user.email
            old_name = target_user.name
            old_phone = target_user.phone_number
            old_role = target_user.role
            
            # Get new values
            new_username = request.POST.get('username', target_user.username)
            new_email = request.POST.get('email', target_user.email)
            new_name = request.POST.get('name', target_user.name)
            new_phone = request.POST.get('phone_number', target_user.phone_number)
            
            role_id = request.POST.get('role')
            new_role = Roles.objects.get(id=role_id) if role_id else None
            
            # Update user fields
            target_user.username = new_username
            target_user.email = new_email
            target_user.name = new_name
            target_user.phone_number = new_phone
            target_user.role = new_role
            
            # Track password change
            password_changed = False
            password = request.POST.get('password')
            if password:
                confirm_password = request.POST.get('confirm_password')
                if password != confirm_password:
                    messages.error(request, 'Passwords do not match!')
                    return redirect('user_list')
                target_user.password = password
                password_changed = True
            
            target_user.save()
            
            # Log user update
            changes = []
            if old_username != new_username:
                changes.append(f"Username: {old_username} → {new_username}")
            if old_email != new_email:
                changes.append(f"Email: {old_email} → {new_email}")
            if old_name != new_name:
                changes.append(f"Name: {old_name} → {new_name}")
            if old_phone != new_phone:
                changes.append(f"Phone: {old_phone} → {new_phone}")
            if old_role != new_role:
                old_role_name = old_role.role_name if old_role else "None"
                new_role_name = new_role.role_name if new_role else "None"
                changes.append(f"Role: {old_role_name} → {new_role_name}")
            if password_changed:
                changes.append("Password changed")
            
            if changes:
                change_details = ", ".join(changes)
                user_logs.objects.create(
                    user=logged_in_user.username,
                    action='Updated User',
                    details=f'User: {old_username} | Changes: {change_details}'
                )
            else:
                user_logs.objects.create(
                    user=logged_in_user.username,
                    action='Updated User',
                    details=f'User: {old_username}'
                )
            
            messages.success(request, f'User "{target_user.username}" updated successfully!')
            return redirect('user_list')
            
        except Exception as e:
            messages.error(request, f'Error updating user: {str(e)}')
            return redirect('user_list')
    
    return redirect('user_list')


def delete_user(request, user_id):
    logged_in_user, role_permissions = get_logged_in_user(request)
    if not logged_in_user:
        return redirect('login')
    
    if request.method == 'POST':
        try:
            target_user = get_object_or_404(Custom_User, id=user_id)
            username = target_user.username
            role_name = target_user.role.role_name if target_user.role else "None"
            
            # Log before deletion
            user_logs.objects.create(
                user=logged_in_user.username,
                action='Deleted User',
                details=f'Username: {username}, Role: {role_name}'
            )
            
            target_user.delete()
            
            messages.success(request, f'User "{username}" deleted successfully!')
            return redirect('user_list')
            
        except Exception as e:
            messages.error(request, f'Error deleting user: {str(e)}')
            return redirect('user_list')
    
    return redirect('user_list')


from django.shortcuts import render, redirect
from django.db.models import Q
from .models import BillingRecord, Custom_User
from .utils import get_filtered_companies, get_company_filter_context
from django.db.models import F, ExpressionWrapper, DecimalField, Q, Case, When, Value
from django.shortcuts import render, redirect
from datetime import date

from django.db.models import F, ExpressionWrapper, DecimalField, Q, Case, When, Value, CharField, BooleanField
from django.shortcuts import render, redirect
from datetime import date

def billing_management(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    
    username = request.session.get('username')
    if not username:
        return redirect('login')
    
    try:
        user = Custom_User.objects.get(username=username)
    except Custom_User.DoesNotExist:
        return redirect('login')
    
    # Get filtered companies
    filtered_companies = get_filtered_companies(request)
    company_ids = list(filtered_companies.values_list('id', flat=True))
    
    if company_ids:
        billing_filter = (
            Q(customer__company_id__in=company_ids) |
            Q(customer_location__customer__company_id__in=company_ids)
        )
    else:
        billing_filter = Q(pk__in=[])
    
    # Get filtered billing records - rely on model properties
    billing_records = BillingRecord.objects.filter(billing_filter).select_related(
        'customer',
        'customer__company',
        'customer_location',
        'customer_location__customer',
        'customer_location__customer__company'
    ).order_by('-billing_start_date', '-billing_date')
    
    context = {
        'user': user,
        'role_permissions': role_permissions,
        'billing_records': billing_records,
        'today': date.today(),
        **get_company_filter_context(request)
    }
    
    return render(request, 'billingrecords.html', context)




def state_list(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    
    # Show ALL states (recommended)
    states = State.objects.all().order_by('name')
    
    context = {
        'user': user,
        'role_permissions': role_permissions,
        'states': states,
        **get_company_filter_context(request)  # Shows filter in navbar
    }
    return render(request, 'states.html', context)


def add_state(request):
    username = request.session.get('username')
    user = Custom_User.objects.get(username=username) if username else None
    
    if request.method == 'POST':
        state_name = request.POST.get('state_name', '').strip()
        
        if not state_name:
            messages.error(request, 'State name is required.')
            return redirect('state_list')
        if State.objects.filter(name__iexact=state_name).exists():
            messages.error(request, f'State "{state_name}" already exists.')
            return redirect('state_list')
        
        try:
            State.objects.create(name=state_name)
            
            # Log state creation
            if user:
                user_logs.objects.create(
                    user=user.username,
                    action='Created State',
                    details=f'State Name: {state_name}'
                )
            
            messages.success(request, f'State "{state_name}" added successfully.')
        except Exception as e:
            messages.error(request, f'Error adding state: {str(e)}')
        
        return redirect('state_list')
    
    return redirect('state_list')
def update_state(request, state_id):
    """Update an existing state"""
    username = request.session.get('username')
    user = Custom_User.objects.get(username=username) if username else None
    
    state = get_object_or_404(State, id=state_id)
    
    if request.method == 'POST':
        old_name = state.name
        state_name = request.POST.get('state_name', '').strip()
        
        if not state_name:
            messages.error(request, 'State name is required.')
            return redirect('state_list')
        if State.objects.filter(name__iexact=state_name).exclude(id=state_id).exists():
            messages.error(request, f'State "{state_name}" already exists.')
            return redirect('state_list')
        
        try:
            state.name = state_name
            state.save()
            
            # Log state update
            if user:
                if old_name != state_name:
                    user_logs.objects.create(
                        user=user.username,
                        action='Updated State',
                        details=f'State: {old_name} | Changes: Name: {old_name} → {state_name}'
                    )
                else:
                    user_logs.objects.create(
                        user=user.username,
                        action='Updated State',
                        details=f'State: {state_name}'
                    )
            
            messages.success(request, f'State "{state_name}" updated successfully.')
        except Exception as e:
            messages.error(request, f'Error updating state: {str(e)}')
        
        return redirect('state_list')
    
    return redirect('state_list')


def delete_state(request, state_id):
    username = request.session.get('username')
    user = Custom_User.objects.get(username=username) if username else None
    
    state = get_object_or_404(State, id=state_id)
    
    if request.method == 'POST':
        state_name = state.name
        try:
            # Log before deletion
            if user:
                user_logs.objects.create(
                    user=user.username,
                    action='Deleted State',
                    details=f'State Name: {state_name}'
                )
            
            state.delete()
            messages.success(request, f'State "{state_name}" deleted successfully.')
        except Exception as e:
            messages.error(request, f'Error deleting state: {str(e)}')
    
    return redirect('state_list')





def city_list(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    cities = City.objects.select_related('state').all().order_by('name')
    states = State.objects.all().order_by('name')
    
    context = {
        'user': user,
        'role_permissions': role_permissions,
        'cities': cities,
        'states': states,
        **get_company_filter_context(request)
    }
    return render(request, 'cities.html', context)
def add_city(request):
    username = request.session.get('username')
    user = Custom_User.objects.get(username=username) if username else None
    
    if request.method == 'POST':
        city_name = request.POST.get('city_name', '').strip()
        state_id = request.POST.get('state')
        
        if not city_name:
            messages.error(request, 'City name is required.')
            return redirect('city_list')
        
        if not state_id:
            messages.error(request, 'State is required.')
            return redirect('city_list')
        
        try:
            state = State.objects.get(id=state_id)
            
            if City.objects.filter(name__iexact=city_name, state=state).exists():
                messages.error(request, f'City "{city_name}" already exists in {state.name}.')
                return redirect('city_list')
            
            City.objects.create(name=city_name, state=state)
            
            # Log city creation
            if user:
                user_logs.objects.create(
                    user=user.username,
                    action='Created City',
                    details=f'City Name: {city_name} in State: {state.name}'
                )
            
            messages.success(request, f'City "{city_name}" added successfully.')
        except State.DoesNotExist:
            messages.error(request, 'Selected state does not exist.')
        except Exception as e:
            messages.error(request, f'Error adding city: {str(e)}')
        
        return redirect('city_list')
    
    return redirect('city_list')


def update_city(request, city_id):
    username = request.session.get('username')
    user = Custom_User.objects.get(username=username) if username else None
    
    city = get_object_or_404(City, id=city_id)
    
    if request.method == 'POST':
        old_name = city.name
        old_state = city.state
        
        city_name = request.POST.get('city_name', '').strip()
        state_id = request.POST.get('state')
        
        if not city_name:
            messages.error(request, 'City name is required.')
            return redirect('city_list')
        
        if not state_id:
            messages.error(request, 'State is required.')
            return redirect('city_list')
        
        try:
            state = State.objects.get(id=state_id)
            
            if City.objects.filter(name__iexact=city_name, state=state).exclude(id=city_id).exists():
                messages.error(request, f'City "{city_name}" already exists in {state.name}.')
                return redirect('city_list')
            
            city.name = city_name
            city.state = state
            city.save()
            
            # Log city update
            if user:
                changes = []
                if old_name != city_name:
                    changes.append(f"Name: {old_name} → {city_name}")
                if old_state.id != state.id:
                    changes.append(f"State: {old_state.name} → {state.name}")
                
                if changes:
                    change_details = ", ".join(changes)
                    user_logs.objects.create(
                        user=user.username,
                        action='Updated City',
                        details=f'City: {old_name} | Changes: {change_details}'
                    )
                else:
                    user_logs.objects.create(
                        user=user.username,
                        action='Updated City',
                        details=f'City: {city_name}'
                    )
            
            messages.success(request, f'City "{city_name}" updated successfully.')
        except State.DoesNotExist:
            messages.error(request, 'Selected state does not exist.')
        except Exception as e:
            messages.error(request, f'Error updating city: {str(e)}')
        
        return redirect('city_list')
    
    return redirect('city_list')


def delete_city(request, city_id):
    username = request.session.get('username')
    user = Custom_User.objects.get(username=username) if username else None
    
    city = get_object_or_404(City, id=city_id)
    
    if request.method == 'POST':
        city_name = city.name
        state_name = city.state.name
        try:
            # Log before deletion
            if user:
                user_logs.objects.create(
                    user=user.username,
                    action='Deleted City',
                    details=f'City Name: {city_name} in State: {state_name}'
                )
            
            city.delete()
            messages.success(request, f'City "{city_name}" deleted successfully.')
        except Exception as e:
            messages.error(request, f'Error deleting city: {str(e)}')
    
    return redirect('city_list')


def area_list(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    areas = Area.objects.select_related('city__state').all().order_by('name')
    cities = City.objects.select_related('state').all().order_by('name')
    
    context = {
        'user': user,
        'role_permissions': role_permissions,
        'areas': areas,
        'cities': cities,
         **get_company_filter_context(request)
    }
    return render(request, 'areas.html', context)

def add_area(request):
    """Add a new area"""
    username = request.session.get('username')
    user = Custom_User.objects.get(username=username) if username else None
    
    if request.method == 'POST':
        area_name = request.POST.get('area_name', '').strip()
        city_id = request.POST.get('city')
        pincode = request.POST.get('pincode', '').strip()
        
        if not area_name:
            messages.error(request, 'Area name is required.')
            return redirect('area_list')
        
        if not city_id:
            messages.error(request, 'City is required.')
            return redirect('area_list')
        
        try:
            city = City.objects.get(id=city_id)
            
            # Check if area already exists in this city
            if Area.objects.filter(name__iexact=area_name, city=city).exists():
                messages.error(request, f'Area "{area_name}" already exists in {city.name}.')
                return redirect('area_list')
            
            Area.objects.create(name=area_name, city=city, pincode=pincode)
            
            # Log area creation
            if user:
                user_logs.objects.create(
                    user=user.username,
                    action='Created Area',
                    details=f'Area Name: {area_name} in City: {city.name}'
                )
            
            messages.success(request, f'Area "{area_name}" added successfully.')
        except City.DoesNotExist:
            messages.error(request, 'Selected city does not exist.')
        except Exception as e:
            messages.error(request, f'Error adding area: {str(e)}')
        
        return redirect('area_list')
    
    return redirect('area_list')


def update_area(request, area_id):
    """Update an existing area"""
    username = request.session.get('username')
    user = Custom_User.objects.get(username=username) if username else None
    
    area = get_object_or_404(Area, id=area_id)
    
    if request.method == 'POST':
        old_name = area.name
        old_city = area.city
        old_pincode = area.pincode
        
        area_name = request.POST.get('area_name', '').strip()
        city_id = request.POST.get('city')
        pincode = request.POST.get('pincode', '').strip()
        
        if not area_name:
            messages.error(request, 'Area name is required.')
            return redirect('area_list')
        
        if not city_id:
            messages.error(request, 'City is required.')
            return redirect('area_list')
        
        try:
            city = City.objects.get(id=city_id)
            
            # Check if another area with same name exists in this city
            if Area.objects.filter(name__iexact=area_name, city=city).exclude(id=area_id).exists():
                messages.error(request, f'Area "{area_name}" already exists in {city.name}.')
                return redirect('area_list')
            
            area.name = area_name
            area.city = city
            area.pincode = pincode
            area.save()
            
            # Log area update
            if user:
                changes = []
                if old_name != area_name:
                    changes.append(f"Name: {old_name} → {area_name}")
                if old_city.id != city.id:
                    changes.append(f"City: {old_city.name} → {city.name}")
                if old_pincode != pincode:
                    changes.append(f"Pincode: {old_pincode} → {pincode}")
                
                if changes:
                    change_details = ", ".join(changes)
                    user_logs.objects.create(
                        user=user.username,
                        action='Updated Area',
                        details=f'Area: {old_name} | Changes: {change_details}'
                    )
                else:
                    user_logs.objects.create(
                        user=user.username,
                        action='Updated Area',
                        details=f'Area: {area_name}'
                    )
            
            messages.success(request, f'Area "{area_name}" updated successfully.')
        except City.DoesNotExist:
            messages.error(request, 'Selected city does not exist.')
        except Exception as e:
            messages.error(request, f'Error updating area: {str(e)}')
        
        return redirect('area_list')
    
    return redirect('area_list')


def delete_area(request, area_id):
    """Delete an area"""
    username = request.session.get('username')
    user = Custom_User.objects.get(username=username) if username else None
    
    area = get_object_or_404(Area, id=area_id)
    
    if request.method == 'POST':
        area_name = area.name
        city_name = area.city.name
        try:
            # Log before deletion
            if user:
                user_logs.objects.create(
                    user=user.username,
                    action='Deleted Area',
                    details=f'Area Name: {area_name} in City: {city_name}'
                )
            
            area.delete()
            messages.success(request, f'Area "{area_name}" deleted successfully.')
        except Exception as e:
            messages.error(request, f'Error deleting area: {str(e)}')
    
    return redirect('area_list')




def get_cities(request, state_id):
    cities = City.objects.filter(state_id=state_id).values("id", "name")
    return JsonResponse(list(cities), safe=False)


def get_areas(request, city_id):
    areas = Area.objects.filter(city_id=city_id).values("id", "name")
    return JsonResponse(list(areas), safe=False)



def customer_locations(request, customer_id):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')

    username = request.session.get('username')
    user = Custom_User.objects.get(username=username)
    if not user:
        return redirect('login')

    customer = get_object_or_404(Customer, id=customer_id)

    if request.method == "POST":
        location_name    = request.POST.get("customer_name")
        address          = request.POST.get("address")
        area_id          = request.POST.get("area")
        city_id          = request.POST.get("city")
        state_id         = request.POST.get("state")
        pincode          = request.POST.get("pincode")
        location_contact = request.POST.get("contact")
        location_email   = request.POST.get("email")
        status           = request.POST.get("status")
        is_active        = (status == "Active")
        subscription_plan_id = request.POST.get("subscription_plan")
        location_link    = request.POST.get("location_link")
        start_date       = request.POST.get("start_date")

        # ── Custom subscription amount ────────────────────────────────────
        custom_subscription_amount = request.POST.get("custom_subscription_amount")
        if custom_subscription_amount and custom_subscription_amount.strip():
            try:
                custom_subscription_amount = Decimal(custom_subscription_amount)
                if custom_subscription_amount < 0:
                    messages.error(request, 'Custom subscription amount cannot be negative')
                    return redirect('customer_locations', customer_id=customer.id)
            except (ValueError, InvalidOperation):
                messages.error(request, 'Invalid custom subscription amount')
                return redirect('customer_locations', customer_id=customer.id)
        else:
            custom_subscription_amount = None

        # ── Outstanding amount ────────────────────────────────────────────
        outstanding_amount = request.POST.get("outstanding_amount")
        if outstanding_amount and outstanding_amount.strip():
            try:
                outstanding_amount = Decimal(outstanding_amount)
                if outstanding_amount < 0:
                    messages.error(request, 'Outstanding amount cannot be negative')
                    return redirect('customer_locations', customer_id=customer.id)
            except (ValueError, InvalidOperation):
                messages.error(request, 'Invalid outstanding amount')
                return redirect('customer_locations', customer_id=customer.id)
        else:
            outstanding_amount = Decimal('0.00')

        # ── GST fields ────────────────────────────────────────────────────
        # The template sends gst_registered=1 when the toggle is on;
        # if the key is absent the location is not GST registered.
        gst_registered = request.POST.get("gst_registered") == "1"
        gst_number     = request.POST.get("gst_number", "").strip() or None

        # If the toggle was accidentally submitted without a number, clear it
        if not gst_registered:
            gst_number = None
        # ─────────────────────────────────────────────────────────────────

        # ── Subscription plan ─────────────────────────────────────────────
        subscription_plan = None
        if subscription_plan_id:
            subscription_plan = SubscriptionPlan.objects.get(id=subscription_plan_id)

        # ── Create location ───────────────────────────────────────────────
        customer_location = CustomerLocation.objects.create(
            customer=customer,
            location_name=location_name,
            address=address,
            area_id=area_id   if area_id   else None,
            city_id=city_id   if city_id   else None,
            state_id=state_id if state_id  else None,
            pincode=pincode,
            location_contact=location_contact,
            location_email=location_email,
            is_active=is_active,
            subscription_plan=subscription_plan,
            custom_subscription_amount=custom_subscription_amount,
            outstanding_amount=outstanding_amount,
            location_link=location_link if location_link else None,
            start_date=start_date       if start_date    else None,
            # ── new GST fields ───────────────────────────────────────────
            gst_registered=gst_registered,
            gst_number=gst_number,
        )

        # ── Camera serials (Live Monitoring only) ─────────────────────────
        if subscription_plan and 'live' in subscription_plan.name.lower():
            camera_serials = request.POST.getlist("camera_serials[]")
            for serial in camera_serials:
                if serial.strip():
                    customer_cameras.objects.create(
                        customer=customer,
                        seriak_number=serial,
                        customer_location=customer_location,
                    )

        # ── Audit log ─────────────────────────────────────────────────────
        user_logs.objects.create(
            user=user.username,
            action='Created Customer Location',
            details=(
                'Location: ' + location_name
                + ' created under Customer: ' + customer.name
                + (' | GST: ' + (gst_number or 'Registered') if gst_registered else ' | GST: Not Registered')
            ),
        )

        # ── Auto-generate bills ───────────────────────────────────────────
        try:
            generate_automatic_bill_locations(
                entity=customer_location,
                entity_type="Customer_Locations",
                user=user,
            )

            if custom_subscription_amount:
                amount_msg = 'Custom amount: Rs.' + str(custom_subscription_amount)
            else:
                amount_msg = 'Plan amount: Rs.' + str(subscription_plan.price) if subscription_plan else ''

            success_msg = 'Customer location added successfully. ' + amount_msg

            if outstanding_amount > 0:
                success_msg += (
                    ' Outstanding amount Rs.' + str(outstanding_amount)
                    + ' will be added to the first bill.'
                )

            if gst_registered:
                success_msg += ' GST registered.'

            messages.success(request, success_msg)

        except Exception as e:
            messages.warning(
                request,
                'Customer location added but billing failed: ' + str(e),
            )

        return redirect('customer_locations', customer_id=customer.id)

    # ── GET — render page ─────────────────────────────────────────────────
    customer_locations = CustomerLocation.objects.filter(customer=customer)
    states             = State.objects.all()
    citys              = City.objects.all()
    areas              = Area.objects.all()
    plans              = SubscriptionPlan.objects.all()
    camera_serials     = customer_cameras.objects.all()

    selected_gst_type = request.session.get('selected_gst_type', 'gst')

    companies = Company.objects.filter(status='Active')
    if selected_gst_type == 'gst':
        companies = companies.filter(gst_registered=True)
    elif selected_gst_type == 'non-gst':
        companies = companies.filter(gst_registered=False)
    companies = companies.order_by('name')

    filtered_companies = get_filtered_companies(request)
    company_ids        = list(filtered_companies.values_list('id', flat=True))

    context = {
        'customer':             customer,
        'subscription_plans':   plans,
        'states':               states,
        'citys':                citys,
        'areas':                areas,
        'customer_locations':   customer_locations,
        'camera_serials':       camera_serials,
        'companies':            companies,
        'filtered_company_ids': company_ids,
        'selected_gst_type':    selected_gst_type,
        'user':                 user,
        'role_permissions':     role_permissions,
    }

    context.update(get_company_filter_context(request))

    return render(request, 'cust_locations.html', context)



def _generate_location_bills_bg(location_ids, user_id):
    """
    Run generate_automatic_bill_locations for every newly created location.
    Runs in a daemon thread — manages its own DB connection.
    """
    print(f"[BG LOC BILLING] 🚀 Thread started for {len(location_ids)} location(s)")
    close_old_connections()

    user = None
    if user_id:
        try:
            user = Custom_User.objects.get(id=user_id)
        except Exception as exc:
            print(f"[BG LOC BILLING] ⚠️  Could not fetch user id={user_id}: {exc}")

    for loc_id in location_ids:
        try:
            close_old_connections()

            location = CustomerLocation.objects.select_related(
                'customer', 'subscription_plan', 'state', 'city', 'area'
            ).get(id=loc_id)

            print(
                f"[BG LOC BILLING] Processing: {location} | "
                f"start_date={location.start_date} | plan={location.subscription_plan}"
            )

            if not location.start_date:
                print(f"[BG LOC BILLING] ⏭  Skipping — no start_date")
                continue

            if not location.subscription_plan:
                print(f"[BG LOC BILLING] ⏭  Skipping — no subscription_plan")
                continue

            first_day_current_month = date.today().replace(day=1)
            if location.start_date >= first_day_current_month:
                print(
                    f"[BG LOC BILLING] ⏭  Skipping — start_date "
                    f"{location.start_date} is in current/future month"
                )
                continue

            print(f"[BG LOC BILLING] ✅ Generating bills for {location} …")
            generate_automatic_bill_locations(
                entity=location,
                entity_type="Customer_Locations",
                user=user,
            )
            print(f"[BG LOC BILLING] ✅ Done: {location}")

        except CustomerLocation.DoesNotExist:
            print(f"[BG LOC BILLING] ❌ Location ID {loc_id} not found")
        except Exception as exc:
            import traceback
            print(f"[BG LOC BILLING] ❌ Error for location ID {loc_id}: {exc}")
            traceback.print_exc()

    close_old_connections()
    print("[BG LOC BILLING] 🏁 Thread finished")



def bulk_upload_locations(request, customer_id):
    """
    Accepts an Excel file with the same column layout as the template.
    Creates CustomerLocation records for the given customer, then fires
    background bill generation.

    URL: POST /customers/<customer_id>/locations/bulk-upload/
    """
    if request.method != 'POST' or not request.FILES.get('excel_file'):
        messages.error(request, 'No file uploaded.')
        return redirect('customer_locations', customer_id=customer_id)

    customer = get_object_or_404(Customer, id=customer_id)

    # resolve logged-in user
    user, _ = get_logged_in_user(request)
    db_user = None
    if user:
        try:
            db_user = Custom_User.objects.get(username=user)
        except Custom_User.DoesNotExist:
            pass

    try:
        df = pd.read_excel(request.FILES['excel_file'], header=0)
        df.columns = df.columns.str.strip()
        df = df.dropna(how='all')

        # Column presence check
        required = ['State', 'City', 'Area']
        missing  = [c for c in required if c not in df.columns]
        if missing:
            messages.error(request, f'Missing required columns: {", ".join(missing)}')
            return redirect('customer_locations', customer_id=customer_id)

        success_count    = 0
        error_count      = 0
        errors           = []
        new_location_ids = []

        for idx, row in df.iterrows():
            row_num = idx + 2  # 1-indexed, header is row 1
            try:
                # ── Location name (optional; fall back to "Location N") ────
                location_name = ''
                if 'Customer' in df.columns and pd.notna(row.get('Customer')):
                    location_name = str(row['Customer']).strip()
                if not location_name or location_name.lower() == 'nan':
                    location_name = f'Location {row_num - 1}'

                # ── State ─────────────────────────────────────────────────
                state_name = str(row['State']).strip()
                if not state_name or state_name.lower() == 'nan':
                    raise ValueError("State is required")
                state, _ = State.objects.get_or_create(name__iexact=state_name, defaults={'name': state_name})

                # ── City ──────────────────────────────────────────────────
                city_name = str(row['City']).strip()
                if not city_name or city_name.lower() == 'nan':
                    raise ValueError("City is required")
                city = City.objects.filter(name__iexact=city_name, state=state).first()
                if not city:
                    city = City.objects.create(name=city_name, state=state)

                # ── Area ──────────────────────────────────────────────────
                area_name = str(row['Area']).strip()
                if not area_name or area_name.lower() == 'nan':
                    raise ValueError("Area is required")
                area = Area.objects.filter(name__iexact=area_name, city=city).first()
                if not area:
                    area = Area.objects.create(name=area_name, city=city)

                # ── Contact / Email ───────────────────────────────────────
                phone = ''
                if 'Phone' in df.columns and pd.notna(row.get('Phone')):
                    phone = str(row['Phone']).strip().replace(' ', '').replace('-', '')
                    if phone.lower() == 'nan':
                        phone = ''

                # auto-generate email from location name
                location_email = (
                    location_name.replace(' ', '_').lower() + '@location.com'
                )

                # ── Address / Pincode ─────────────────────────────────────
                address = ''
                if 'Address' in df.columns and pd.notna(row.get('Address')):
                    address = str(row['Address']).strip()
                    if address.lower() == 'nan':
                        address = ''

                pincode = ''
                if 'Pincode' in df.columns and pd.notna(row.get('Pincode')):
                    pincode = str(row['Pincode']).strip()
                    if pincode.lower() == 'nan':
                        pincode = ''

                # ── Location Link ─────────────────────────────────────────
                location_link = None
                if 'Location Link' in df.columns and pd.notna(row.get('Location Link')):
                    ll = str(row['Location Link']).strip()
                    if ll and ll.lower() not in ['nan', 'none', '']:
                        if not ll.startswith(('http://', 'https://')):
                            ll = 'https://' + ll
                        location_link = ll

                # ── Start Date ────────────────────────────────────────────
                start_date = None
                if 'Service Start' in df.columns and pd.notna(row.get('Service Start')):
                    try:
                        start_date = pd.to_datetime(row['Service Start'], dayfirst=True).date()
                    except Exception:
                        start_date = None

                # ── Subscription Plan ─────────────────────────────────────
                subscription_plan = None
                if 'Accounting' in df.columns and pd.notna(row.get('Accounting')):
                    plan_name = str(row['Accounting']).strip()
                    if plan_name and plan_name.lower() != 'nan':
                        subscription_plan = SubscriptionPlan.objects.filter(
                            name__icontains=plan_name
                        ).first()
                        if not subscription_plan:
                            print(f"[LOC UPLOAD] ⚠️  Row {row_num}: Plan '{plan_name}' not found in DB")

                # ── Status ────────────────────────────────────────────────
                is_active = True
                if 'Service' in df.columns and pd.notna(row.get('Service')):
                    is_active = str(row['Service']).strip().lower() in ['active', 'running']

                # ── Financial ─────────────────────────────────────────────
                outstanding_amount = Decimal('0.00')
                if 'Outstanding Amount' in df.columns and pd.notna(row.get('Outstanding Amount')):
                    try:
                        outstanding_amount = Decimal(str(row['Outstanding Amount']))
                    except Exception:
                        outstanding_amount = Decimal('0.00')

                custom_subscription_amount = None
                if 'Custom Amount' in df.columns and pd.notna(row.get('Custom Amount')):
                    try:
                        custom_subscription_amount = Decimal(str(row['Custom Amount']))
                    except Exception:
                        custom_subscription_amount = None

                # ── GST ───────────────────────────────────────────────────
                gst_registered = False
                gst_number     = None

                if 'GST Registered' in df.columns and pd.notna(row.get('GST Registered')):
                    if str(row['GST Registered']).strip().lower() in ['yes', 'y', 'true', '1', 'registered']:
                        gst_registered = True

                if 'GST Number' in df.columns and pd.notna(row.get('GST Number')):
                    raw_gst = str(row['GST Number']).strip().upper()
                    if raw_gst and raw_gst.lower() not in ['nan', 'none', '']:
                        if len(raw_gst) != 15:
                            error_count += 1
                            errors.append(
                                f"Row {row_num}: Invalid GST number '{raw_gst}' (must be 15 characters)"
                            )
                            continue
                        gst_registered = True
                        gst_number     = raw_gst

                # ── Tax percentage ────────────────────────────────────────
                tax_percentage = None
                if 'Tax Percentage' in df.columns and pd.notna(row.get('Tax Percentage')):
                    try:
                        tax_percentage = float(row['Tax Percentage'])
                    except Exception:
                        tax_percentage = None

                # ── Remarks ───────────────────────────────────────────────
                remarks = ''
                if 'Remarks' in df.columns and pd.notna(row.get('Remarks')):
                    remarks = str(row['Remarks']).strip()
                    if remarks.lower() == 'nan':
                        remarks = ''

                # ── CREATE ────────────────────────────────────────────────
                loc = CustomerLocation.objects.create(
                    customer                   = customer,
                    location_name              = location_name,
                    address                    = address,
                    area                       = area,
                    city                       = city,
                    state                      = state,
                    pincode                    = pincode,
                    location_contact           = phone,
                    location_email             = location_email,
                    is_active                  = is_active,
                    subscription_plan          = subscription_plan,
                    custom_subscription_amount = custom_subscription_amount,
                    outstanding_amount         = outstanding_amount,
                    location_link              = location_link,
                    start_date                 = start_date,
                    gst_registered             = gst_registered,
                    gst_number                 = gst_number,
                    remarks                    = remarks,
                )

                # audit
                user_logs.objects.create(
                    user   = user if user else 'system',
                    action = 'Bulk Upload - Created Customer Location',
                    details= f'Location: {location_name} created under Customer: {customer.name} via bulk upload',
                )

                if start_date and subscription_plan:
                    new_location_ids.append(loc.id)
                    print(f"[LOC UPLOAD] ✅ Created & queued: {loc}")
                else:
                    print(f"[LOC UPLOAD] ✅ Created (not queued — missing start_date or plan): {loc}")

                success_count += 1

            except Exception as exc:
                error_count += 1
                import traceback
                traceback.print_exc()
                errors.append(f"Row {row_num}: {exc}")

        # ── Result messages ───────────────────────────────────────────────
        if success_count:
            messages.success(request, f'✅ Successfully imported {success_count} location(s)!')
        if error_count:
            summary = f'⚠️ Failed to import {error_count} location(s). '
            summary += 'Errors: ' + ' | '.join(errors[:5])
            if len(errors) > 5:
                summary += f' … and {len(errors) - 5} more.'
            messages.warning(request, summary)

        # ── Fire background billing ───────────────────────────────────────
        if new_location_ids:
            bg_user_id = db_user.id if db_user else None
            t = threading.Thread(
                target   = _generate_location_bills_bg,
                args     = (new_location_ids, bg_user_id),
                daemon   = True,
                name     = "BulkLocationBillingThread",
            )
            t.start()
            messages.info(
                request,
                f'🔄 Bill generation started in background for '
                f'{len(new_location_ids)} eligible location(s).'
            )

    except Exception as exc:
        import traceback
        traceback.print_exc()
        messages.error(request, f'❌ Error processing file: {exc}')

    return redirect('customer_locations', customer_id=customer_id)

from datetime import datetime, timedelta, date
from decimal import Decimal
from io import BytesIO
import os
import calendar

from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.utils import timezone
from django.utils.html import strip_tags
from django.db import transaction

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, 
    Spacer, Image, KeepTogether
)

# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def number_to_words(number):
    """Convert number to words (Indian numbering system)"""
    ones  = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine']
    tens  = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']
    teens = ['Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen',
             'Sixteen', 'Seventeen', 'Eighteen', 'Nineteen']

    def convert_below_hundred(n):
        if n < 10:   return ones[n]
        elif n < 20: return teens[n - 10]
        else:        return tens[n // 10] + (' ' + ones[n % 10] if n % 10 != 0 else '')

    def convert_below_thousand(n):
        if n < 100:
            return convert_below_hundred(n)
        return ones[n // 100] + ' Hundred' + (' ' + convert_below_hundred(n % 100) if n % 100 != 0 else '')

    if number == 0:
        return 'Zero Only'

    integer_part = int(number)
    decimal_part = int(round((number - integer_part) * 100))
    result = []

    if integer_part >= 10000000:
        result.append(convert_below_thousand(integer_part // 10000000) + ' Crore')
        integer_part %= 10000000
    if integer_part >= 100000:
        result.append(convert_below_thousand(integer_part // 100000) + ' Lakh')
        integer_part %= 100000
    if integer_part >= 1000:
        result.append(convert_below_thousand(integer_part // 1000) + ' Thousand')
        integer_part %= 1000
    if integer_part > 0:
        result.append(convert_below_thousand(integer_part))

    words = ' '.join(result)
    if decimal_part > 0:
        words += ' and ' + convert_below_hundred(decimal_part) + ' Paise'
    return words + ' Only'


def number_to_words_loc(amount):
    """Convert amount to words for Indian currency"""
    try:
        from num2words import num2words
        rupees = int(amount)
        paise  = int((amount - rupees) * 100)
        words  = num2words(rupees, lang='en_IN').title() + ' Rupees'
        if paise > 0:
            words += ' and ' + num2words(paise, lang='en_IN').title() + ' Paise'
        return words + ' Only'
    except ImportError:
        return 'Rupees ' + f'{amount:.2f}' + ' Only'


# =============================================================================
# SHARED GST HELPER
# =============================================================================

def _resolve_gst_flags(entity, entity_type, organization):
    
    if entity_type in ('customer', 'Customer_Locations'):
        # ── Location-level GST takes priority ──────────────────────────────
        # For Customer_Locations entity IS the location, so check it first.
        # For plain customers entity IS the customer.
        is_gst_applicable = getattr(entity, 'gst_registered', False)
    else:
        # Company / Company_Locations always subject to GST
        is_gst_applicable = True

    is_same_state = False
    if is_gst_applicable and organization and getattr(entity, 'state', None):
        org_state = getattr(organization, 'state', None)
        if org_state:
            is_same_state = (org_state.id == entity.state.id)
        elif hasattr(organization, 'state_name'):
            is_same_state = (str(organization.state_name).lower() == str(entity.state).lower())

    if not is_gst_applicable:
        gst_type = 'NON-GST'
    elif is_same_state:
        gst_type = 'INTRA-STATE'
    else:
        gst_type = 'INTER-STATE'

    return is_gst_applicable, is_same_state, gst_type


def _calc_gst_loc(base_amount, is_same_state):
    """Return (cgst, sgst, igst, total_gst)"""
    base_amount = Decimal(str(base_amount))
    if is_same_state:
        cgst = (base_amount * Decimal('0.09')).quantize(Decimal('0.01'))
        sgst = (base_amount * Decimal('0.09')).quantize(Decimal('0.01'))
        igst = Decimal('0.00')
    else:
        cgst = Decimal('0.00')
        sgst = Decimal('0.00')
        igst = (base_amount * Decimal('0.18')).quantize(Decimal('0.01'))
    return cgst, sgst, igst, cgst + sgst + igst


# =============================================================================
# ADVANCE ADJUSTMENT
# =============================================================================
def adjust_advance_amount_loc(customer, bill_amount):
    """Adjust advance amount against bill amount. Works for both billing paths."""
    customer.refresh_from_db()

    bill_amount    = Decimal(str(bill_amount))
    advance_amount = Decimal(str(customer.advance_amount or 0))

    if advance_amount <= 0:
        return {
            'original_amount':   bill_amount,
            'advance_used':      Decimal('0.00'),
            'final_amount':      bill_amount,
            'remaining_advance': Decimal('0.00'),
            'paid_amount':       Decimal('0.00'),
            'balance_amount':    bill_amount,
        }

    advance_used      = min(advance_amount, bill_amount)
    final_amount      = bill_amount - advance_used
    remaining_advance = advance_amount - advance_used

    customer.advance_amount = remaining_advance
    customer.save(update_fields=['advance_amount'])
    customer.refresh_from_db()

    return {
        'original_amount':   bill_amount,
        'advance_used':      advance_used,
        'final_amount':      final_amount,
        'remaining_advance': remaining_advance,
        'paid_amount':       advance_used,
        'balance_amount':    final_amount,
    }

# Alias used in the locations billing path
adjust_advance_amount_loc = adjust_advance_amount


def calculate_prorated_amount_loc(monthly_price, start_date, month_start, month_end):
    """Calculate prorated amount for a single month"""
    monthly_price   = Decimal(str(monthly_price))
    effective_start = max(start_date, month_start)

    if effective_start > month_end:
        return {'amount': Decimal('0.00'), 'days_used': 0, 'days_in_month': 0, 'is_prorated': False}

    days_in_month = (month_end - month_start).days + 1

    if effective_start == month_start:
        return {'amount': monthly_price, 'days_used': days_in_month,
                'days_in_month': days_in_month, 'is_prorated': False}

    days_used       = (month_end - effective_start).days + 1
    prorated_amount = (monthly_price / Decimal(str(days_in_month))) * Decimal(str(days_used))

    return {
        'amount':        prorated_amount.quantize(Decimal('0.01')),
        'days_used':     days_used,
        'days_in_month': days_in_month,
        'is_prorated':   True,
    }


def get_months_between_loc(start_date, end_date):
    """Get list of (month_start, month_end) tuples between dates"""
    months  = []
    current = start_date.replace(day=1)

    while current <= end_date:
        if current.month == 12:
            month_end = current.replace(year=current.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = current.replace(month=current.month + 1, day=1) - timedelta(days=1)

        month_end = min(month_end, end_date)
        months.append((current, month_end))

        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1, day=1)
        else:
            current = current.replace(month=current.month + 1, day=1)

    return months

def generate_automatic_bill_locations(entity, entity_type, user):
    if not entity.start_date or not entity.subscription_plan:
        return []

    entity.refresh_from_db()

    start_date = (
        datetime.strptime(entity.start_date, '%Y-%m-%d').date()
        if isinstance(entity.start_date, str) else entity.start_date
    )

    today                   = datetime.now().date()
    first_day_current_month = today.replace(day=1)
    last_day_last_month     = first_day_current_month - timedelta(days=1)

    if start_date >= first_day_current_month:
        return []

    try:
        if entity_type == 'Customer_Locations':
            organization = entity.customer.company
            if not organization:
                raise Exception('No company associated with customer ' + entity.customer.name)
            customer = entity.customer
        else:
            organization = entity.company
            customer = None
    except Exception as e:
        raise Exception('Error getting organization for billing: ' + str(e))

    # ── GST flags ─────────────────────────────────────────────────────────────
    # Always check the COMPANY's gst_registered, not the customer/location flag.
    is_gst_applicable = getattr(organization, 'gst_registered', False)

    is_same_state = False
    if is_gst_applicable and organization and entity.state:
        if hasattr(organization, 'state') and organization.state:
            is_same_state = (organization.state.id == entity.state.id)
        elif hasattr(organization, 'state_name'):
            is_same_state = (str(organization.state_name).lower() == str(entity.state).lower())
    # ─────────────────────────────────────────────────────────────────────────

    months_to_bill = get_months_between_loc(start_date, last_day_last_month)

    billing_records = []
    pdf_buffers     = []

    monthly_price = (
        Decimal(str(entity.custom_subscription_amount))
        if hasattr(entity, 'custom_subscription_amount') and entity.custom_subscription_amount
        else Decimal(str(entity.subscription_plan.price))
    )

    outstanding_added  = False
    outstanding_amount = Decimal(str(
        getattr(entity, 'outstanding_amount', 0) or 0
    ))

    for month_index, (month_start, month_end) in enumerate(months_to_bill):
        entity.refresh_from_db()

        proration           = calculate_prorated_amount_loc(monthly_price, start_date, month_start, month_end)
        subscription_amount = proration['amount']

        if subscription_amount <= 0:
            continue

        outstanding_in_this_bill = Decimal('0.00')
        if not outstanding_added and outstanding_amount > 0:
            outstanding_in_this_bill = outstanding_amount
            outstanding_added = True

        # ── GST Calculation ─────────────────────────────────────────────────
        cgst = sgst = igst = Decimal('0.00')
        if is_gst_applicable:
            cgst, sgst, igst, total_gst = _calc_gst_loc(subscription_amount, is_same_state)
            gst_type = 'INTRA-STATE' if is_same_state else 'INTER-STATE'
        else:
            total_gst = Decimal('0.00')
            gst_type  = 'NON-GST'
        # ────────────────────────────────────────────────────────────────────

        total_bill_amount = subscription_amount + total_gst + outstanding_in_this_bill

        if entity_type == 'Customer_Locations' and customer:
            advance_adjustment = adjust_advance_amount_loc(customer, total_bill_amount)
            paid_amount     = advance_adjustment['paid_amount']
            balance_amount  = advance_adjustment['balance_amount']
            discount_amount = advance_adjustment['advance_used']
        else:
            paid_amount = discount_amount = Decimal('0.00')
            balance_amount = total_bill_amount
            advance_adjustment = {
                'original_amount':   total_bill_amount,
                'advance_used':      Decimal('0.00'),
                'final_amount':      total_bill_amount,
                'remaining_advance': Decimal('0.00'),
                'paid_amount':       Decimal('0.00'),
                'balance_amount':    total_bill_amount,
            }

        notes_parts = [
            entity_type.replace('_', ' ').title(),
            'subscription bill',
            month_start.strftime('%B %Y'),
        ]
        if proration['is_prorated']:      notes_parts.append('Prorated')
        if outstanding_in_this_bill > 0:  notes_parts.append('Plus Outstanding Amount')

        billing_kwargs = {
            'amount':             total_bill_amount,
            'discount_amount':    discount_amount,
            'gst_amount':         total_gst,
            'gst_type':           gst_type,
            'billing_date':       today,
            'billing_start_date': max(start_date, month_start),
            'billing_end_date':   month_end,
            'due_date':           today + timedelta(days=15),
            'paid':               balance_amount <= 0,
            'paid_amount':        paid_amount,
            'balance_amount':     balance_amount,
            'payment_date':       today if paid_amount > 0 else None,           # ← advance sets pay date
            'payment_mode':       'Advance Adjustment' if paid_amount > 0 else None,  # ← advance sets mode
            'notes':              ' '.join(notes_parts),
        }

        if entity_type == 'Customer_Locations':
            billing_kwargs['customer_location'] = entity
            billing_kwargs['customer']          = None

        from .models import BillingRecord, BillItem
        billing_record = BillingRecord.objects.create(**billing_kwargs)

        # Subscription line item
        proration_text = (
            ' Prorated ' + str(proration['days_used'])
            + ' days of ' + str(proration['days_in_month']) + ' days'
            if proration['is_prorated'] else ''
        )
        BillItem.objects.create(
            billing_record=billing_record,
            item_name=(entity.subscription_plan.name + ' - ' + month_start.strftime('%B %Y')),
            description=(
                'Subscription period: '
                + max(start_date, month_start).strftime('%d-%b-%Y')
                + ' to ' + month_end.strftime('%d-%b-%Y')
                + proration_text
            ),
            quantity=1,
            unit_price=subscription_amount,
            tax_percentage=18.00 if is_gst_applicable else 0.00,
        )

        # Outstanding line item (first bill only, no tax)
        if outstanding_in_this_bill > 0:
            BillItem.objects.create(
                billing_record=billing_record,
                item_name='Outstanding Amount',
                description='Previous outstanding balance. No GST applicable.',
                quantity=1,
                unit_price=outstanding_in_this_bill,
                tax_percentage=0,
            )
            if hasattr(entity, 'outstanding_amount'):
                entity.outstanding_amount = Decimal('0.00')
                entity.save(update_fields=['outstanding_amount'])

        amounts = {
            'subscription_amount': f'{subscription_amount:.2f}',
            'outstanding_amount':  f'{outstanding_in_this_bill:.2f}',
            'base_amount':         f'{subscription_amount:.2f}',
            'cgst':                f'{cgst:.2f}',
            'sgst':                f'{sgst:.2f}',
            'igst':                f'{igst:.2f}',
            'total_gst':           f'{total_gst:.2f}',
            'subtotal':            f'{total_bill_amount:.2f}',
            'advance_used':        f'{discount_amount:.2f}',
            'total_amount':        f'{balance_amount:.2f}',
            'is_gst_applicable':   is_gst_applicable,
        }

        gst_info = {
            'is_same_state':     is_same_state,
            'is_gst_applicable': is_gst_applicable,
        }

        pdf_buffer = generate_invoice_pdf_locations(
            billing_record=billing_record,
            subscription_plan=entity.subscription_plan,
            entity=entity,
            gst_details=gst_info,
            amounts=amounts,
            organization=organization,
            billing_period=(
                month_start.strftime('%d-%b-%Y')
                + ' to ' + month_end.strftime('%d-%b-%Y')
            ),
            proration_info=proration,
            entity_type=entity_type,
            advance_adjustment=advance_adjustment,
            outstanding_amount=outstanding_in_this_bill,
        )

        billing_records.append(billing_record)
        pdf_buffers.append({
            'buffer':   pdf_buffer,
            'filename': (
                'Invoice_IN-' + str(billing_record.id)
                + '_' + month_start.strftime('%b_%Y') + '.pdf'
            ),
            'month':  month_start.strftime('%B %Y'),
            'amount': balance_amount,
        })

        # ── Data logs (Customer_Locations only) ───────────────────────────
        if entity_type == 'Customer_Locations':
            from .models import data_logs

            bill_status = 'Paid' if balance_amount <= 0 else 'Pending'

            outstanding_detail = (
                'Outstanding: ' + str(outstanding_in_this_bill) + ', '
                if outstanding_in_this_bill > 0 else ''
            )

            # ── LOG 1: Bill generation entry (always created) ───────────────
            data_logs.objects.create(
                user=user,
                timestamp=timezone.now(),
                customer=entity.customer,
                location=entity,
                billing_record=billing_record,
                payment_amount=Decimal('0.00'),
                balance_amount=balance_amount,
                billing_period_start=max(start_date, month_start),
                billing_period_end=month_end,
                is_payment=False,
                is_approved=False,
                status=bill_status,
                payment_mode=None,
                payment_date=None,
                total_paid=Decimal('0.00'),
                action='Auto-generated bill for ' + month_start.strftime('%B %Y'),
                details=(
                    'Subscription: ' + entity.subscription_plan.name
                    + ', Amount: ' + str(subscription_amount)
                    + ', ' + outstanding_detail
                    + 'Total: ' + str(total_bill_amount)
                    + ', Balance: ' + str(balance_amount)
                ),
            )

            # ── LOG 2: Advance adjustment log (only if advance was used) ────
            if paid_amount > 0:
                data_logs.objects.create(
                    user=user,
                    timestamp=timezone.now(),
                    customer=entity.customer,
                    location=entity,
                    billing_record=billing_record,
                    payment_amount=paid_amount,
                    balance_amount=balance_amount,
                    billing_period_start=max(start_date, month_start),
                    billing_period_end=month_end,
                    is_payment=True,             # ← advance is auto-approved
                    status=bill_status,
                    payment_mode='Advance Adjustment',
                    payment_date=today,             # ← payment date is today
                    total_paid=paid_amount,
                    action='Advance adjusted for ' + month_start.strftime('%B %Y'),
                    details=(
                        'Advance Adjustment for Invoice: '
                        + str(billing_record.invoice_number or billing_record.id)
                        + ', Subscription: ' + entity.subscription_plan.name
                        + ', Bill Total: Rs.' + str(total_bill_amount)
                        + ', Advance Used: Rs.' + str(paid_amount)
                        + ', Remaining Advance: Rs.' + str(advance_adjustment['remaining_advance'])
                        + ', Balance After Adjustment: Rs.' + str(balance_amount)
                    ),
                )
        # ─────────────────────────────────────────────────────────────────────

    # Send consolidated email
    if billing_records and entity.location_email and organization:
        if entity_type == 'Company_Locations':
            recipient_name = entity.company.name + ' - ' + entity.location_name
        elif entity_type == 'Customer_Locations':
            recipient_name = entity.customer.name + ' - ' + entity.location_name
        else:
            recipient_name = entity.location_name

        send_consolidated_billing_email_location(
            recipient_email=entity.location_email,
            recipient_name=recipient_name,
            billing_records=billing_records,
            pdf_buffers=pdf_buffers,
            organization=organization,
            customer=customer,
        )

    return billing_records


# =============================================================================
# EMAIL — CUSTOMER BILLING
# =============================================================================

def send_consolidated_billing_email(recipient_email, recipient_name, generated_bills,
                                    subscription_plan, company, entity=None):

    is_gst_applicable = generated_bills[0]['gst_details'].get('is_gst_applicable', True)

    if len(generated_bills) == 1:
        br      = generated_bills[0]['billing_record']
        subject = ('Tax Invoice' if is_gst_applicable else 'Invoice') + ' IN-' + str(br.id) + ' - ' + subscription_plan.name + ' Plan'
    else:
        nums    = ['IN-' + str(b['billing_record'].id) for b in generated_bills]
        subject = ('Tax Invoices' if is_gst_applicable else 'Invoices') + ' ' + ', '.join(nums) + ' - ' + subscription_plan.name + ' Plan'

    total_billed  = sum(Decimal(b['amounts']['subtotal'])      for b in generated_bills)
    total_paid    = sum(Decimal(b['amounts']['advance_used'])  for b in generated_bills)
    total_balance = sum(Decimal(b['amounts']['total_amount'])  for b in generated_bills)

    invoices_html = ''
    for idx, bill in enumerate(generated_bills, 1):
        br          = bill['billing_record']
        amounts     = bill['amounts']
        gst_details = bill['gst_details']
        gst_app     = gst_details.get('is_gst_applicable', True)

        gst_breakdown = ''
        if gst_app:
            if gst_details['is_same_state']:
                gst_breakdown = (
                    '<div class="detail-row"><span class="detail-label">CGST (9%):</span>'
                    '<span>Rs. ' + amounts['cgst'] + '</span></div>'
                    '<div class="detail-row"><span class="detail-label">SGST (9%):</span>'
                    '<span>Rs. ' + amounts['sgst'] + '</span></div>'
                )
            else:
                gst_breakdown = (
                    '<div class="detail-row"><span class="detail-label">IGST (18%):</span>'
                    '<span>Rs. ' + amounts['igst'] + '</span></div>'
                )

        advance_section = ''
        if Decimal(amounts['advance_used']) > 0:
            advance_section = (
                '<div class="detail-row" style="border-top:2px solid #dee2e6;margin-top:10px;padding-top:10px;">'
                '<span class="detail-label">Subtotal:</span><span>Rs. ' + amounts['subtotal'] + '</span></div>'
                '<div class="detail-row"><span class="detail-label" style="color:#28a745;">Advance Adjusted:</span>'
                '<span style="color:#28a745;">- Rs. ' + amounts['advance_used'] + '</span></div>'
            )

        period_desc = (
            bill['billing_start'].strftime('%d-%b-%Y')
            + ' to ' + bill['billing_end'].strftime('%d-%b-%Y')
        )
        if bill.get('days_count'):
            period_desc += ' (' + str(bill['days_count']) + ' days)'

        invoice_label = 'Tax Invoice' if gst_app else 'Invoice'

        invoices_html += (
            '<div class="invoice-details">'
            '<h3 style="color:#3d6a8e;margin-top:0;">' + invoice_label + ' ' + str(idx) + ': IN-' + str(br.id) + '</h3>'
            '<div class="detail-row"><span class="detail-label">Billing Period:</span><span>' + period_desc + '</span></div>'
            '<div class="detail-row"><span class="detail-label">Invoice Date:</span><span>' + br.billing_date.strftime('%d-%b-%Y') + '</span></div>'
            '<div class="detail-row"><span class="detail-label">Due Date:</span><span>' + br.due_date.strftime('%d-%b-%Y') + '</span></div>'
            '<div class="detail-row"><span class="detail-label">Base Amount:</span><span>Rs. ' + amounts['base_amount'] + '</span></div>'
            + gst_breakdown
            + advance_section +
            '<div class="detail-row" style="border-top:2px solid #3d6a8e;margin-top:10px;padding-top:10px;">'
            '<span class="detail-label"><strong>Amount Due:</strong></span>'
            '<span><strong>Rs. ' + amounts['total_amount'] + '</strong></span></div>'
            '</div>'
        )

    gst_info_block = (
        '<div class="gst-info"><strong>GST Information:</strong><br>'
        + ('Same State Transaction - CGST (9%) + SGST (9%) = 18%'
           if generated_bills[0]['gst_details']['is_same_state']
           else 'Interstate Transaction - IGST (18%)')
        + '</div>'
        if is_gst_applicable else
        '<div class="gst-info" style="background-color:#f0f0f0;border-left-color:#999;">'
        '<strong>Note:</strong> This customer is not GST registered. No GST has been applied.</div>'
    )

    advance_balance_html = ''
    if entity and hasattr(entity, 'advance_amount') and entity.advance_amount and entity.advance_amount > 0:
        advance_balance_html = (
            '<div class="summary-row" style="margin-top:10px;font-size:14px;">'
            '<span>Remaining Advance Balance:</span>'
            '<span style="color:#28a745;">Rs. ' + f'{entity.advance_amount:.2f}' + '</span></div>'
        )

    paid_row_html = ''
    if total_paid > 0:
        paid_row_html = (
            '<div class="summary-row"><span><strong>Paid (Advance):</strong></span>'
            '<span style="color:#28a745;">Rs. ' + f'{total_paid:.2f}' + '</span></div>'
        )

    invoice_header_label = 'Tax Invoice' if is_gst_applicable else 'Invoice'
    plural = 's' if len(generated_bills) > 1 else ''

    html_content = f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><style>
body{{font-family:Arial,sans-serif;line-height:1.6;color:#333;max-width:700px;margin:0 auto;padding:20px;}}
.header{{background-color:#3d6a8e;color:white;padding:20px;text-align:center;border-radius:5px 5px 0 0;}}
.content{{background-color:#f8f9fa;padding:30px;border:1px solid #dee2e6;}}
.invoice-details{{background-color:white;padding:20px;border-radius:5px;margin:20px 0;border-left:4px solid #3d6a8e;}}
.detail-row{{display:flex;justify-content:space-between;padding:8px 0;border-bottom:1px solid #e9ecef;}}
.detail-row:last-child{{border-bottom:none;}}
.detail-label{{font-weight:bold;color:#6c757d;}}
.summary-box{{background-color:#e7f1ff;padding:20px;border-radius:5px;margin:20px 0;}}
.summary-row{{display:flex;justify-content:space-between;padding:8px 0;}}
.total-amount{{font-size:24px;font-weight:bold;color:#3d6a8e;}}
.footer{{text-align:center;color:#6c757d;font-size:12px;margin-top:30px;padding-top:20px;border-top:1px solid #dee2e6;}}
.attachment-note{{background-color:#fff3cd;border-left:4px solid #ffc107;padding:15px;margin:20px 0;}}
.gst-info{{background-color:#d1ecf1;border-left:4px solid #0c5460;padding:15px;margin:20px 0;}}
</style></head>
<body>
<div class="header"><h1>{company.name}</h1><p>{invoice_header_label}{plural} Notification</p></div>
<div class="content">
<p>Dear {recipient_name},</p>
<p>{'Invoices have' if len(generated_bills) > 1 else 'An invoice has'} been generated for your subscription with {company.name}.</p>
{gst_info_block}
{invoices_html}
<div class="summary-box">
<div class="summary-row"><span><strong>Total Billed:</strong></span><span>Rs. {total_billed:.2f}</span></div>
{paid_row_html}
<div class="summary-row" style="border-top:2px solid #3d6a8e;padding-top:12px;margin-top:12px;">
<span class="total-amount">Total Amount Due:</span><span class="total-amount">Rs. {total_balance:.2f}</span></div>
{advance_balance_html}
</div>
<div class="attachment-note"><strong>{len(generated_bills)} PDF Invoice{plural} Attached</strong><br>
Please find your detailed invoice{plural} attached to this email.</div>
<p><strong>Payment Instructions:</strong></p>
<ul>
<li>Payment is due by {generated_bills[-1]['billing_record'].due_date.strftime('%d-%b-%Y')}</li>
<li>Please include invoice reference{plural} in the payment description</li>
{'<li>Scan the QR code in the attached PDF for quick payment</li>' if hasattr(company, 'qr_code') and company.qr_code else ''}
<li>Contact us at {company.email} or {company.contact} if you have any questions</li>
</ul>
<p><strong>Subscription Plan:</strong> {subscription_plan.name}</p>
<p>Thank you for choosing {company.name}!</p>
<p>Best regards,<br><strong>{company.name} Team</strong></p>
</div>
<div class="footer">
<p>This is an automated message. Please do not reply to this email.</p>
<p>{company.name} | {company.contact} | {company.email}</p>
<p>Copyright {datetime.now().year} {company.name}. All rights reserved.</p>
</div>
</body></html>"""

    text_content = strip_tags(html_content)
    email = EmailMultiAlternatives(
        subject=subject,
        body=text_content,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[recipient_email],
    )
    email.attach_alternative(html_content, 'text/html')
    for bill in generated_bills:
        email.attach(
            'Invoice_IN-' + str(bill['billing_record'].id) + '.pdf',
            bill['pdf_buffer'].getvalue(),
            'application/pdf',
        )
    email.send()


# =============================================================================
# EMAIL — LOCATION BILLING
# =============================================================================

def send_consolidated_billing_email_location(recipient_email, recipient_name, billing_records,
                                             pdf_buffers, organization, customer=None):
    subject = 'Monthly Invoices - ' + str(len(billing_records)) + ' Month(s)'

    total_billed  = sum(Decimal(str(br.amount))             for br in billing_records)
    total_paid    = sum(Decimal(str(br.paid_amount or 0))   for br in billing_records)
    total_balance = sum(Decimal(str(pdf['amount']))         for pdf in pdf_buffers)

    month_summary = ''
    for i, pdf_data in enumerate(pdf_buffers, 1):
        br = billing_records[i - 1]
        month_summary += (
            '<tr>'
            '<td style="padding:8px;border-bottom:1px solid #e9ecef;">' + str(i) + '</td>'
            '<td style="padding:8px;border-bottom:1px solid #e9ecef;">' + pdf_data['month'] + '</td>'
            '<td style="padding:8px;border-bottom:1px solid #e9ecef;">IN-' + str(br.id) + '</td>'
            '<td style="padding:8px;border-bottom:1px solid #e9ecef;text-align:right;">Rs. ' + f'{br.amount:.2f}' + '</td>'
            '<td style="padding:8px;border-bottom:1px solid #e9ecef;text-align:right;">Rs. ' + f'{(br.paid_amount or 0):.2f}' + '</td>'
            '<td style="padding:8px;border-bottom:1px solid #e9ecef;text-align:right;">Rs. ' + f'{pdf_data["amount"]:.2f}' + '</td>'
            '</tr>'
        )

    remaining_advance_html = ''
    if customer and customer.advance_amount and customer.advance_amount > 0:
        remaining_advance_html = (
            '<div class="summary-row" style="margin-top:10px;font-size:14px;">'
            '<span>Remaining Advance Balance:</span>'
            '<span style="color:#28a745;">Rs. ' + f'{customer.advance_amount:.2f}' + '</span></div>'
        )

    attachments_html = ''.join('<li>' + pdf['filename'] + '</li>' for pdf in pdf_buffers)

    html_content = f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><style>
body{{font-family:Arial,sans-serif;line-height:1.6;color:#333;max-width:750px;margin:0 auto;padding:20px;}}
.header{{background-color:#3d6a8e;color:white;padding:20px;text-align:center;border-radius:5px 5px 0 0;}}
.content{{background-color:#f8f9fa;padding:30px;border:1px solid #dee2e6;}}
table{{width:100%;border-collapse:collapse;background:white;margin:20px 0;}}
th{{background-color:#3d6a8e;color:white;padding:12px;text-align:left;}}
.summary-box{{background-color:#e7f1ff;padding:20px;border-radius:5px;margin:20px 0;}}
.summary-row{{display:flex;justify-content:space-between;padding:8px 0;}}
.total-amount{{font-size:24px;font-weight:bold;color:#3d6a8e;}}
</style></head>
<body>
<div class="header"><h1>{organization.name}</h1><p>Consolidated Monthly Invoices</p></div>
<div class="content">
<p>Dear {recipient_name},</p>
<p>Please find attached {len(billing_records)} invoice(s) for your subscription with {organization.name}.</p>
<h3>Invoice Summary</h3>
<table>
<thead><tr><th>#</th><th>Month</th><th>Invoice No.</th>
<th style="text-align:right;">Total</th><th style="text-align:right;">Paid</th>
<th style="text-align:right;">Balance</th></tr></thead>
<tbody>{month_summary}</tbody>
</table>
<div class="summary-box">
<div class="summary-row"><span><strong>Total Billed:</strong></span><span>Rs. {total_billed:.2f}</span></div>
<div class="summary-row"><span><strong>Paid (Advance):</strong></span><span style="color:#28a745;">Rs. {total_paid:.2f}</span></div>
<div class="summary-row" style="border-top:2px solid #3d6a8e;padding-top:12px;margin-top:12px;">
<span class="total-amount">Total Balance:</span><span class="total-amount">Rs. {total_balance:.2f}</span></div>
{remaining_advance_html}
</div>
<p><strong>Attached Documents:</strong></p><ul>{attachments_html}</ul>
<p><strong>Payment Instructions:</strong></p>
<ul>
<li>Payment is due within 15 days from invoice date</li>
<li>Please reference the invoice numbers in your payment</li>
<li>Contact us at {organization.email} or {organization.contact} for any queries</li>
</ul>
<p>Thank you for your continued business!</p>
<p>Best regards,<br><strong>{organization.name} Team</strong></p>
</div>
<div style="text-align:center;color:#6c757d;font-size:12px;margin-top:30px;padding-top:20px;border-top:1px solid #dee2e6;">
<p>This is an automated message. Please do not reply to this email.</p>
<p>{organization.name} | {organization.contact} | {organization.email}</p>
<p>Copyright {datetime.now().year} {organization.name}. All rights reserved.</p>
</div>
</body></html>"""

    text_content = strip_tags(html_content)
    email = EmailMultiAlternatives(
        subject=subject,
        body=text_content,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[recipient_email],
    )
    email.attach_alternative(html_content, 'text/html')
    for pdf_data in pdf_buffers:
        email.attach(pdf_data['filename'], pdf_data['buffer'].getvalue(), 'application/pdf')
    email.send()


# =============================================================================
# PDF — CUSTOMER BILLING
# =============================================================================

def generate_invoice_pdf(billing_record, subscription_plan, entity, gst_details, amounts,
                         company, billing_period, months_count, days_count=None,
                         advance_adjustment=None, monthly_price=None):
    """Generate PDF invoice. GST columns omitted for non-GST customers."""
    buffer = BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=A4,
                               rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    PAGE_W = A4[0] - doc.leftMargin - doc.rightMargin
    styles = getSampleStyleSheet()

    small_left  = ParagraphStyle('small_left',  parent=styles['Normal'], fontSize=9,  leading=12, alignment=TA_LEFT)
    small_right = ParagraphStyle('small_right', parent=styles['Normal'], fontSize=9,  leading=12, alignment=TA_RIGHT)
    desc_style  = ParagraphStyle('desc',        parent=styles['Normal'], fontSize=8,  leading=10, alignment=TA_LEFT)

    elements          = []
    is_gst_applicable = gst_details.get('is_gst_applicable', True)

    def safe_str(value):
        if not value: return ''
        for attr in ('name', 'city_name', 'state_name'):
            if hasattr(value, attr): return str(getattr(value, attr))
        return str(value)

    # Logo
    logo = ''
    if getattr(company, 'logo', None):
        try:
            if hasattr(company.logo, 'path') and os.path.exists(company.logo.path):
                logo = Image(company.logo.path, 1.2 * inch, 1.2 * inch, kind='proportional')
        except Exception as e:
            print('Logo error: ' + str(e))

    # Header
    invoice_title = 'TAX INVOICE' if is_gst_applicable else 'INVOICE'
    company_name_block = Paragraph(
        '<font size="16"><b>' + company.name + '</b></font><br/>'
        '<font size="9">' + safe_str(getattr(company, 'address', '')) + '<br/>'
        'Contact: ' + safe_str(getattr(company, 'contact', '')) + '<br/>'
        'GSTIN: ' + safe_str(getattr(company, 'gst_number', '')) + '</font>',
        small_left
    )
    invoice_block = Paragraph(
        '<b>' + invoice_title + '</b><br/>'
        '<font size="9">Original for Recipient</font><br/><br/>'
        '<b>IN-' + str(billing_record.id) + '</b>',
        ParagraphStyle('hdr', fontSize=14, alignment=TA_RIGHT, fontName='Helvetica-Bold')
    )
    header = Table([[logo, company_name_block, invoice_block]],
                   colWidths=[1.5*inch, PAGE_W*0.45, PAGE_W*0.30])
    header.setStyle(TableStyle([
        ('VALIGN',       (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING',  (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    elements += [header, Spacer(1, 10)]

    # Amount Due bar
    final_amount = amounts.get('total_amount', '0.00')
    amount_due = Table([[
        Paragraph('<b>Amount Due</b>', small_left),
        Paragraph('<b>Rs. ' + final_amount + '</b>',
                  ParagraphStyle('amt', fontSize=14, alignment=TA_RIGHT,
                                 textColor=colors.white, fontName='Helvetica-Bold'))
    ]], colWidths=[PAGE_W*0.55, PAGE_W*0.45])
    amount_due.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), colors.HexColor('#2f5f7a')),
        ('TEXTCOLOR',     (0,0), (-1,-1), colors.white),
        ('LEFTPADDING',   (0,0), (-1,-1), 10),
        ('RIGHTPADDING',  (0,0), (-1,-1), 10),
        ('TOPPADDING',    (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    elements += [amount_due, Spacer(1, 12)]

    # Company info + dates
    company_block = Paragraph(
        '<b>' + company.name + '</b><br/>'
        + safe_str(getattr(company, 'address', '')) + '<br/>'
        'GSTIN: ' + safe_str(getattr(company, 'gst_number', '')),
        small_left
    )
    date_block = Paragraph(
        'Issue Date: ' + billing_record.billing_date.strftime('%d-%b-%Y') + '<br/>'
        'Due Date: '   + billing_record.due_date.strftime('%d-%b-%Y'),
        small_right
    )
    company_info = Table([[company_block, date_block]], colWidths=[PAGE_W*0.62, PAGE_W*0.38])
    company_info.setStyle(TableStyle([('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0)]))
    elements += [company_info, Spacer(1, 10)]

    # Client details
    client_name_block = Paragraph(
        '<b>Client Details</b><br/>'
        '<b>' + entity.name + '</b><br/>'
        'Email: ' + safe_str(getattr(entity, 'email', '')),
        small_left
    )
    billing_address_block = Paragraph(
        '<b>Billing Address</b><br/>'
        + safe_str(entity.address) + '<br/>'
        + safe_str(entity.city) + ', ' + safe_str(entity.state) + '<br/>'
        + safe_str(entity.pincode),
        small_right
    )
    client_row = Table([[client_name_block, billing_address_block]],
                       colWidths=[PAGE_W*0.5, PAGE_W*0.5])
    client_row.setStyle(TableStyle([
        ('VALIGN',       (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING',  (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('ALIGN',        (1,0),  (1, 0), 'RIGHT'),
    ]))
    elements += [client_row, Spacer(1, 12)]

    # Billing table
    headers = ['S.No', 'Item Description', 'Price', 'Taxable']
    if is_gst_applicable:
        headers += (['CGST @ 9%', 'SGST @ 9%'] if gst_details['is_same_state'] else ['IGST @ 18%'])
    headers.append('Amount')

    period_desc = subscription_plan.name
    if monthly_price and hasattr(entity, 'custom_subscription_amount') and entity.custom_subscription_amount:
        period_desc += ' (Custom: Rs.' + str(monthly_price) + '/month)'
    if days_count:
        period_desc += ' (' + str(days_count) + ' days)'
    elif months_count > 1:
        period_desc += ' (' + str(months_count) + ' months)'

    row = ['1', Paragraph(period_desc + '<br/>Period: ' + billing_period, desc_style),
           amounts['base_amount'], amounts['base_amount']]
    if is_gst_applicable:
        row += ([amounts['cgst'], amounts['sgst']] if gst_details['is_same_state'] else [amounts['igst']])
    row.append(amounts['subtotal'])

    if is_gst_applicable:
        col_ratios = [0.07, 0.33, 0.12, 0.13, 0.12, 0.12, 0.11] if gst_details['is_same_state'] \
               else [0.07, 0.38, 0.13, 0.14, 0.17, 0.11]
    else:
        col_ratios = [0.07, 0.43, 0.17, 0.18, 0.15]

    col_widths  = [PAGE_W * r for r in col_ratios[:len(headers)]]
    items_table = Table([headers, row], colWidths=col_widths, repeatRows=1)
    items_table.setStyle(TableStyle([
        ('GRID',       (0,0), (-1,-1), 0.7, colors.grey),
        ('BACKGROUND', (0,0), (-1, 0), colors.HexColor('#2f5f7a')),
        ('TEXTCOLOR',  (0,0), (-1, 0), colors.white),
        ('FONTNAME',   (0,0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('ALIGN',      (2,0), (-1,-1), 'RIGHT'),
    ]))
    elements += [items_table, Spacer(1, 4)]

    # Summary rows
    outstanding_val = Decimal(str(amounts.get('outstanding_charge', '0.00')))
    advance_val     = Decimal(str(amounts.get('advance_used',       '0.00')))
    summary_rows    = []

    summary_rows.append(['Subtotal (incl. GST):' if is_gst_applicable else 'Subtotal:', 'Rs. ' + amounts['subtotal']])
    if outstanding_val > 0:
        summary_rows.append(['Outstanding Balance (no GST):', 'Rs. ' + amounts['outstanding_charge']])
    if advance_val > 0:
        summary_rows.append(['Advance Adjusted:', '- Rs. ' + amounts['advance_used']])
    summary_rows.append(['Balance Due:', 'Rs. ' + amounts['total_amount']])

    num_cols      = len(headers)
    padded        = [[''] * (num_cols - 2) + [lbl, val] for lbl, val in summary_rows]
    summary_table = Table(padded, colWidths=col_widths)

    sstyle = [
        ('ALIGN',       (-2,0), (-1,-1), 'RIGHT'),
        ('FONTSIZE',    ( 0,0), (-1,-1), 8),
        ('TOPPADDING',  ( 0,0), (-1,-1), 3),
        ('BOTTOMPADDING',( 0,0),(-1,-1), 3),
    ]
    if outstanding_val > 0:
        sstyle += [
            ('TEXTCOLOR', (-2,1), (-1,1), colors.HexColor('#c0392b')),
            ('FONTNAME',  (-2,1), (-1,1), 'Helvetica-Bold'),
        ]
    if advance_val > 0:
        adv_idx = 2 if outstanding_val > 0 else 1
        sstyle += [('TEXTCOLOR', (-2, adv_idx), (-1, adv_idx), colors.HexColor('#27ae60'))]

    last = len(summary_rows) - 1
    sstyle += [
        ('FONTNAME',  (-2, last), (-1, last), 'Helvetica-Bold'),
        ('LINEABOVE', (-2, last), (-1, last), 1.5, colors.HexColor('#2f5f7a')),
        ('FONTSIZE',  (-2, last), (-1, last), 9),
    ]
    summary_table.setStyle(TableStyle(sstyle))
    elements += [summary_table, Spacer(1, 12)]

    # Bank details + amount in words
    bank_lines = []
    if company.bank_name:     bank_lines.append('<b>Bank:</b> ' + company.bank_name)
    if company.account_number:
        bank_lines.append('<b>Account Holder:</b> ' + company.name)
        bank_lines.append('<b>A/C:</b> ' + company.account_number)
    if company.ifsc_code:     bank_lines.append('<b>IFSC:</b> ' + company.ifsc_code)
    if company.branch_name:   bank_lines.append('<b>Branch:</b> ' + company.branch_name)

    bank_block   = Paragraph('<br/>'.join(bank_lines) if bank_lines else 'Bank details not available', small_left)
    amount_words = Paragraph('<b>Amount in Words</b><br/>' + number_to_words(float(final_amount)), small_left)
    elements.append(Table([[bank_block, amount_words]], colWidths=[PAGE_W*0.55, PAGE_W*0.45]))
    elements.append(Spacer(1, 10))

    # QR + Payment terms
    qr_img = ''
    if getattr(company, 'qr_code', None):
        try:
            if hasattr(company.qr_code, 'path') and os.path.exists(company.qr_code.path):
                qr_img = Image(company.qr_code.path, 1.6*inch, 1.6*inch, kind='proportional')
        except Exception as e:
            print('QR error: ' + str(e))

    payment_terms = Paragraph(
        '<b>Payment Terms</b><br/>'
        'Payment due within 15 days<br/>'
        'Late payment charges applicable<br/>'
        'Mention Invoice Number during payment',
        ParagraphStyle('pt', parent=small_left, alignment=TA_RIGHT)
    )

    if qr_img:
        elements.append(KeepTogether([Table([[qr_img, payment_terms]], colWidths=[PAGE_W*0.35, PAGE_W*0.65])]))
    else:
        elements.append(payment_terms)

    elements.append(Spacer(1, 16))
    elements.append(Paragraph(
        '<b>For ' + company.name + '</b><br/><br/>Authorized Signatory',
        ParagraphStyle('sign', alignment=TA_RIGHT, fontSize=10)
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer


# =============================================================================
# PDF — LOCATION BILLING
# =============================================================================

def generate_invoice_pdf_locations(billing_record, subscription_plan, entity, gst_details, amounts,
                                   organization, billing_period, proration_info, entity_type,
                                   advance_adjustment=None, outstanding_amount=None):
    """Generate PDF invoice with line items. GST columns omitted for non-GST locations."""
    buffer = BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=A4,
                               rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    PAGE_W = A4[0] - doc.leftMargin - doc.rightMargin
    styles = getSampleStyleSheet()

    small_left  = ParagraphStyle('small_left',  parent=styles['Normal'], fontSize=9,  leading=12, alignment=TA_LEFT)
    small_right = ParagraphStyle('small_right', parent=styles['Normal'], fontSize=9,  leading=12, alignment=TA_RIGHT)
    desc_style  = ParagraphStyle('desc',        parent=styles['Normal'], fontSize=8,  leading=10, alignment=TA_LEFT)

    elements          = []
    is_gst_applicable = gst_details.get('is_gst_applicable', True)

    # Logo
    logo = None
    try:
        if hasattr(organization, 'logo') and organization.logo.name:
            logo_path = organization.logo.path
            if os.path.exists(logo_path):
                logo = Image(logo_path, width=1.2*inch, height=1.2*inch, kind='proportional')
    except Exception as e:
        print('Logo error: ' + str(e))

    # Header
    invoice_title = 'TAX INVOICE' if is_gst_applicable else 'INVOICE'
    org_name_block = Paragraph(
        '<font size="16"><b>' + organization.name + '</b></font><br/>'
        '<font size="9">'
        + getattr(organization, 'address', '') + '<br/>'
        'Contact: ' + getattr(organization, 'contact', '') + '<br/>'
        'GSTIN: '   + getattr(organization, 'gst_number', '') + '</font>',
        small_left
    )
    invoice_block = Paragraph(
        '<b>' + invoice_title + '</b><br/>'
        '<font size="9">Original for Recipient</font><br/><br/>'
        '<b>IN-' + str(billing_record.id) + '</b>',
        ParagraphStyle('hdr', fontSize=14, alignment=TA_RIGHT, fontName='Helvetica-Bold')
    )
    header = Table([[logo, org_name_block, invoice_block]],
                   colWidths=[1.5*inch, PAGE_W*0.45, PAGE_W*0.30])
    header.setStyle(TableStyle([
        ('VALIGN',       (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING',  (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    elements += [header, Spacer(1, 10)]

    # Amount Due bar
    final_amount = amounts.get('total_amount', '0.00')
    amount_due   = Table([[
        Paragraph('<b>Amount Due</b>', small_left),
        Paragraph('<b>Rs. ' + final_amount + '</b>',
                  ParagraphStyle('amt', fontSize=14, alignment=TA_RIGHT,
                                 textColor=colors.white, fontName='Helvetica-Bold'))
    ]], colWidths=[PAGE_W*0.55, PAGE_W*0.45])
    amount_due.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), colors.HexColor('#2f5f7a')),
        ('TEXTCOLOR',     (0,0), (-1,-1), colors.white),
        ('LEFTPADDING',   (0,0), (-1,-1), 10),
        ('RIGHTPADDING',  (0,0), (-1,-1), 10),
        ('TOPPADDING',    (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    elements += [amount_due, Spacer(1, 12)]

    # Org info + dates
    org_block  = Paragraph(
        '<b>' + organization.name + '</b><br/>'
        + getattr(organization, 'address', '') + '<br/>'
        'GSTIN: ' + getattr(organization, 'gst_number', ''),
        small_left
    )
    date_block = Paragraph(
        'Issue Date: ' + billing_record.billing_date.strftime('%d-%b-%Y') + '<br/>'
        'Due Date: '   + billing_record.due_date.strftime('%d-%b-%Y'),
        small_right
    )
    org_info = Table([[org_block, date_block]], colWidths=[PAGE_W*0.62, PAGE_W*0.38])
    org_info.setStyle(TableStyle([('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0)]))
    elements += [org_info, Spacer(1, 10)]

    # Client details
    if entity_type == 'Customer_Locations':
        client_name = entity.customer.name + ' - ' + entity.location_name
    elif entity_type == 'Company_Locations':
        client_name = entity.company.name  + ' - ' + entity.location_name
    else:
        client_name = entity.location_name

    client_name_block = Paragraph(
        '<b>Client Details</b><br/>'
        '<b>' + client_name + '</b><br/>'
        'Email: ' + getattr(entity, 'location_email', ''),
        small_left
    )
    billing_address_block = Paragraph(
        '<b>Billing Address</b><br/>'
        + str(entity.address) + '<br/>'
        + str(entity.city) + ', ' + str(entity.state) + '<br/>'
        + str(entity.pincode),
        small_right
    )
    client_row = Table([[client_name_block, billing_address_block]], colWidths=[PAGE_W*0.5, PAGE_W*0.5])
    client_row.setStyle(TableStyle([
        ('VALIGN',       (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING',  (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('ALIGN',        (1,0),  (1, 0), 'RIGHT'),
    ]))
    elements += [client_row, Spacer(1, 12)]

    # Billing table with line items
    headers = ['S.No', 'Item Description', 'Price', 'Taxable']
    if is_gst_applicable:
        headers += (['CGST @ 9%', 'SGST @ 9%'] if gst_details['is_same_state'] else ['IGST @ 18%'])
    headers.append('Amount')

    line_items = billing_record.bill_items.all()
    table_data = [headers]

    for idx, item in enumerate(line_items, 1):
        item_desc    = item.item_name
        if item.description:
            item_desc += "<br/><font size='7'>" + item.description + '</font>'
        taxable      = Decimal(str(item.total_price))

        if is_gst_applicable and item.tax_percentage > 0:
            if gst_details['is_same_state']:
                c = (taxable * Decimal('0.09')).quantize(Decimal('0.01'))
                s = (taxable * Decimal('0.09')).quantize(Decimal('0.01'))
                row = [str(idx), Paragraph(item_desc, desc_style),
                       f'{taxable:.2f}', f'{taxable:.2f}',
                       f'{c:.2f}', f'{s:.2f}', f'{taxable+c+s:.2f}']
            else:
                g = (taxable * Decimal('0.18')).quantize(Decimal('0.01'))
                row = [str(idx), Paragraph(item_desc, desc_style),
                       f'{taxable:.2f}', f'{taxable:.2f}', f'{g:.2f}', f'{taxable+g:.2f}']

        elif is_gst_applicable and item.tax_percentage == 0:
            # GST customer but this item is tax-free (e.g. outstanding)
            if gst_details['is_same_state']:
                row = [str(idx), Paragraph(item_desc, desc_style),
                       f'{taxable:.2f}', '0.00', '0.00', '0.00', f'{taxable:.2f}']
            else:
                row = [str(idx), Paragraph(item_desc, desc_style),
                       f'{taxable:.2f}', '0.00', '0.00', f'{taxable:.2f}']

        else:
            # Non-GST location: no tax columns
            row = [str(idx), Paragraph(item_desc, desc_style),
                   f'{taxable:.2f}', f'{taxable:.2f}', f'{taxable:.2f}']

        table_data.append(row)

    if is_gst_applicable:
        col_ratios = [0.07, 0.33, 0.12, 0.13, 0.12, 0.12, 0.11] if gst_details['is_same_state'] \
               else [0.07, 0.38, 0.13, 0.14, 0.17, 0.11]
    else:
        col_ratios = [0.07, 0.43, 0.17, 0.18, 0.15]

    col_widths  = [PAGE_W * r for r in col_ratios[:len(headers)]]
    items_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    items_table.setStyle(TableStyle([
        ('GRID',       (0,0), (-1,-1), 0.7, colors.grey),
        ('BACKGROUND', (0,0), (-1, 0), colors.HexColor('#2f5f7a')),
        ('TEXTCOLOR',  (0,0), (-1, 0), colors.white),
        ('FONTNAME',   (0,0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('ALIGN',      (2,0), (-1,-1), 'RIGHT'),
        ('VALIGN',     (0,0), (-1,-1), 'TOP'),
    ]))
    elements += [items_table, Spacer(1, 8)]

    # Advance adjustment
    if advance_adjustment and Decimal(str(advance_adjustment.get('advance_used', 0))) > 0:
        adv_data = [
            [''] * (len(headers)-2) + ['Subtotal:',       'Rs. ' + amounts.get('subtotal',      '0.00')],
            [''] * (len(headers)-2) + ['Advance Adjusted:', '- Rs. ' + amounts.get('advance_used', '0.00')],
            [''] * (len(headers)-2) + ['Balance Due:',    'Rs. ' + amounts.get('total_amount',  '0.00')],
        ]
        adv_table = Table(adv_data, colWidths=col_widths)
        adv_table.setStyle(TableStyle([
            ('ALIGN',    (-2,0), (-1,-1), 'RIGHT'),
            ('FONTNAME', (-2,-1),(-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', ( 0, 0),(-1,-1), 8),
            ('LINEABOVE',(-2,-1),(-1,-1), 1.5, colors.HexColor('#2f5f7a')),
            ('TEXTCOLOR',(-2, 1),(-1, 1), colors.green),
        ]))
        elements += [adv_table, Spacer(1, 12)]

    # Bank details + amount in words
    bank_block   = Paragraph(
        '<b>Account Holder:</b> V K CONTROL SYSTEM PRIVATE LIMITED<br/>'
        '<b>Bank:</b> ICICI Bank Ltd.<br/>'
        '<b>A/C:</b> 696905600545<br/>'
        '<b>IFSC:</b> ICIC0006969',
        small_left
    )
    amount_words = Paragraph(
        '<b>Amount in Words</b><br/>' + number_to_words_loc(float(final_amount)),
        small_left
    )
    elements.append(Table([[bank_block, amount_words]], colWidths=[PAGE_W*0.55, PAGE_W*0.45]))
    elements.append(Spacer(1, 10))

    # QR + Payment terms
    qr_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'sachine_qr.jpg')
    qr_img  = Image(qr_path, 1.6*inch, 1.6*inch) if os.path.exists(qr_path) else ''

    payment_terms = Paragraph(
        '<b>Payment Terms</b><br/>'
        'Payment due within 15 days<br/>'
        'Late payment charges applicable<br/>'
        'Mention Invoice Number during payment',
        ParagraphStyle('pt', parent=small_left, alignment=TA_RIGHT)
    )

    if qr_img:
        elements.append(KeepTogether([Table([[qr_img, payment_terms]], colWidths=[PAGE_W*0.35, PAGE_W*0.65])]))
    else:
        elements.append(payment_terms)

    elements.append(Spacer(1, 16))
    elements.append(Paragraph(
        '<b>For ' + organization.name + '</b><br/><br/>Authorized Signatory',
        ParagraphStyle('sign', alignment=TA_RIGHT, fontSize=10)
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer
def update_customer_location(request, location_id):
    """
    Update customer location with automatic subscription change tracking.
    Includes GST registration fields.
    """
    username = request.session.get('username')
    user = Custom_User.objects.get(username=username)
    if not user:
        return redirect('login')

    if request.method == 'POST':
        try:
            location = get_object_or_404(CustomerLocation, id=location_id)

            # ── Snapshot old values ───────────────────────────────────────
            old_location_name     = location.location_name
            old_subscription_plan = location.subscription_plan
            old_custom_amount     = location.custom_subscription_amount
            old_status            = "Active" if location.is_active else "Inactive"
            old_gst_registered    = location.gst_registered
            old_gst_number        = location.gst_number or ""

            # ── Subscription plan ─────────────────────────────────────────
            subscription_plan_id  = request.POST.get('subscription_plan')
            new_subscription_plan = None
            if subscription_plan_id:
                new_subscription_plan = SubscriptionPlan.objects.get(id=subscription_plan_id)

            # ── Custom subscription amount ────────────────────────────────
            new_custom_amount = request.POST.get('custom_subscription_amount', '').strip()
            if new_custom_amount:
                try:
                    new_custom_amount = Decimal(new_custom_amount)
                except (ValueError, InvalidOperation):
                    new_custom_amount = None
            else:
                new_custom_amount = None

            # ── GST fields ────────────────────────────────────────────────
            new_gst_registered = request.POST.get("gst_registered") == "1"
            new_gst_number     = request.POST.get("gst_number", "").strip() or None

            # Clear GST number if toggle is off
            if not new_gst_registered:
                new_gst_number = None
            # ─────────────────────────────────────────────────────────────

            # ── Effective amounts for change detection ────────────────────
            old_amount = (
                old_custom_amount if old_custom_amount
                else (old_subscription_plan.price if old_subscription_plan else Decimal('0.00'))
            )
            new_amount = (
                new_custom_amount if new_custom_amount
                else (new_subscription_plan.price if new_subscription_plan else Decimal('0.00'))
            )

            plan_changed   = False
            amount_changed = False
            is_upgrade     = False
            upgrade_amount = Decimal('0.00')

            if old_subscription_plan and new_subscription_plan:
                if old_subscription_plan.id != new_subscription_plan.id:
                    plan_changed = True
                if new_amount != old_amount:
                    amount_changed = True
                if new_amount > old_amount:
                    is_upgrade     = True
                    upgrade_amount = new_amount - old_amount

            # ── Apply field updates ───────────────────────────────────────
            new_location_name = request.POST.get('customer_name', location.location_name)
            status            = request.POST.get('status')
            new_status        = "Active" if status == 'Active' else "Inactive"

            location.location_name              = new_location_name
            location.address                    = request.POST.get('address',  location.address)
            location.area_id                    = request.POST.get('area',     location.area_id)
            location.city_id                    = request.POST.get('city',     location.city_id)
            location.state_id                   = request.POST.get('state',    location.state_id)
            location.pincode                    = request.POST.get('pincode',  location.pincode)
            location.location_contact           = request.POST.get('contact',  location.location_contact)
            location.location_email             = request.POST.get('email',    location.location_email)
            location.remarks                    = request.POST.get('remarks',  location.remarks)
            location.is_active                  = (status == 'Active')
            location.subscription_plan          = new_subscription_plan
            location.custom_subscription_amount = new_custom_amount
            location.start_date                 = request.POST.get('start_date', location.start_date)
            location.location_link              = request.POST.get('location_link') or None

            # ── GST fields ────────────────────────────────────────────────
            location.gst_registered = new_gst_registered
            location.gst_number     = new_gst_number
            # ─────────────────────────────────────────────────────────────

            # save() triggers SubscriptionChange creation if plan/amount changed
            location.save()

            # ── Build change summary for user_logs ────────────────────────
            changes = []

            if old_location_name != new_location_name:
                changes.append(
                    "Name: " + old_location_name + " to " + new_location_name
                )
            if plan_changed:
                changes.append(
                    "Plan: "
                    + (old_subscription_plan.name if old_subscription_plan else "None")
                    + " to "
                    + (new_subscription_plan.name if new_subscription_plan else "None")
                )
            if amount_changed and not plan_changed:
                changes.append(
                    "Amount: Rs." + str(old_amount) + " to Rs." + str(new_amount)
                )
            if old_status != new_status:
                changes.append("Status: " + old_status + " to " + new_status)

            # GST change tracking
            if old_gst_registered != new_gst_registered:
                changes.append(
                    "GST: "
                    + ("Not Registered" if old_gst_registered else "Registered")
                    + " to "
                    + ("Registered" if new_gst_registered else "Not Registered")
                )
            elif new_gst_registered and old_gst_number != (new_gst_number or ""):
                changes.append(
                    "GST Number: " + (old_gst_number or "None")
                    + " to " + (new_gst_number or "None")
                )

            # ── Audit log (user_logs) ─────────────────────────────────────
            log_details = (
                "Location: " + old_location_name
                + " under Customer: " + location.customer.name
                + (" | Changes: " + ", ".join(changes) if changes else "")
            )
            user_logs.objects.create(
                user=user.username,
                action='Updated Customer Location',
                details=log_details,
            )

            # ── data_logs for subscription change ─────────────────────────
            if plan_changed or amount_changed:
                change_type   = "Upgrade" if is_upgrade else "Change"
                old_plan_name = old_subscription_plan.name if old_subscription_plan else "None"
                new_plan_name = new_subscription_plan.name if new_subscription_plan else "None"

                details = (
                    "LOCATION SUBSCRIPTION " + change_type.upper() + "\n\n"
                    "Customer: " + location.customer.name + "\n"
                    "Location: " + location.location_name + "\n"
                    "Date: " + timezone.now().strftime('%d-%m-%Y %H:%M:%S') + "\n\n"
                    "Old Subscription:\n"
                    "  Plan: " + old_plan_name + "\n"
                    "  Amount: Rs." + str(old_amount) + " per month\n\n"
                    "New Subscription:\n"
                    "  Plan: " + new_plan_name + "\n"
                    "  Amount: Rs." + str(new_amount) + " per month\n\n"
                    "Change Details:\n"
                    "  " + ("Amount Increase" if is_upgrade else "Amount Change")
                    + ": Rs." + str(abs(new_amount - old_amount)) + " per month\n"
                    "  Effective From: " + timezone.now().date().strftime('%d-%m-%Y') + "\n\n"
                    "Notes:\n"
                    "- This change will be reflected in the next billing cycle\n"
                    "- Pro-rated billing will be applied for the current month\n"
                    "- Previous unbilled days will use old plan rate\n"
                    "- Remaining days in month will use new plan rate"
                )

                data_logs.objects.create(
                    user=user,
                    timestamp=timezone.now(),
                    customer=location.customer,
                    location=location,
                    action=(
                        "Location Service " + change_type
                        + ": " + old_plan_name + " to " + new_plan_name
                    ),
                    details=details,
                    is_upgrade=is_upgrade,
                    old_subscription_plan=old_subscription_plan,
                    new_subscription_plan=new_subscription_plan,
                    upgrade_amount=upgrade_amount if is_upgrade else (new_amount - old_amount),
                    status=change_type,
                    billing_period_start=timezone.now().date(),
                    billing_period_end=None,
                )

                if is_upgrade:
                    messages.success(
                        request,
                        'Location "' + old_location_name + '" upgraded successfully from '
                        + old_plan_name + ' (Rs.' + str(old_amount) + '/month) to '
                        + new_plan_name + ' (Rs.' + str(new_amount) + '/month). '
                        'Next bill will be pro-rated for the change.',
                    )
                else:
                    messages.success(
                        request,
                        'Location "' + old_location_name + '" subscription changed from '
                        + old_plan_name + ' to ' + new_plan_name
                        + '. Next bill will reflect the pro-rated change.',
                    )

            else:
                # Regular update — no subscription change
                data_logs.objects.create(
                    user=user,
                    timestamp=timezone.now(),
                    customer=location.customer,
                    location=location,
                    action="Location Updated: " + old_location_name,
                    details="Location details updated by " + user.username,
                    status='Updated',
                )
                messages.success(
                    request,
                    'Location "' + old_location_name + '" updated successfully!',
                )

            return redirect('customer_locations', customer_id=location.customer.id)

        except Exception as e:
            messages.error(request, 'Error updating location: ' + str(e))
            import traceback
            traceback.print_exc()
            return redirect('customer_locations', customer_id=location.customer.id)

    # ── GET — render edit form ────────────────────────────────────────────
    else:
        location           = get_object_or_404(CustomerLocation, id=location_id)
        subscription_plans = SubscriptionPlan.objects.all()
        states             = State.objects.all()

        context = {
            'location':           location,
            'subscription_plans': subscription_plans,
            'states':             states,
        }
        return render(request, 'update_location.html', context)

def delete_customer_location(request, location_id):
    username=request.session.get('username')
    user=Custom_User.objects.get(username=username)
    if not user:
        return redirect('login')
    if request.method == 'POST':
        try:
            location = get_object_or_404(CustomerLocation, id=location_id)
            location_name = location.location_name
            customer_name = location.customer.name
            customer_id = location.customer.id
            
            # Log before deletion
            user_logs.objects.create(
                user=user.username,
                action='Deleted Customer Location',
                details=f'Location: {location_name} deleted under Customer: {customer_name}'
            )
            
            location.delete()
            
            messages.success(request, f'Location "{location_name}" deleted successfully!')
            return redirect('customer_locations', customer_id=customer_id)
        except Exception as e:
            messages.error(request, f'Error deleting location: {str(e)}')
            return redirect('customer_locations', customer_id=customer_id)
    
    return redirect('customer_locations', customer_id=customer_id)





def company_locations(request, company_id):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')

    company = get_object_or_404(Company, id=company_id)

    if request.method == "POST":
        location_name = request.POST.get("customer_name")
        address = request.POST.get("address")
        area_id = request.POST.get("area")
        city_id = request.POST.get("city")
        state_id = request.POST.get("state")
        pincode = request.POST.get("pincode")
        subscription_plan_id = request.POST.get("subscription_plan")
        start_date = request.POST.get("start_date")
        location_contact = request.POST.get("contact")
        location_email = request.POST.get("email")
        location_link= request.POST.get("location_link")
        processing_fee = request.POST.get("processing_fee")
        is_active = request.POST.get("is_active") == "on"

        company_location=CompanyLocation.objects.create(
            company=company,
            location_name=location_name,
            address=address,
            area_id=area_id or None,
            city_id=city_id or None,
            state_id=state_id or None,
            pincode=pincode,
            subscription_plan_id=subscription_plan_id or None,
            start_date=start_date or None,
            location_contact=location_contact,
            location_email=location_email,
            is_active=is_active,
            location_link=location_link or None,
            processing_fee=processing_fee or None,
        )
      

        return redirect('company_locations', company_id=company.id)

    # GET request
    locations = CompanyLocation.objects.filter(company=company)
    states = State.objects.all()
    cities = City.objects.all()
    areas = Area.objects.all()
    plans = SubscriptionPlan.objects.all()

    context = {
        "company": company,
        "locations": locations,
        "states": states,
        "cities": cities,
        "areas": areas,
        "subscription_plans": plans,
        'user': user,
        'role_permissions': role_permissions,
    }

    return render(request, "company_locations.html", context)



def update_company_location(request, location_id):
    username = request.session.get('username')
    user = Custom_User.objects.get(username=username)
    if not user:
        return redirect('login')

    if request.method == 'POST':
        try:
            location = get_object_or_404(CompanyLocation, id=location_id)

            location.location_name = request.POST.get('customer_name', location.location_name)
            location.address = request.POST.get('address', location.address)
            location.area_id = request.POST.get('area') or location.area_id
            location.city_id = request.POST.get('city') or location.city_id
            location.state_id = request.POST.get('state') or location.state_id
            location.pincode = request.POST.get('pincode', location.pincode)

            location.subscription_plan_id = request.POST.get('subscription_plan') or None
            location.start_date = request.POST.get('start_date') or location.start_date

            location.location_contact = request.POST.get('contact', location.location_contact)
            location.location_email = request.POST.get('email', location.location_email)
            location.location_contact_person = request.POST.get(
                'location_contact_person', 
                location.location_contact_person
            )
            location.location_link= request.POST.get('location_link') or None   
            location.is_active = request.POST.get('status') == 'Active'
            processing_fee = request.POST.get('processing_fee')
            location.processing_fee = processing_fee if processing_fee else location.processing_fee
            location.save()

            messages.success(request, f'Location "{location.location_name}" updated successfully!')
            return redirect('company_locations', company_id=location.company.id)

        except Exception as e:
            messages.error(request, f'Error updating location: {str(e)}')
            return redirect('company_locations', company_id=location.company.id)


def delete_company_location(request, location_id):
    
    username = request.session.get('username')
    user = Custom_User.objects.get(username=username)
    if not user:
        return redirect('login')

    if request.method == 'POST':
        try:
            location = get_object_or_404(CompanyLocation, id=location_id)
            name = location.location_name
            company_id = location.company.id
            location.delete()

            messages.success(request, f'Location "{name}" deleted successfully!')
            return redirect('company_locations', company_id=company_id)

        except Exception as e:
            messages.error(request, f'Error deleting location: {str(e)}')
            return redirect('company_locations', company_id=company_id)

    return redirect('company_locations', company_id=company_id)



def settings_view(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    
    state_count = State.objects.count()
    city_count = City.objects.count()
    area_count = Area.objects.count()
    context = {'state_count': state_count, 
               'city_count': city_count, 
               'area_count': area_count,  
                **get_company_filter_context(request),
                'user': user,
                'role_permissions': role_permissions,
    }
    return render(request, 'settings.html', context)

def organization_details(request):
    username = request.session.get('username')
    user = Custom_User.objects.get(username=username)

    if not user:
        return redirect('login')

    # Try to get the first organization, or create a new one if none exists
    organization = Organization.objects.first()

    if organization is None:
        # If no organization exists, create a new one
        organization = Organization(name='', address='', contact='', email='', gst_number='', logo=None, cin_number='')
        organization.save()

    if request.method == 'POST':
        # If the request method is POST, we update the organization details
        # Get data from the request and save the organization
        
        organization.name = request.POST.get('companyName', '')
        organization.gst_number = request.POST.get('gstNumber', '')
        organization.address = request.POST.get('address', '')
        organization.contact = request.POST.get('phoneNumber', '')
        organization.email = request.POST.get('email', '')
        organization.cin_number = request.POST.get('cinNumber', '')
        
        # Handle logo upload if file is provided
        if 'logoFile' in request.FILES:
            organization.logo = request.FILES['logoFile']
        
        # Save the organization details
        organization.save()

        # Redirect after saving
        return redirect('organization_management')  # Redirect to the same page or to another page after saving

    context = {
        'organization': organization
    }

    return render(request, 'companydetails.html', context)



from datetime import datetime



from django.shortcuts import render, redirect
from datetime import datetime




from datetime import datetime, timedelta
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Sum
from django.db import models
from django.core.mail import EmailMessage
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io
from .models import Custom_User, Customer, Company, CustomerLocation, data_logs
from .utils import get_filtered_companies, get_company_filter_context


def send_email_report(request):
    """
    Send payment report as Excel attachment via email.
    """
    try:
        entity = request.POST.get('entity')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        
        if not entity or not entity.startswith('customer_'):
            return JsonResponse({'success': False, 'message': 'Invalid customer selected'}, status=400)
        
        customer_id = entity.replace('customer_', '')
        
        try:
            customer = Customer.objects.select_related('company').get(id=customer_id)
        except Customer.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Customer not found'}, status=404)
        
        # Get customer email
        customer_email = customer.email
        if not customer_email:
            return JsonResponse({'success': False, 'message': 'Customer has no email address'}, status=400)
        
        # Auto-set dates if not provided
        if not start_date_str and customer.start_date:
            start_date_str = customer.start_date.strftime('%Y-%m-%d')
        
        if not end_date_str:
            from datetime import date
            end_date_str = date.today().strftime('%Y-%m-%d')
        
        # Get data logs for this customer
        data_logs_query = data_logs.objects.filter(
            Q(customer_id=customer_id, location__isnull=True) |
            Q(location__customer_id=customer_id)
        ).select_related('customer', 'location', 'user', 'billing_record')
        
        # Apply date filters
        if start_date_str:
            start_date_obj = datetime.strptime(start_date_str, '%Y-%m-%d')
            data_logs_query = data_logs_query.filter(timestamp__gte=start_date_obj)
        
        if end_date_str:
            end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1)
            data_logs_query = data_logs_query.filter(timestamp__lt=end_date_obj)
        
        data_logs_list = data_logs_query.order_by('-timestamp')
        
        if not data_logs_list.exists():
            return JsonResponse({'success': False, 'message': 'No data found for this customer'}, status=400)
        
        # Calculate total outstanding
        total_outstanding = data_logs_query.aggregate(
            total=Sum('balance_amount')
        )['total'] or 0
        
        # Create Excel file in memory
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payment Report"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:L1')
        title_cell = ws['A1']
        title_cell.value = f"Payment Report - {customer.name}"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Customer Info
        ws.merge_cells('A2:L2')
        info_cell = ws['A2']
        company_name = customer.company.name if customer.company else 'N/A'
        info_cell.value = f"Company: {company_name} | Period: {start_date_str} to {end_date_str}"
        info_cell.alignment = Alignment(horizontal="center")
        
        # Outstanding Balance
        ws.merge_cells('A3:L3')
        balance_cell = ws['A3']
        balance_cell.value = f"Total Outstanding Balance: ₹{total_outstanding:.2f}"
        balance_cell.font = Font(bold=True, size=12, color="FF0000" if total_outstanding > 0 else "00AA00")
        balance_cell.alignment = Alignment(horizontal="center")
        
        # Headers (Row 5)
        headers = [
            'Date & Time',
            'Action',
            'Location',
            'Billing Period',
            'Payment Amount',
            'Total Paid',
            'Balance',
            'Payment Mode',
            'Payment Date',
            'Status',
            'User',
            'Notes'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Data rows
        row_num = 6
        for log in data_logs_list:
            billing_period = f"{log.billing_period_start.strftime('%d/%m/%Y') if log.billing_period_start else 'N/A'} - {log.billing_period_end.strftime('%d/%m/%Y') if log.billing_period_end else 'N/A'}"
            
            ws.cell(row=row_num, column=1, value=log.timestamp.strftime('%Y-%m-%d %H:%M:%S')).border = border
            ws.cell(row=row_num, column=2, value=log.action or 'N/A').border = border
            ws.cell(row=row_num, column=3, value=log.location.location_name if log.location else 'Main').border = border
            ws.cell(row=row_num, column=4, value=billing_period).border = border
            ws.cell(row=row_num, column=5, value=float(log.payment_amount) if log.payment_amount else 0).border = border
            ws.cell(row=row_num, column=6, value=float(log.total_paid) if log.total_paid else 0).border = border
            ws.cell(row=row_num, column=7, value=float(log.balance_amount) if log.balance_amount else 0).border = border
            ws.cell(row=row_num, column=8, value=log.payment_mode or 'N/A').border = border
            ws.cell(row=row_num, column=9, value=log.payment_date.strftime('%d/%m/%Y') if log.payment_date else 'N/A').border = border
            ws.cell(row=row_num, column=10, value=log.status or 'N/A').border = border
            ws.cell(row=row_num, column=11, value=log.user.name if (log.user and log.user.name) else (log.user.username if log.user else 'N/A')).border = border
            ws.cell(row=row_num, column=12, value=log.details[:100] if log.details else 'N/A').border = border
            
            row_num += 1
        
        # Adjust column widths
        column_widths = {
            'A': 20, 'B': 30, 'C': 20, 'D': 25, 'E': 15,
            'F': 15, 'G': 15, 'H': 15, 'I': 15, 'J': 15,
            'K': 20, 'L': 30
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create email
        subject = f"Payment Report - {customer.name}"
        body = f"""
Dear {customer.name},

Please find attached your payment report for the period {start_date_str} to {end_date_str}.

Report Summary:
- Total Outstanding Balance: ₹{total_outstanding:.2f}
- Number of Records: {data_logs_list.count()}

If you have any questions, please don't hesitate to contact us.

Best regards,
{settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'Billing Team'}
        """.strip()
        
        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'noreply@example.com',
            to=[customer_email],
        )
        
        # Attach Excel file
        filename = f"payment_report_{customer.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        email.attach(filename, excel_file.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Send email
        try:
            email.send(fail_silently=False)
        except Exception as email_error:
            import traceback
            traceback.print_exc()
            raise  # Re-raise to be caught by outer try-except
        
        return JsonResponse({
            'success': True,
            'message': f'Report sent successfully to {customer_email}'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': f'Failed to send email: {str(e)}'
        }, status=500)
def reports(request):
    
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')

    try:
        user = Custom_User.objects.get(username=user)
    except Custom_User.DoesNotExist:
        return redirect('login')
    
    if request.method == 'POST' and request.POST.get('email_report') == 'true':
        return send_email_report(request)

    filtered_companies = get_filtered_companies(request)
    company_ids = list(filtered_companies.values_list('id', flat=True))
    
    selected_gst_type = request.session.get('selected_gst_type', 'gst')
    selected_company_ids = request.session.get('selected_company_ids', [])
    filters_active = selected_gst_type or selected_company_ids
    
    if company_ids:
        customers = Customer.objects.filter(
            company_id__in=company_ids
        ).prefetch_related(
            'contacts', 'locations', 'locations__area',
            'locations__city', 'locations__state'
        ).select_related('company').order_by('name')
        customer_locations = CustomerLocation.objects.filter(
            customer__company_id__in=company_ids
        ).select_related('customer', 'area', 'city', 'state').order_by('location_name')
    elif filters_active:
        customers = Customer.objects.none()
        customer_locations = CustomerLocation.objects.none()
    else:
        customers = Customer.objects.prefetch_related(
            'contacts', 'locations', 'locations__area',
            'locations__city', 'locations__state'
        ).select_related('company').all()
        customer_locations = CustomerLocation.objects.select_related(
            'customer', 'area', 'city', 'state'
        ).all()
        
    companies = filtered_companies.order_by('name')
    
    entity           = request.GET.get('entity')
    location_id      = request.GET.get('location')
    start_date_param = request.GET.get('start_date')
    end_date_param   = request.GET.get('end_date')
    export_excel     = request.GET.get('export_excel')
    
    start_date        = start_date_param
    end_date          = end_date_param
    selected_customer = None
    
    if entity and entity.startswith('customer_'):
        customer_id = entity.replace('customer_', '')
        try:
            selected_customer = Customer.objects.get(id=customer_id)
            if not start_date_param and selected_customer.start_date:
                start_date = selected_customer.start_date.strftime('%Y-%m-%d')
            if not end_date_param:
                from datetime import date
                end_date = date.today().strftime('%Y-%m-%d')
        except Customer.DoesNotExist:
            pass

    filtered_data_logs      = None
    customer_ledger_balance = None
    customer_advance_amount = None

    if entity or location_id or start_date or end_date:
        
        if company_ids:
            data_logs_query = data_logs.objects.filter(
                customer__company_id__in=company_ids
            ).select_related(
                'customer', 'customer__company', 'location',
                'location__customer', 'user', 'billing_record'
            ).prefetch_related('billing_record__bill_items')
        elif filters_active:
            data_logs_query = data_logs.objects.none()
        else:
            data_logs_query = data_logs.objects.all().select_related(
                'customer', 'customer__company', 'location',
                'location__customer', 'user', 'billing_record'
            ).prefetch_related('billing_record__bill_items')
        
        if entity and entity.startswith('customer_'):
            customer_id = entity.replace('customer_', '')
            try:
                customer = Customer.objects.get(id=customer_id)

                if company_ids and customer.company_id not in company_ids:
                    filtered_data_logs = data_logs.objects.none()
                elif filters_active and not company_ids:
                    filtered_data_logs = data_logs.objects.none()
                else:
                    data_logs_query = data_logs_query.filter(
                        Q(customer_id=customer_id, location__isnull=True) |
                        Q(location__customer_id=customer_id)
                    )

                    # ── What locations does this customer have? ────────────
                    locations    = list(CustomerLocation.objects.filter(customer_id=customer_id))
                    location_ids = [l.id for l in locations]

                    # ── Balance calculation (union of customer_id + location bills) ──
                    bal_c = BillingRecord.objects.filter(
                        Q(customer_id=customer_id) |
                        (Q(customer_location_id__in=location_ids) if location_ids else Q()),
                        balance_amount__gt=0
                    ).distinct().aggregate(total=Sum('balance_amount'))['total'] or Decimal('0.00')

                    billing_balance = bal_c

                    unapproved_payments = data_logs.objects.filter(
                        Q(customer_id=customer_id, location__isnull=True) |
                        Q(location__customer_id=customer_id)
                    ).filter(
                        is_payment=True,
                        is_approved=False
                    ).aggregate(total=Sum('payment_amount'))['total'] or Decimal('0.00')

                    customer_ledger_balance = billing_balance + unapproved_payments

                    # ── Advance amount ─────────────────────────────────────
                    # Advance and outstanding are mutually exclusive:
                    # if advance > 0, outstanding is always 0 and vice-versa.
                    advance = customer.advance_amount or Decimal('0.00')
                    if advance > 0:
                        customer_advance_amount  = advance
                        customer_ledger_balance  = Decimal('0.00')
                    else:
                        customer_advance_amount = Decimal('0.00')

            except Customer.DoesNotExist:
                filtered_data_logs = data_logs.objects.none()

        elif location_id:
            data_logs_query = data_logs_query.filter(location_id=location_id)
        
        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
                data_logs_query = data_logs_query.filter(timestamp__gte=start_date_obj)
            except ValueError:
                pass

        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                data_logs_query = data_logs_query.filter(timestamp__lt=end_date_obj)
            except ValueError:
                pass
        
        if filtered_data_logs is None:
            filtered_data_logs = data_logs_query.order_by('-timestamp')
        
        if export_excel == 'true' and filtered_data_logs and filtered_data_logs.exists():
            return export_to_excel(filtered_data_logs, entity, location_id, start_date, end_date)

    context = {
        'customers'              : customers,
        'companies'              : companies,
        'customer_locations'     : customer_locations,
        'data_logs'              : filtered_data_logs,
        'selected_entity'        : entity,
        'selected_location'      : location_id,
        'start_date'             : start_date,
        'end_date'               : end_date,
        'customer_ledger_balance': customer_ledger_balance,
        'customer_advance_amount': customer_advance_amount,
        **get_company_filter_context(request),
        'user'                   : user,
        'role_permissions'       : role_permissions,
    }

    return render(request, 'reports.html', context)
def export_to_excel(data_logs_queryset, entity, location_id, start_date, end_date):
    
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from datetime import datetime

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Logs Report"

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Report Title
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = "Data Logs Report"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Filter Information
    ws.merge_cells('A2:K2')
    filter_info = []
    if start_date:
        filter_info.append(f"From: {start_date}")
    if end_date:
        filter_info.append(f"To: {end_date}")
    if entity:
        filter_info.append(f"Entity: {entity}")
    if location_id:
        filter_info.append(f"Location ID: {location_id}")
    
    filter_cell = ws['A2']
    filter_cell.value = " | ".join(filter_info) if filter_info else "All Records"
    filter_cell.alignment = Alignment(horizontal="center")

    # Headers (Row 4)
    headers = [
        'Date & Time',
        'Action',
        'Customer',
        'Company',
        'Location',
        'Payment Amount',
        'Total Paid',
        'Balance',
        'Payment Mode',
        'Status',
        'User'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Data rows
    row_num = 5
    for log in data_logs_queryset:
        ws.cell(row=row_num, column=1, value=log.timestamp.strftime('%Y-%m-%d %H:%M:%S')).border = border
       
        ws.cell(row=row_num, column=3, value=log.customer.name if log.customer else 'N/A').border = border
        ws.cell(row=row_num, column=4, value=log.customer.company.name if (log.customer and log.customer.company) else 'N/A').border = border
        ws.cell(row=row_num, column=5, value=log.location.location_name if log.location else 'Main').border = border
        ws.cell(row=row_num, column=6, value=float(log.payment_amount) if log.payment_amount else 0).border = border
        ws.cell(row=row_num, column=7, value=float(log.total_paid) if log.total_paid else 0).border = border
        ws.cell(row=row_num, column=8, value=float(log.balance_amount) if log.balance_amount else 0).border = border
        ws.cell(row=row_num, column=9, value=log.payment_mode or 'N/A').border = border
        ws.cell(row=row_num, column=10, value=log.status or 'N/A').border = border
        ws.cell(row=row_num, column=11, value=log.user.name if (log.user and log.user.name) else (log.user.username if log.user else 'N/A')).border = border
        
        row_num += 1

    # Adjust column widths
    column_widths = {
        'A': 20,  # Date & Time
        'C': 25,  # Customer
        'D': 25,  # Company
        'E': 20,  # Location
        'F': 15,  # Payment Amount
        'G': 15,  # Total Paid
        'H': 15,  # Balance
        'I': 15,  # Payment Mode
        'J': 15,  # Status
        'K': 20,  # User
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'data_logs_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

def get_customer_locations(request):
    """AJAX endpoint to get locations for a selected customer"""
    customer_id = request.GET.get('customer_id')
    
    if customer_id:
        locations = CustomerLocation.objects.filter(
            customer_id=customer_id
        ).values('id', 'location_name', 'address')
        
        return JsonResponse({
            'locations': list(locations)
        })
    
    return JsonResponse({'locations': []})

def export_to_excel(data_logs_queryset, entity, location_id, start_date, end_date):
    """Export filtered data logs to Excel file"""
    
    try:
        # Convert queryset to list to ensure data is fetched
        data_logs_list = list(data_logs_queryset)
        
        if not data_logs_list:
            # Return empty response if no data
            response = HttpResponse("No data to export", status=400)
            return response
        
        # Create a workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reports"
        
        # Define header style
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Define headers
        headers = [
            "Customer Name", "Location", "Company", "Address", "Area", "City", "State",
            "Statement Period", "Service Status", "Payment Mode", "Pay Date", 
            "Total", "Paid", "Pending Balance", "Created"
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data rows
        for row_num, log in enumerate(data_logs_list, 2):
            ws.cell(row=row_num, column=1).value = log.customer.name if log.customer else "N/A"
            ws.cell(row=row_num, column=2).value = log.location.location_name if log.location else "N/A"
            ws.cell(row=row_num, column=3).value = log.customer.company.name if (log.customer and log.customer.company) else "N/A"
            
            # Use location address if available, otherwise customer address
            address = ""
            if log.location:
                address = log.location.address
            elif log.customer:
                address = log.customer.address
            ws.cell(row=row_num, column=4).value = address
            
            # Area, City, State from location or customer
            if log.location:
                ws.cell(row=row_num, column=5).value = str(log.location.area) if log.location.area else ""
                ws.cell(row=row_num, column=6).value = str(log.location.city) if log.location.city else ""
                ws.cell(row=row_num, column=7).value = str(log.location.state) if log.location.state else ""
            elif log.customer:
                ws.cell(row=row_num, column=5).value = str(log.customer.area) if log.customer.area else ""
                ws.cell(row=row_num, column=6).value = str(log.customer.city) if log.customer.city else ""
                ws.cell(row=row_num, column=7).value = str(log.customer.state) if log.customer.state else ""
            
            ws.cell(row=row_num, column=8).value = f"{log.billing_period_start.strftime('%d/%m/%Y')} - {log.billing_period_end.strftime('%d/%m/%Y')}" if log.billing_period_start and log.billing_period_end else "N/A"
            
            # Service status
            status = "N/A"
            if log.location:
                status = "Active" if log.location.is_active else "Inactive"
            elif log.customer:
                status = log.customer.status if log.customer.status else "N/A"
            ws.cell(row=row_num, column=9).value = status
            
            ws.cell(row=row_num, column=10).value = log.payment_mode if log.payment_mode else "N/A"
            ws.cell(row=row_num, column=11).value = log.payment_date.strftime("%d/%m/%Y") if log.payment_date else "N/A"
            ws.cell(row=row_num, column=12).value = float(log.payment_amount) if log.payment_amount else 0.00
            ws.cell(row=row_num, column=13).value = float(log.total_paid) if log.total_paid else 0.00
            ws.cell(row=row_num, column=14).value = float(log.balance_amount) if log.balance_amount else 0.00
            ws.cell(row=row_num, column=15).value = log.timestamp.strftime("%d/%m/%Y") if log.timestamp else "N/A"
            
            # Apply alignment to data cells
            for col in range(1, 16):
                ws.cell(row=row_num, column=col).alignment = Alignment(vertical="center", horizontal="center")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25  # Customer Name
        ws.column_dimensions['B'].width = 25  # Location
        ws.column_dimensions['C'].width = 25  # Company
        ws.column_dimensions['D'].width = 30  # Address
        ws.column_dimensions['E'].width = 15  # Area
        ws.column_dimensions['F'].width = 15  # City
        ws.column_dimensions['G'].width = 15  # State
        ws.column_dimensions['H'].width = 25  # Statement Period
        ws.column_dimensions['I'].width = 15  # Service Status
        ws.column_dimensions['J'].width = 15  # Payment Mode
        ws.column_dimensions['K'].width = 15  # Pay Date
        ws.column_dimensions['L'].width = 15  # Total
        ws.column_dimensions['M'].width = 15  # Paid
        ws.column_dimensions['N'].width = 18  # Pending Balance
        ws.column_dimensions['O'].width = 20  # Created
        
        # Prepare the response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Generate filename with date range if provided
        filename = "reports"
        if start_date and end_date:
            filename += f"_{start_date}_to_{end_date}"
        elif start_date:
            filename += f"_from_{start_date}"
        elif end_date:
            filename += f"_until_{end_date}"
        
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        # Save workbook to response
        wb.save(response)
        
        return response
        
    except Exception as e:
        # Log the error for debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error exporting to Excel: {str(e)}", exc_info=True)
        
        return HttpResponse(f"Error exporting data: {str(e)}", status=500)
    
import threading
import logging
from django.db import connection, close_old_connections

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Background worker
# ─────────────────────────────────────────────────────────────────────────────

def _generate_bills_for_customers(customer_ids, user_id):
    """
    Run generate_automatic_bill for every newly uploaded customer.
    Executes in a background thread — MUST manage its own DB connection.

    Key fixes vs the broken version:
      1. close_old_connections() at the top — forces Django to open a fresh
         connection for this thread (the one inherited from the request thread
         gets closed when the response is sent).
      2. Pass user_id (int) not the ORM object — re-fetch inside the thread
         so it uses the new connection.
      3. Wrap each customer in its own try/except so one failure doesn't
         abort the rest.
      4. Call close_old_connections() again at the end to release the
         connection back to the pool.
    """

    print(f"[BG BILLING] 🚀 Thread started for {len(customer_ids)} customer(s)")

    # ── CRITICAL: open a fresh DB connection for this thread ──────────────────
    close_old_connections()

    # Re-fetch user inside this thread's connection context
    user = None
    if user_id:
        try:
            user = Custom_User.objects.get(id=user_id)
        except Exception as e:
            print(f"[BG BILLING] ⚠️ Could not fetch user id={user_id}: {e}")

    for customer_id in customer_ids:
        try:
            # Refresh connection before each customer (long-running thread safety)
            close_old_connections()

            customer = Customer.objects.select_related(
                'company', 'state', 'company__state', 'subscription_plan'
            ).get(id=customer_id)

            print(f"[BG BILLING] Processing: {customer.name} | start_date={customer.start_date} | plan={customer.subscription_plan}")

            # Guard: needs start_date AND subscription_plan
            if not customer.start_date:
                print(f"[BG BILLING] ⏭ Skipping {customer.name} — no start_date")
                continue

            if not customer.subscription_plan:
                print(f"[BG BILLING] ⏭ Skipping {customer.name} — no subscription_plan")
                continue

            # Guard: start_date must be BEFORE current month for bills to generate
            from datetime import date
            first_day_current_month = date.today().replace(day=1)
            if customer.start_date >= first_day_current_month:
                print(
                    f"[BG BILLING] ⏭ Skipping {customer.name} — "
                    f"start_date {customer.start_date} is in current/future month, no past bills to generate"
                )
                continue

            print(f"[BG BILLING] ✅ Generating bills for {customer.name}...")

            generate_automatic_bill(
                entity=customer,
                entity_type="customer",
                user=user,
            )

            print(f"[BG BILLING] ✅ Done: {customer.name}")

        except Customer.DoesNotExist:
            print(f"[BG BILLING] ❌ Customer ID {customer_id} not found in DB")
        except Exception as e:
            import traceback
            print(f"[BG BILLING] ❌ Error for customer ID {customer_id}: {e}")
            traceback.print_exc()

    # Release the DB connection when done
    close_old_connections()
    print("[BG BILLING] 🏁 Thread finished")


# ─────────────────────────────────────────────────────────────────────────────
# Main view
# ─────────────────────────────────────────────────────────────────────────────

def bulk_upload_customers(request):
    """
    Handle bulk upload from Excel file.
    After upload, fires background thread to auto-generate bills.
    """
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']

            df = pd.read_excel(excel_file, header=0)
            df.columns = df.columns.str.strip()

            required_columns = ['Customer', 'Phone', 'State', 'City', 'Area']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                messages.error(request, f'Missing required columns: {", ".join(missing_columns)}')
                return redirect('customer_management')

            df = df.dropna(how='all')
            df = df[df['Customer'].notna()]
            df = df[df['Customer'].astype(str).str.strip() != '']

            success_count    = 0
            error_count      = 0
            errors           = []
            new_customer_ids = []   # IDs eligible for bill generation

            for index, row in df.iterrows():
                try:
                    # ── Customer Name ──────────────────────────────────────────
                    customer_name = str(row['Customer']).strip()

                    if customer_name.replace('+', '').replace('-', '').replace(' ', '').isdigit():
                        error_count += 1
                        errors.append(f"Row {index + 2}: Name looks like a phone number: {customer_name}")
                        continue

                    if not customer_name or customer_name.lower() == 'nan':
                        error_count += 1
                        errors.append(f"Row {index + 2}: Customer name is empty")
                        continue

                    # ── Company ────────────────────────────────────────────────
                    company = None
                    if 'Company' in df.columns and pd.notna(row['Company']):
                        company_name = str(row['Company']).strip()
                        if company_name and company_name.lower() not in ['nan', 'none', '']:
                            company = Company.objects.filter(name__iexact=company_name).first()
                            if not company:
                                company = Company.objects.filter(name__icontains=company_name).first()
                            if not company:
                                error_count += 1
                                errors.append(
                                    f"Row {index + 2}: Company '{company_name}' not found. "
                                    "Please create the company first."
                                )
                                continue

                    # ── State ──────────────────────────────────────────────────
                    state_name = str(row['State']).strip()
                    if not state_name or state_name.lower() == 'nan':
                        raise ValueError("State is required")
                    state = State.objects.filter(name__iexact=state_name).first()
                    if not state:
                        state = State.objects.create(name=state_name)

                    # ── City ───────────────────────────────────────────────────
                    city_name = str(row['City']).strip()
                    if not city_name or city_name.lower() == 'nan':
                        raise ValueError("City is required")
                    city = City.objects.filter(name__iexact=city_name, state=state).first()
                    if not city:
                        city = City.objects.create(name=city_name, state=state)

                    # ── Area ───────────────────────────────────────────────────
                    area_name = str(row['Area']).strip()
                    if not area_name or area_name.lower() == 'nan':
                        raise ValueError("Area is required")
                    area = Area.objects.filter(name__iexact=area_name, city=city).first()
                    if not area:
                        area = Area.objects.create(name=area_name, city=city)

                    # ── Service Start Date ─────────────────────────────────────
                    start_date = None
                    if 'Service Start' in df.columns and pd.notna(row['Service Start']):
                        try:
                            start_date = pd.to_datetime(row['Service Start'], dayfirst=True).date()
                        except Exception:
                            start_date = datetime.now().date()

                    # ── Subscription Plan ──────────────────────────────────────
                    subscription_plan = None
                    if 'Accounting' in df.columns and pd.notna(row['Accounting']):
                        service_type = str(row['Accounting']).strip()
                        if service_type and service_type.lower() != 'nan':
                            try:
                                subscription_plan = SubscriptionPlan.objects.filter(
                                    name__icontains=service_type
                                ).first()
                                if not subscription_plan:
                                    print(f"[UPLOAD] ⚠️ Row {index+2}: Plan '{service_type}' not found in DB")
                            except Exception:
                                pass

                    # ── Previous Subscription Plan ─────────────────────────────
                    previous_subscription_plan = None
                    if 'Previous Plan' in df.columns and pd.notna(row['Previous Plan']):
                        prev_plan_name = str(row['Previous Plan']).strip()
                        if prev_plan_name and prev_plan_name.lower() not in ['nan', 'none', '']:
                            try:
                                previous_subscription_plan = SubscriptionPlan.objects.filter(
                                    name__icontains=prev_plan_name
                                ).first()
                            except Exception:
                                pass

                    # ── Upgrade Date ───────────────────────────────────────────
                    upgrade_date = None
                    if 'Upgrade Date' in df.columns and pd.notna(row['Upgrade Date']):
                        try:
                            upgrade_date = pd.to_datetime(row['Upgrade Date'], dayfirst=True).date()
                        except Exception:
                            pass

                    # ── Address + Pincode ──────────────────────────────────────
                    address = ''
                    if 'Address' in df.columns and pd.notna(row['Address']):
                        address = str(row['Address']).strip()
                        if address.lower() == 'nan':
                            address = ''

                    pincode = ''
                    if 'Pincode' in df.columns and pd.notna(row['Pincode']):
                        pincode = str(row['Pincode']).strip()
                        if pincode.lower() == 'nan':
                            pincode = ''

                    # ── Location Link ──────────────────────────────────────────
                    location_link = None
                    if 'Location Link' in df.columns and pd.notna(row['Location Link']):
                        location_link = str(row['Location Link']).strip()
                        if location_link and location_link.lower() not in ['nan', 'none', '']:
                            if not location_link.startswith(('http://', 'https://')):
                                location_link = 'https://' + location_link
                        else:
                            location_link = None

                    # ── Status ─────────────────────────────────────────────────
                    status = 'In-Active'
                    if 'Service' in df.columns and pd.notna(row['Service']):
                        if str(row['Service']).strip().lower() in ['running', 'active']:
                            status = 'Active'

                    # ── Financial Fields ───────────────────────────────────────
                    advance_amount = None
                    if 'Advance Amt' in df.columns and pd.notna(row['Advance Amt']):
                        try:
                            advance_amount = float(row['Advance Amt'])
                        except Exception:
                            advance_amount = None

                    outstanding_amount = 0.00
                    if 'Outstanding Amount' in df.columns and pd.notna(row['Outstanding Amount']):
                        try:
                            outstanding_amount = float(row['Outstanding Amount'])
                        except Exception:
                            outstanding_amount = 0.00

                    custom_subscription_amount = None
                    if 'Custom Amount' in df.columns and pd.notna(row['Custom Amount']):
                        try:
                            custom_subscription_amount = float(row['Custom Amount'])
                        except Exception:
                            custom_subscription_amount = None

                    # ── GST Fields ─────────────────────────────────────────────
                    gst_registered = False
                    gst_number     = None

                    if 'GST Registered' in df.columns and pd.notna(row['GST Registered']):
                        if str(row['GST Registered']).strip().lower() in ['yes', 'y', 'true', '1', 'registered']:
                            gst_registered = True

                    if 'GST Number' in df.columns and pd.notna(row['GST Number']):
                        raw_gst = str(row['GST Number']).strip().upper()
                        if raw_gst and raw_gst.lower() not in ['nan', 'none', '']:
                            gst_registered = True
                            if len(raw_gst) != 15:
                                error_count += 1
                                errors.append(
                                    f"Row {index + 2}: Invalid GST number '{raw_gst}' (must be 15 characters)"
                                )
                                continue
                            gst_number = raw_gst
                        else:
                            gst_number = None

                    # ── Tax Percentage ─────────────────────────────────────────
                    tax_percentage = None
                    if 'Tax Percentage' in df.columns and pd.notna(row['Tax Percentage']):
                        try:
                            tax_percentage = float(row['Tax Percentage'])
                        except Exception:
                            tax_percentage = None

                    # ── Email (auto-generated) ─────────────────────────────────
                    email = f"{customer_name.replace(' ', '_').lower()}@customer.com"

                    # ── Duplicate Checks ───────────────────────────────────────
                    existing_customer = Customer.objects.filter(name__iexact=customer_name).first()
                    if existing_customer:
                        if company and existing_customer.company == company:
                            error_count += 1
                            errors.append(
                                f"Row {index + 2}: Customer '{customer_name}' already exists "
                                f"under company '{company.name}'"
                            )
                            continue
                        elif not company:
                            error_count += 1
                            errors.append(f"Row {index + 2}: Customer '{customer_name}' already exists")
                            continue

                    if gst_number:
                        existing_gst = Customer.objects.filter(gst_number=gst_number).first()
                        if existing_gst:
                            error_count += 1
                            errors.append(
                                f"Row {index + 2}: GST number '{gst_number}' already exists "
                                f"for customer '{existing_gst.name}'"
                            )
                            continue

                    phone = None
                    if pd.notna(row['Phone']):
                        phone = str(row['Phone']).strip().replace(' ', '').replace('-', '')
                        if phone and phone.lower() != 'nan':
                            existing_contact = CustomerContact.objects.filter(
                                phone_number=phone
                            ).first()
                            if existing_contact:
                                error_count += 1
                                errors.append(
                                    f"Row {index + 2}: Phone '{phone}' already exists "
                                    f"for customer '{existing_contact.customer.name}'"
                                )
                                continue

                    # ── Remarks ────────────────────────────────────────────────
                    remarks_parts = []
                    if 'Rate' in df.columns and pd.notna(row['Rate']):
                        rate = str(row['Rate']).strip()
                        if rate and rate.lower() != 'nan':
                            remarks_parts.append(f"Rate: Rs.{rate}")
                    if 'Created' in df.columns and pd.notna(row['Created']):
                        created_val = str(row['Created']).strip()
                        if created_val and created_val.lower() != 'nan':
                            remarks_parts.append(f"Created: {created_val}")
                    if 'Remarks' in df.columns and pd.notna(row['Remarks']):
                        custom_remarks = str(row['Remarks']).strip()
                        if custom_remarks and custom_remarks.lower() != 'nan':
                            remarks_parts.append(custom_remarks)
                    remarks = ' | '.join(remarks_parts) if remarks_parts else ''

                    # ── Create Customer ────────────────────────────────────────
                    customer = Customer.objects.create(
                        name                       = customer_name,
                        email                      = email,
                        address                    = address,
                        company                    = company,
                        state                      = state,
                        city                       = city,
                        area                       = area,
                        pincode                    = pincode,
                        location_link              = location_link,
                        start_date                 = start_date,
                        status                     = status,
                        subscription_plan          = subscription_plan,
                        previous_subscription_plan = previous_subscription_plan,
                        upgrade_date               = upgrade_date,
                        advance_amount             = advance_amount,
                        outstanding_amount         = outstanding_amount,
                        custom_subscription_amount = custom_subscription_amount,
                        gst_registered             = gst_registered,
                        gst_number                 = gst_number,
                        tax_percentage             = tax_percentage,
                        remarks                    = remarks,
                    )

                    if phone:
                        CustomerContact.objects.create(customer=customer, phone_number=phone)

                    # Only queue for billing if they have both required fields
                    if start_date and subscription_plan:
                        new_customer_ids.append(customer.id)
                        print(f"[UPLOAD] ✅ Created & queued for billing: {customer.name} (id={customer.id})")
                    else:
                        print(f"[UPLOAD] ✅ Created (not queued — missing start_date or plan): {customer.name}")

                    success_count += 1

                except Exception as e:
                    error_count += 1
                    import traceback
                    print(f"[UPLOAD] ❌ Row {index+2} error: {e}")
                    traceback.print_exc()
                    errors.append(f"Row {index + 2} ({row.get('Customer', 'Unknown')}): {str(e)}")

            # ── Result Messages ────────────────────────────────────────────────
            if success_count > 0:
                messages.success(request, f'✅ Successfully imported {success_count} customer(s)!')

            if error_count > 0:
                error_summary = f'⚠️ Failed to import {error_count} customer(s). '
                if len(errors) <= 5:
                    error_summary += 'Errors: ' + ' | '.join(errors)
                else:
                    error_summary += 'First 5 errors: ' + ' | '.join(errors[:5])
                messages.warning(request, error_summary)

            # ── 🚀 Fire background bill generation ────────────────────────────
            if new_customer_ids:
                # ✅ Pass user ID (int), not the ORM object — re-fetched inside thread
                logged_in_username = get_logged_in_user(request)[0]
                bg_user_id = None
                try:
                    bg_user_id = Custom_User.objects.get(username=logged_in_username).id
                except Exception as e:
                    print(f"[UPLOAD] ⚠️ Could not get user id for billing thread: {e}")

                print(f"[UPLOAD] 🚀 Firing billing thread for IDs: {new_customer_ids}")

                billing_thread = threading.Thread(
                    target=_generate_bills_for_customers,
                    args=(new_customer_ids, bg_user_id),
                    daemon=True,
                    name="BulkBillingThread",
                )
                billing_thread.start()

                messages.info(
                    request,
                    f'🔄 Bill generation started in the background for '
                    f'{len(new_customer_ids)} eligible customer(s). '
                    f'Bills and invoices will be emailed automatically.'
                )
            else:
                print("[UPLOAD] ℹ️ No customers queued for billing (all missing start_date or plan)")

        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f'❌ Error processing file: {str(e)}')

    return redirect('customer_management')


from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from .models import Company, Customer, BillingRecord

def export_companies_excel(request):
    """Export all companies to Excel"""
    companies = Company.objects.all().select_related('area', 'city', 'state', 'subscription_plan')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ['S.No', 'Company Name', 'Contact', 'Email', 'GST Number', 'Address', 
               'Area', 'City', 'State', 'Pincode', 'Contact Person', 'Contact Person Email', 
               'Contact Person Phone', 'Start Date', 'Status', 'Subscription Plan']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_num, company in enumerate(companies, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=company.name)
        ws.cell(row=row_num, column=3, value=company.contact)
        ws.cell(row=row_num, column=4, value=company.email)
        ws.cell(row=row_num, column=5, value=company.gst_number or 'N/A')
        ws.cell(row=row_num, column=6, value=company.address)
        ws.cell(row=row_num, column=7, value=company.area.name if company.area else '')
        ws.cell(row=row_num, column=8, value=company.city.name if company.city else '')
        ws.cell(row=row_num, column=9, value=company.state.name if company.state else '')
        ws.cell(row=row_num, column=10, value=company.pincode or '')
        ws.cell(row=row_num, column=11, value=company.Contact_person or '')
        ws.cell(row=row_num, column=12, value=company.contact_person_email or '')
        ws.cell(row=row_num, column=13, value=company.contact_person_phone or '')
        ws.cell(row=row_num, column=14, value=company.start_date.strftime('%Y-%m-%d') if company.start_date else '')
        ws.cell(row=row_num, column=15, value=company.status or '')
        ws.cell(row=row_num, column=16, value=company.subscription_plan.name if company.subscription_plan else '')
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Companies_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


def export_customers_excel(request, status=None):
    """Export customers to Excel (all, active, or inactive)"""
    if status == 'active':
        customers = Customer.objects.filter(status='Active').select_related('area', 'city', 'state', 'company', 'subscription_plan')
        filename = 'Active_Customers'
    elif status == 'inactive':
        customers = Customer.objects.exclude(status='Active').select_related('area', 'city', 'state', 'company', 'subscription_plan')
        filename = 'Inactive_Customers'
    else:
        customers = Customer.objects.all().select_related('area', 'city', 'state', 'company', 'subscription_plan')
        filename = 'All_Customers'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Header styling
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ['S.No', 'Customer Name', 'Company', 'Email', 'GST Number', 'Address', 
               'Area', 'City', 'State', 'Pincode', 'Phone Numbers', 'Start Date', 
               'Status', 'Subscription Plan', 'Tax %']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_num, customer in enumerate(customers, 2):
        # Get all phone numbers
        phone_numbers = ', '.join([contact.phone_number for contact in customer.contacts.all()])
        
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=customer.name)
        ws.cell(row=row_num, column=3, value=customer.company.name if customer.company else '')
        ws.cell(row=row_num, column=4, value=customer.email)
        ws.cell(row=row_num, column=5, value=customer.gst_number or 'N/A')
        ws.cell(row=row_num, column=6, value=customer.address)
        ws.cell(row=row_num, column=7, value=customer.area.name if customer.area else '')
        ws.cell(row=row_num, column=8, value=customer.city.name if customer.city else '')
        ws.cell(row=row_num, column=9, value=customer.state.name if customer.state else '')
        ws.cell(row=row_num, column=10, value=customer.pincode or '')
        ws.cell(row=row_num, column=11, value=phone_numbers)
        ws.cell(row=row_num, column=12, value=customer.start_date.strftime('%Y-%m-%d') if customer.start_date else '')
        ws.cell(row=row_num, column=13, value=customer.status or '')
        ws.cell(row=row_num, column=14, value=customer.subscription_plan.name if customer.subscription_plan else '')
        ws.cell(row=row_num, column=15, value=float(customer.tax_percentage) if customer.tax_percentage else '')
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


def export_invoices_excel(request):
    """Export all billing records/invoices to Excel"""
    invoices = BillingRecord.objects.all().select_related(
        'customer', 'company', 'customer_location', 'company_location'
    ).order_by('-billing_date')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    
    # Header styling
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    
    # Headers
    headers = ['S.No', 'Invoice Number', 'Client Name', 'Client Type', 'Location', 
               'Amount (₹)', 'Billing Date', 'Due Date', 'Payment Status', 
               'Payment Date', 'Payment Mode', 'Transaction ID', 'Notes']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_num, invoice in enumerate(invoices, 2):
        client_type = 'Customer' if invoice.customer else 'Company'
        payment_status = 'Paid' if invoice.paid else 'Unpaid'
        
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=invoice.invoice_number or 'N/A')
        ws.cell(row=row_num, column=3, value=invoice.client_name)
        ws.cell(row=row_num, column=4, value=client_type)
        ws.cell(row=row_num, column=5, value=invoice.location_name)
        ws.cell(row=row_num, column=6, value=float(invoice.amount))
        ws.cell(row=row_num, column=7, value=invoice.billing_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=8, value=invoice.due_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=9, value=payment_status)
        ws.cell(row=row_num, column=10, value=invoice.payment_date.strftime('%Y-%m-%d') if invoice.payment_date else '')
        ws.cell(row=row_num, column=11, value=invoice.payment_mode or '')
        ws.cell(row=row_num, column=12, value=invoice.transaction_id or '')
        ws.cell(row=row_num, column=13, value=invoice.notes or '')
        
        # Color code based on payment status
        if invoice.paid:
            for col in range(1, 14):
                ws.cell(row=row_num, column=col).fill = PatternFill(
                    start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"
                )
        else:
            for col in range(1, 14):
                ws.cell(row=row_num, column=col).fill = PatternFill(
                    start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"
                )
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 40)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=All_Invoices_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


def export_financial_summary_excel(request, report_type):
    """Export financial summaries (total_amount, collected, outstanding)"""
    wb = Workbook()
    ws = wb.active
    
    if report_type == 'total_amount':
        ws.title = "Total Amount"
        invoices = BillingRecord.objects.all().select_related(
            'customer', 
            'customer_location'
        )
        filename = 'Total_Amount_Summary'
        header_color = "E74C3C"
        
    elif report_type == 'collected':
        ws.title = "Collected Amount"
        # Only show approved payments
        invoices = BillingRecord.objects.filter(paid=True).exclude(
            data_logs__is_payment=True,
            data_logs__is_approved=False
        ).select_related(
            'customer', 
            'customer_location'
        )
        filename = 'Collected_Amount_Summary'
        header_color = "1ABC9C"
        
    elif report_type == 'outstanding':
        ws.title = "Outstanding Amount"
        invoices = BillingRecord.objects.filter(paid=False).select_related(
            'customer', 
            'customer_location'
        )
        filename = 'Outstanding_Amount_Summary'
        header_color = "F39C12"
    
    # Header styling
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ['S.No', 'Invoice Number', 'Client Name', 'Location', 'Amount (₹)', 
               'Billing Date', 'Due Date', 'Payment Status', 'Payment Date', 'Payment Mode']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    total_amount = 0
    for row_num, invoice in enumerate(invoices, 2):
        total_amount += invoice.amount
        
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=invoice.invoice_number or 'N/A')
        ws.cell(row=row_num, column=3, value=invoice.client_name)
        ws.cell(row=row_num, column=4, value=invoice.location_name)
        ws.cell(row=row_num, column=5, value=float(invoice.amount))
        ws.cell(row=row_num, column=6, value=invoice.billing_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=7, value=invoice.due_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=8, value='Paid' if invoice.paid else 'Unpaid')
        ws.cell(row=row_num, column=9, value=invoice.payment_date.strftime('%Y-%m-%d') if invoice.payment_date else '')
        ws.cell(row=row_num, column=10, value=invoice.payment_mode or '')
    
    # Add total row
    last_row = invoices.count() + 2
    ws.cell(row=last_row, column=4, value='TOTAL:').font = Font(bold=True)
    ws.cell(row=last_row, column=5, value=float(total_amount)).font = Font(bold=True)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 40)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response




import pandas as pd
from django.shortcuts import redirect
from django.contrib import messages
from datetime import datetime
from django.db.models import Q

import pandas as pd
from datetime import datetime, date
from django.shortcuts import redirect
from django.contrib import messages
from django.db.models import Q
from decimal import Decimal
def bulk_upload_billing_records(request):
    """
    Bulk upload billing records from Excel file - UPDATE ONLY MODE
    Updates existing BillingRecord and data_logs with payment information
    Does NOT create new billing records
    """

    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file)
            df.columns = df.columns.str.strip()

            required_columns = [
                'CUSTOMER NAME',
                'LOCATION',
                'BILLING START DATE',
                'BILLING END DATE',
                'AMOUNT PAID',
                'PAYMENT DATE',
                'PAYMENT MODE',
            ]

            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                messages.error(request, f'Missing required columns: {", ".join(missing_columns)}')
                return redirect('billing_records')

            success_count   = 0
            error_count     = 0
            not_found_count = 0
            errors          = []

            for index, row in df.iterrows():
                try:
                    # ── Customer ──────────────────────────────────
                    customer_name = str(row['CUSTOMER NAME']).strip()
                    if not customer_name or customer_name.lower() == 'nan':
                        error_count += 1
                        errors.append(f"Row {index + 2}: Empty customer name")
                        continue

                    customer = (
                        Customer.objects.filter(name__iexact=customer_name).first()
                        or Customer.objects.filter(name__icontains=customer_name).first()
                    )
                    if not customer:
                        not_found_count += 1
                        errors.append(f"Row {index + 2}: Customer '{customer_name}' not found")
                        continue

                    # ── Location (optional) ───────────────────────
                    location_name     = str(row['LOCATION']).strip() if pd.notna(row['LOCATION']) else None
                    customer_location = None

                    if location_name and location_name.lower() not in ['nan', 'none', '']:
                        customer_location = (
                            CustomerLocation.objects.filter(
                                customer=customer,
                                location_name__iexact=location_name
                            ).first()
                            or CustomerLocation.objects.filter(
                                customer=customer,
                                location_name__icontains=location_name
                            ).first()
                        )
                        if not customer_location:
                            error_count += 1
                            errors.append(
                                f"Row {index + 2}: Location '{location_name}' not found "
                                f"for customer '{customer_name}'"
                            )
                            continue

                    # ── Dates ─────────────────────────────────────
                    billing_start_date = billing_end_date = payment_date = None
                    try:
                        if pd.notna(row['BILLING START DATE']):
                            billing_start_date = pd.to_datetime(
                                row['BILLING START DATE'], dayfirst=True
                            ).date()
                        if pd.notna(row['BILLING END DATE']):
                            billing_end_date = pd.to_datetime(
                                row['BILLING END DATE'], dayfirst=True
                            ).date()
                        if pd.notna(row['PAYMENT DATE']):
                            payment_date = pd.to_datetime(
                                row['PAYMENT DATE'], dayfirst=True
                            ).date()
                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {index + 2}: Invalid date format - {str(e)}")
                        continue

                    if not billing_start_date or not billing_end_date:
                        error_count += 1
                        errors.append(f"Row {index + 2}: Billing period dates are required")
                        continue

                    # ── Amount Paid ───────────────────────────────
                    paid_amount = Decimal('0')
                    if pd.notna(row['AMOUNT PAID']):
                        try:
                            paid_amount = Decimal(str(row['AMOUNT PAID']).replace(',', ''))
                        except Exception:
                            error_count += 1
                            errors.append(f"Row {index + 2}: Invalid amount paid")
                            continue

                    # ── Payment Mode ──────────────────────────────
                    payment_mode_map = {
                        'cash':        'cash',
                        'online':      'upi',
                        'upi':         'upi',
                        'card':        'card',
                        'net banking': 'netbanking',
                        'netbanking':  'netbanking',
                        'cheque':      'cheque',
                    }
                    payment_mode = 'cash'
                    if pd.notna(row['PAYMENT MODE']):
                        payment_mode = payment_mode_map.get(
                            str(row['PAYMENT MODE']).strip().lower(), 'cash'
                        )

                    # ── Optional fields ───────────────────────────
                    transaction_id = ''
                    if 'TRANSACTION ID' in df.columns and pd.notna(row.get('TRANSACTION ID')):
                        transaction_id = str(row['TRANSACTION ID']).strip()

                    remarks = ''
                    if 'REMARKS' in df.columns and pd.notna(row.get('REMARKS')):
                        remarks = str(row['REMARKS']).strip()

                    # ── NEW: Receipt Number ───────────────────────
                    receipt_number = ''
                    if 'RECEIPT NUMBER' in df.columns and pd.notna(row.get('RECEIPT NUMBER')):
                        receipt_number = str(row['RECEIPT NUMBER']).strip()

                    # ── Derived flags ─────────────────────────────
                    is_paid = paid_amount > 0 and payment_date is not None

                    # ── Find existing BillingRecord ───────────────
                    if customer_location:
                        billing_record = BillingRecord.objects.filter(
                            customer_location=customer_location,
                            billing_start_date=billing_start_date,
                            billing_end_date=billing_end_date
                        ).first()
                    else:
                        billing_record = BillingRecord.objects.filter(
                            customer=customer,
                            customer_location__isnull=True,
                            billing_start_date=billing_start_date,
                            billing_end_date=billing_end_date
                        ).first()

                    if not billing_record:
                        not_found_count += 1
                        location_info = f" - {location_name}" if location_name else ""
                        errors.append(
                            f"Row {index + 2}: No billing record found for "
                            f"'{customer_name}{location_info}' "
                            f"for period {billing_start_date} to {billing_end_date}"
                        )
                        continue

                    # ── Recalculate balances ───────────────────────
                    total_amount       = billing_record.amount        or Decimal('0')
                    discount_amount    = billing_record.discount_amount or Decimal('0')
                    existing_paid      = billing_record.paid_amount   or Decimal('0')
                    new_total_paid     = existing_paid + paid_amount
                    balance_amount     = max(total_amount - new_total_paid - discount_amount, Decimal('0'))
                    is_fully_paid      = balance_amount <= 0

                    # ── Update BillingRecord ──────────────────────
                    billing_record.paid          = is_fully_paid
                    billing_record.paid_amount   = new_total_paid
                    billing_record.balance_amount = balance_amount
                    if is_paid:
                        billing_record.payment_date = payment_date
                        billing_record.payment_mode = payment_mode
                    if transaction_id:
                        billing_record.transaction_id = transaction_id
                    if remarks:
                        billing_record.notes = (
                            f"{billing_record.notes} | {remarks}"
                            if billing_record.notes else remarks
                        )
                    billing_record.save()

                    # ── Build details string (includes receipt) ───
                    location_detail = (
                        f" | Location: {customer_location.location_name}"
                        if customer_location else ""
                    )
                    details = (
                        f"Statement Period: {billing_start_date} to {billing_end_date}"
                        f"{location_detail}"
                        f" | Amount Paid: Rs {paid_amount}"
                        f" | Total Paid: Rs {new_total_paid}"
                        f" | Balance: Rs {balance_amount}"
                    )
                    if receipt_number:
                        details += f" | Receipt No: {receipt_number}"
                    if remarks:
                        details += f" | {remarks}"

                    # ── Update or create data_logs ────────────────
                    data_log = data_logs.objects.filter(
                        customer=customer,
                        location=customer_location,
                        billing_period_start=billing_start_date,
                        billing_period_end=billing_end_date,
                        billing_record=billing_record
                    ).first()

                    if data_log:
                        data_log.payment_amount  = paid_amount if is_paid else None
                        data_log.payment_mode    = payment_mode if is_paid else None
                        data_log.payment_date    = payment_date if is_paid else None
                        data_log.status          = 'Paid' if is_fully_paid else 'Pending'
                        data_log.is_payment      = is_paid
                        data_log.is_approved     = is_paid
                        data_log.transaction_id  = transaction_id
                        data_log.payment_notes   = remarks
                        data_log.total_paid      = new_total_paid
                        data_log.balance_amount  = balance_amount
                        data_log.receipt_number  = receipt_number or data_log.receipt_number
                        data_log.details         = details
                        data_log.save()
                    else:
                        action = (
                            "Bulk Upload - Payment Received" if is_paid
                            else "Bulk Upload - Updated"
                        )
                        data_logs.objects.create(
                            user=request.user if hasattr(request, 'user') else None,
                            action=action,
                            details=details,
                            customer=customer,
                            location=customer_location,
                            billing_record=billing_record,
                            payment_amount=paid_amount if is_paid else None,
                            billing_period_start=billing_start_date,
                            billing_period_end=billing_end_date,
                            status='Paid' if is_fully_paid else 'Pending',
                            payment_mode=payment_mode if is_paid else None,
                            payment_date=payment_date if is_paid else None,
                            is_payment=is_paid,
                            is_approved=is_paid,
                            transaction_id=transaction_id,
                            payment_notes=remarks,
                            total_paid=new_total_paid,
                            balance_amount=balance_amount,
                            receipt_number=receipt_number or None,  # NEW
                        )

                    success_count += 1

                except Exception as e:
                    error_count += 1
                    customer_name_safe = row.get('CUSTOMER NAME', 'Unknown')
                    errors.append(f"Row {index + 2} ({customer_name_safe}): {str(e)}")

            # ── Summary messages ──────────────────────────────────
            if success_count > 0:
                messages.success(request, f'Successfully updated {success_count} billing record(s)!')
            if not_found_count > 0:
                messages.warning(request, f'Could not find {not_found_count} billing record(s) in database')
            if error_count > 0:
                error_summary = f'Failed to process {error_count} record(s). '
                shown = errors[:5]
                error_summary += ('Errors: ' if len(errors) <= 5 else 'First 5 errors: ')
                error_summary += ' | '.join(shown)
                messages.error(request, error_summary)

        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')

    return redirect('manual_billing')



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum
from decimal import Decimal, InvalidOperation
import uuid as uuid_lib
from decimal import Decimal
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
def manual_billing_view(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    username = request.session.get("username")
    if not username:
        return redirect("login")

    user = get_object_or_404(Custom_User, username=username)

    filtered_companies = get_filtered_companies(request)
    company_ids = list(filtered_companies.values_list('id', flat=True))

    selected_gst_type    = request.session.get('selected_gst_type', 'gst')
    selected_company_ids = request.session.get('selected_company_ids', [])
    filters_active       = selected_gst_type or selected_company_ids

    if company_ids:
        customers = Customer.objects.filter(
            company_id__in=company_ids
        ).select_related('company').order_by('name')
    elif filters_active:
        customers = Customer.objects.none()
    else:
        customers = Customer.objects.all().select_related('company').order_by('name')

    if request.method == "POST":

        submission_token = request.POST.get("submission_token", "").strip()
        processed_tokens = request.session.get("processed_payment_tokens", [])

        if submission_token and submission_token in processed_tokens:
            messages.warning(request, "This payment was already submitted. Please do not resubmit.")
            return redirect("manual_billing")

        try:
            with transaction.atomic():
                entity_type    = request.POST.get("entity_type")
                entity_id      = request.POST.get("entity_id")
                payment_date   = request.POST.get("payment_date")
                payment_mode   = request.POST.get("payment_mode")
                amount_paid    = request.POST.get("amount")
                discount       = float(request.POST.get("discount", 0))
                notes          = request.POST.get("notes", "")
                receipt_number = request.POST.get("receipt_number", "").strip() or None

                raw_locations      = request.POST.getlist("selected_locations")
                selected_locations = list(dict.fromkeys(raw_locations))

                if not entity_type or not entity_id:
                    messages.error(request, "Please select a customer.")
                    return redirect("manual_billing")

                if not selected_locations:
                    messages.error(request, "Please select at least one location.")
                    return redirect("manual_billing")

                if entity_type == "customer":
                    customer = get_object_or_404(Customer, id=entity_id)
                    if company_ids and customer.company_id not in company_ids:
                        messages.error(request, "Selected customer is not in the filtered companies.")
                        return redirect("manual_billing")
                    elif filters_active and not company_ids:
                        messages.error(request, "No companies match the current filter.")
                        return redirect("manual_billing")

                elif entity_type == "location":
                    location_obj = get_object_or_404(CustomerLocation, id=entity_id)
                    customer     = location_obj.customer
                    if company_ids and customer.company_id not in company_ids:
                        messages.error(request, "Selected location's customer is not in the filtered companies.")
                        return redirect("manual_billing")
                    elif filters_active and not company_ids:
                        messages.error(request, "No companies match the current filter.")
                        return redirect("manual_billing")
                else:
                    messages.error(request, "Invalid entity type.")
                    return redirect("manual_billing")

                # ── Validate all selected locations belong to this customer ─
                valid_location_ids = set(
                    CustomerLocation.objects.filter(
                        customer=customer
                    ).values_list('id', flat=True)
                )
                bill_ids_ordered = []
                seen_bill_ids    = set()

                for loc_id in selected_locations:
                    if loc_id == "main":
                        qs = BillingRecord.objects.filter(
                            customer=customer,
                            customer_location__isnull=True,
                        ).values_list('id', flat=True)

                    else:
                        try:
                            loc_id_int = int(loc_id)
                        except (ValueError, TypeError):
                            continue

                        if loc_id_int not in valid_location_ids:
                            messages.error(
                                request,
                                f"Location {loc_id_int} does not belong to this customer."
                            )
                            return redirect("manual_billing")

                        qs = BillingRecord.objects.filter(
                            customer_location_id=loc_id_int,
                        ).values_list('id', flat=True)

                    for bid in qs:
                        if bid not in seen_bill_ids:
                            seen_bill_ids.add(bid)
                            bill_ids_ordered.append(bid)

                if not bill_ids_ordered:
                    messages.error(request, "No bills found for the selected locations.")
                    return redirect("manual_billing")

                # ── Lock rows ─────────────────────────────────────────────
                locked_bills_qs = list(
                    BillingRecord.objects
                    .select_for_update()
                    .filter(id__in=bill_ids_ordered)
                    .order_by('billing_start_date', 'id')
                )

                seen_locked    = set()
                deduped_locked = []
                for bill in locked_bills_qs:
                    if bill.id not in seen_locked:
                        seen_locked.add(bill.id)
                        deduped_locked.append(bill)

                id_to_bill  = {b.id: b for b in deduped_locked}
                all_records = [id_to_bill[bid] for bid in bill_ids_ordered if bid in id_to_bill]
                prior_paid_snapshot    = {b.id: Decimal(str(b.paid_amount    or 0)) for b in all_records}
                prior_balance_snapshot = {b.id: Decimal(str(b.balance_amount or 0)) for b in all_records}

                billing_records = [
                    b for b in all_records
                    if prior_balance_snapshot[b.id] > Decimal('0.00')
                ]

                if not billing_records:
                    messages.error(request, "No unpaid bills found — they may have just been paid.")
                    return redirect("manual_billing")

                # ── Parse payment amounts ─────────────────────────────────
                # Discount reduces the BILL AMOUNT, not the payment.
                # Example: bill=11800, discount=1800, payment=10000
                #   effective bill = 11800 - 1800 = 10000
                #   payment 10000 >= effective 10000 → fully cleared
                customer_advance = Decimal(str(customer.advance_amount or 0))
                payment_amount   = Decimal(str(amount_paid)) if amount_paid else Decimal('0.00')
                discount_amount  = Decimal(str(discount))

                # Total cash available to distribute = payment + advance (no discount subtraction)
                total_available     = payment_amount + customer_advance
                remaining_available = total_available
                bills_updated       = 0
                total_paid_to_bills = Decimal('0.00')
                processed_in_loop   = set()

                # ── Distribute payment across bills ───────────────────────
                for bill in billing_records:
                    if bill.id in processed_in_loop:
                        continue

                    if remaining_available <= Decimal('0.00'):
                        break

                    bill_balance = prior_balance_snapshot[bill.id]
                    prior_paid   = prior_paid_snapshot[bill.id]

                    if bill_balance <= Decimal('0.00'):
                        continue

                    # Discount reduces what the customer actually needs to pay
                    # effective_balance = how much cash is needed to clear the bill
                    effective_balance = max(Decimal('0.00'), bill_balance - discount_amount)

                    if remaining_available >= effective_balance:
                        # Bill is fully cleared:
                        # cash paid = effective_balance, discount covers the rest
                        paid_now    = effective_balance
                        new_balance = Decimal('0.00')
                        bill.paid   = True
                    else:
                        # Partial payment — no discount applied yet (not enough to clear)
                        paid_now    = remaining_available
                        new_balance = bill_balance - paid_now
                        bill.paid   = False

                    bill.balance_amount  = new_balance
                    bill.paid_amount     = prior_paid + paid_now
                    bill.payment_date    = payment_date
                    bill.payment_mode    = payment_mode
                    bill.discount_amount = discount_amount  # store discount on billing record

                    if bill.customer_id is None and bill.customer_location_id:
                        bill.customer = customer

                    if notes:
                        existing_notes = bill.notes or ''
                        timestamp      = payment_date or 'No date'
                        bill.notes = (
                            existing_notes
                            + '\n[' + timestamp + '] Payment: Rs.'
                            + str(paid_now.quantize(Decimal('0.01')))
                            + ' - ' + notes
                        ).strip()

                    bill.save()

                    processed_in_loop.add(bill.id)
                    remaining_available -= paid_now
                    total_paid_to_bills += paid_now
                    bills_updated       += 1

                # ── Advance balance recalculation ─────────────────────────
                advance_used  = Decimal('0.00')
                advance_added = Decimal('0.00')
                new_advance   = customer_advance

                if payment_amount >= total_paid_to_bills:
                    excess = payment_amount - total_paid_to_bills
                    if excess > Decimal('0.00'):
                        advance_added = excess
                        new_advance   = customer_advance + advance_added
                else:
                    advance_used = total_paid_to_bills - payment_amount
                    if advance_used > customer_advance:
                        advance_used = customer_advance
                    new_advance = customer_advance - advance_used

                customer.advance_amount = new_advance
                customer.save()

                # ── Ledger / audit entries ────────────────────────────────
                for bill in billing_records[:bills_updated]:
                    paid_this_transaction = (
                        Decimal(str(bill.paid_amount or 0))
                        - prior_paid_snapshot[bill.id]
                    )

                    bill_balance_final = Decimal(str(bill.balance_amount))
                    bill_status        = "Fully Paid" if bill.paid else "Partially Paid"
                    location_info      = (
                        bill.customer_location.location_name
                        if bill.customer_location
                        else "Main / No Location"
                    )

                    ledger_detail = (
                        'Payment Entry  : ' + str(payment_date or 'No date')        + '\n'
                        'Invoice        : ' + str(bill.invoice_number or bill.id)   + '\n'
                        'Customer       : ' + customer.name                          + '\n'
                        'Location       : ' + location_info                          + '\n'
                        'Billing Period : ' + str(bill.billing_start_date or 'N/A')
                        + ' to '            + str(bill.billing_end_date   or 'N/A') + '\n'
                        'Recorded By    : ' + user.username                          + '\n'
                        'Payment Mode   : ' + str(payment_mode  or 'Not specified') + '\n'
                        'Receipt No     : ' + str(receipt_number or 'N/A')          + '\n'
                        'Discount       : Rs.' + str(discount_amount.quantize(Decimal('0.01'))) + '\n'  # ← discount
                        'Paid (this txn): Rs.' + str(paid_this_transaction.quantize(Decimal('0.01'))) + '\n'
                        'Total Paid     : Rs.' + str(Decimal(str(bill.paid_amount or 0)).quantize(Decimal('0.01'))) + '\n'
                        'Balance        : Rs.' + str(bill_balance_final.quantize(Decimal('0.01')))    + '\n'
                        'Status         : '   + bill_status + '\n\n'
                        'Advance Summary:\n'
                        '  Prior Advance : Rs.' + str(customer_advance.quantize(Decimal('0.01'))) + '\n'
                        '  Advance Used  : Rs.' + str(advance_used.quantize(Decimal('0.01')))     + '\n'
                        '  Advance Added : Rs.' + str(advance_added.quantize(Decimal('0.01')))    + '\n'
                        '  New Advance   : Rs.' + str(new_advance.quantize(Decimal('0.01')))      + '\n\n'
                        'Notes: ' + (notes or 'None')
                    )

                    data_logs.objects.create(
                        user                 = user,
                        action               = 'Payment Recorded',
                        customer             = customer,
                        location             = bill.customer_location,
                        billing_record       = bill,
                        payment_amount       = paid_this_transaction,
                        total_paid           = Decimal(str(bill.paid_amount or 0)),
                        balance_amount       = bill_balance_final,
                        status               = bill_status,
                        payment_mode         = payment_mode,
                        payment_date         = payment_date,
                        receipt_number       = receipt_number,
                        payment_notes        = notes,
                        discount_given       = discount_amount,  # ← store discount in data_logs
                        is_payment           = True,
                        is_approved          = False,
                        submitted_by         = user,
                        billing_period_start = bill.billing_start_date,
                        billing_period_end   = bill.billing_end_date,
                        details              = ledger_detail,
                    )

                user_logs.objects.create(
                    user    = user.username,
                    action  = 'Recorded Payment',
                    details = (
                        'Customer: '       + customer.name + ', '
                        'Amount Paid: Rs.' + str(payment_amount.quantize(Decimal('0.01'))) + ', '
                        'Discount: Rs.'    + str(discount_amount.quantize(Decimal('0.01'))) + ', '
                        'Bills Updated: '  + str(bills_updated) + ', '
                        'Payment Mode: '   + str(payment_mode or 'Not specified') + ', '
                        'Receipt No: '     + str(receipt_number or 'N/A')
                    )
                )

                if submission_token:
                    processed_tokens.append(submission_token)
                    request.session["processed_payment_tokens"] = processed_tokens[-20:]
                    request.session.modified = True

                message_parts = []

                if payment_amount > 0:
                    message_parts.append(
                        'Payment of Rs.' + str(payment_amount.quantize(Decimal('0.01')))
                        + ' recorded for ' + str(bills_updated) + ' bill(s).'
                    )
                else:
                    message_parts.append(
                        'Payment recorded for ' + str(bills_updated) + ' bill(s) using advance only.'
                    )

                if discount_amount > 0:
                    message_parts.append(
                        'Discount of Rs.' + str(discount_amount.quantize(Decimal('0.01'))) + ' applied.'
                    )

                if advance_used > 0:
                    message_parts.append(
                        'Rs.' + str(advance_used.quantize(Decimal('0.01'))) + ' deducted from advance.'
                    )

                if advance_added > 0:
                    message_parts.append(
                        'Rs.' + str(advance_added.quantize(Decimal('0.01'))) + ' added to advance.'
                    )

                message_parts.append(
                    'Current advance balance: Rs.' + str(new_advance.quantize(Decimal('0.01'))) + '.'
                )

                messages.success(request, ' '.join(message_parts))
                return redirect("manual_billing")

        except ValueError as e:
            messages.error(request, 'Invalid number format: ' + str(e))
            import traceback; traceback.print_exc()
            return redirect("manual_billing")

        except Exception as e:
            messages.error(request, 'Error recording payment: ' + str(e))
            import traceback; traceback.print_exc()
            return redirect("manual_billing")

    # ── GET ───────────────────────────────────────────────────────────────
    context = {
        'user'             : user,
        'role_permissions' : role_permissions,
        'customers'        : customers,
        'submission_token' : str(uuid_lib.uuid4()),
        **get_company_filter_context(request)
    }
    return render(request, 'manualbilling.html', context)

def get_entity_billing_details(request):
    entity_type = request.GET.get('entity_type')
    entity_id   = request.GET.get('entity_id')

    if not entity_type or not entity_id:
        return JsonResponse({'error': 'Missing required parameters'}, status=400)

    try:
        locations_data = []

        if entity_type == 'customer':
            customer = Customer.objects.get(id=entity_id)

            # ── Main customer bills (no location) ─────────────────────────
            # Only bills where customer_location IS NULL AND customer = this customer.
            # Bills created for a location (customer_location_id set) are handled
            # in the per-location loop below — they must NOT appear here.
            main_bills = BillingRecord.objects.filter(
                customer=customer,
                customer_location__isnull=True,
                balance_amount__gt=0
            )

            if main_bills.exists():
                total_due = sum(
                    Decimal(str(bill.balance_amount)) for bill in main_bills
                )
                locations_data.append({
                    'id'    : 'main',
                    'name'  : f'{customer.name} (Direct Customer)',
                    'count' : main_bills.count(),
                    'amount': float(total_due)
                })

            # ── Location-specific bills ────────────────────────────────────
            # FIX: Historical BillingRecords for a CustomerLocation were created
            # with customer_id=NULL — only customer_location_id was set.
            # The old query `BillingRecord.objects.filter(customer_location=location)`
            # already works correctly here (it filters by customer_location_id, so
            # customer_id=NULL rows are included).
            #
            # The real bug was in manual_billing_view: it was querying by
            # customer_location_id WITHOUT an ownership check, so a location bill
            # with customer_id=NULL would be fetched regardless of which customer
            # was selected. That is now fixed via valid_location_ids check.
            #
            # Here we just make sure we count the same bills that
            # manual_billing_view will actually process — i.e. bills found via
            # customer_location_id, NOT via customer= (since they may have NULL there).
            locations = CustomerLocation.objects.filter(customer=customer)
            for location in locations:
                location_bills = BillingRecord.objects.filter(
                    customer_location=location,   # matches by customer_location_id
                    balance_amount__gt=0
                )

                if location_bills.exists():
                    total_due = sum(
                        Decimal(str(bill.balance_amount)) for bill in location_bills
                    )
                    locations_data.append({
                        'id'    : str(location.id),
                        'name'  : location.location_name,
                        'count' : location_bills.count(),
                        'amount': float(total_due)
                    })

        elif entity_type == 'location':
            location      = CustomerLocation.objects.get(id=entity_id)
            location_bills = BillingRecord.objects.filter(
                customer_location=location,
                balance_amount__gt=0
            )

            if location_bills.exists():
                total_due = sum(
                    Decimal(str(bill.balance_amount)) for bill in location_bills
                )
                locations_data.append({
                    'id'    : str(location.id),
                    'name'  : location.location_name,
                    'count' : location_bills.count(),
                    'amount': float(total_due)
                })

        else:
            return JsonResponse({'error': 'Invalid entity type'}, status=400)

        return JsonResponse({
            'locations'      : locations_data,
            'total_locations': len(locations_data)
        })

    except Customer.DoesNotExist:
        return JsonResponse({'error': 'Customer not found'}, status=404)
    except CustomerLocation.DoesNotExist:
        return JsonResponse({'error': 'Location not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Sum, Q, Avg
from django.db.models.functions import TruncMonth, TruncYear
from datetime import datetime, timedelta
from decimal import Decimal
import json

# For PDF Export with Charts
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image, KeepTogether, Frame
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT, TA_JUSTIFY
from reportlab.graphics.shapes import Drawing, Rect, String, Circle, Line
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.legends import Legend
from reportlab.graphics import renderPDF
from io import BytesIO

from .models import (
    Customer, Company, CustomerLocation, CompanyLocation, 
    BillingRecord, SubscriptionPlan
)

from django.http import HttpResponse
from django.shortcuts import render
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncMonth
from datetime import datetime, timedelta
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, 
    Spacer, PageBreak
)
from reportlab.graphics.shapes import Drawing, Rect, String, Line
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart

# Import your models (adjust the import path as needed)
from .models import (
    Customer, Company, CustomerLocation, CompanyLocation,
    BillingRecord, SubscriptionPlan, Organization
)


def subscription_analytics_report(request):
    """Generate comprehensive PDF report with charts - Two Page Layout with Colored KPI Cards"""
    
    # ============== GET DATE RANGE ==============
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=365)
    
    if request.GET.get('start_date'):
        start_date = datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
    if request.GET.get('end_date'):
        end_date = datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
    
    # ============== CREATE PDF ==============
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="subscription_analytics_{start_date}_to_{end_date}.pdf"'
    
    doc = SimpleDocTemplate(
        response, 
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # ============== CUSTOM STYLES ==============
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1F2937'),
        spaceAfter=5,
        fontName='Helvetica-Bold',
        alignment=TA_CENTER
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#6B7280'),
        spaceAfter=15,
        alignment=TA_CENTER
    )
    
    section_title_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=11,
        textColor=colors.HexColor('#374151'),
        spaceAfter=8,
        spaceBefore=5,
        fontName='Helvetica-Bold',
        alignment=TA_LEFT
    )
    
    # ============== CALCULATE DATA - CORRECTED FILTERING ==============
    
    # TOTAL COUNTS (NOT filtered by date - these are cumulative)
    total_customers = Customer.objects.count()
    active_customers = Customer.objects.filter(status='Active').count()
    closed_customers = Customer.objects.exclude(status='Active').count()
    
    total_companies = Company.objects.count()
    active_companies = Company.objects.filter(status='active').count()
    
    active_customer_locations = CustomerLocation.objects.filter(is_active=True).count()
    active_company_locations = CompanyLocation.objects.filter(is_active=True).count()
    total_active_subscriptions = active_customers + active_companies + active_customer_locations + active_company_locations
    
    # INVOICE COUNTS - Filtered by billing_date (within selected date range)
    total_invoices = BillingRecord.objects.filter(
        billing_date__gte=start_date,
        billing_date__lte=end_date
    ).count()
    
    paid_invoices = BillingRecord.objects.filter(
        billing_date__gte=start_date,
        billing_date__lte=end_date,
        paid=True
    ).count()
    
    unpaid_invoices = total_invoices - paid_invoices
    
    # REVENUE - Filtered by billing_date
    total_amount = BillingRecord.objects.filter(
        billing_date__gte=start_date,
        billing_date__lte=end_date
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    total_collected = BillingRecord.objects.filter(
        billing_date__gte=start_date,
        billing_date__lte=end_date,
        paid=True
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    total_outstanding = total_amount - total_collected
    
    # ADVANCE PAYMENTS - Filter by customers added in date range
    advance_payment = Customer.objects.filter(
        start_date__gte=start_date,
        start_date__lte=end_date
    ).aggregate(
        total=Sum('advance_amount')
    )['total'] or 0
    
    # MONTHLY REVENUE - Filtered by billing_date
    monthly_revenue = list(BillingRecord.objects.filter(
        billing_date__gte=start_date,
        billing_date__lte=end_date
    ).annotate(
        month=TruncMonth('billing_date')
    ).values('month').annotate(
        revenue=Sum('amount')
    ).order_by('month'))
    
    # MONTHLY COLLECTION - Filtered by payment_date
    monthly_collection = list(BillingRecord.objects.filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date,
        paid=True
    ).annotate(
        month=TruncMonth('payment_date')
    ).values('month').annotate(
        collected=Sum('amount')
    ).order_by('month'))
    
    # PAYMENT METHODS - Filtered by billing_date
    payment_methods = list(BillingRecord.objects.filter(
        billing_date__gte=start_date,
        billing_date__lte=end_date,
        paid=True
    ).values('payment_mode').annotate(
        count=Count('id'),
        amount=Sum('amount')
    ).order_by('-amount'))
    
    # TOP CUSTOMERS - Filtered by billing_date
    top_customers = list(BillingRecord.objects.filter(
        billing_date__gte=start_date,
        billing_date__lte=end_date,
        paid=True,
        customer__isnull=False
    ).values('customer__name').annotate(
        revenue=Sum('amount'),
        invoice_count=Count('id')
    ).order_by('-revenue')[:10])
    
    # SUBSCRIPTION PLAN DISTRIBUTION (Total counts)
    plan_growth = []
    for plan in SubscriptionPlan.objects.all():
        customer_count = Customer.objects.filter(subscription_plan=plan).count()
        customer_location_count = CustomerLocation.objects.filter(subscription_plan=plan).count()
        company_location_count = CompanyLocation.objects.filter(subscription_plan=plan).count()
        
        total_count = customer_count + customer_location_count + company_location_count
        
        if total_count > 0:
            plan_growth.append({
                'name': plan.name,
                'count': total_count
            })
    
    # STATE-WISE DISTRIBUTION (Total counts)
    state_distribution = list(Customer.objects.values('state__name').annotate(
        customer_count=Count('id')
    ).order_by('-customer_count')[:10])
    
    # ============== PAGE 1: HEADER ==============
    
    try:
        organization = Organization.objects.first()
        org_name = organization.name if organization else "Organization Name"
    except:
        org_name = "Organization Name"
    
    org_name_style = ParagraphStyle(
        'OrgName',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#374151'),
        fontName='Helvetica-Bold',
        alignment=TA_CENTER,
        spaceAfter=3
    )
    
    elements.append(Paragraph(org_name, org_name_style))
    elements.append(Paragraph('Subscription Analytics Report', title_style))
    date_text = f'Period: {start_date.strftime("%d/%m/%Y")} to {end_date.strftime("%d/%m/%Y")}'
    elements.append(Paragraph(date_text, subtitle_style))
    
    # ============== COLORED KPI CARDS ==============
    
    kpi_card_width = 82
    kpi_card_height = 45
    kpi_gap = 5
    
    kpi_data = [
        ('Companies', str(total_companies), '#3B82F6'),
        ('Total Customers', str(total_customers), '#10B981'),
        ('Active', str(active_customers), '#8B5CF6'),
        ('Closed', str(closed_customers), '#EF4444'),
        ('Total Invoices', str(total_invoices), '#F59E0B'),
        ('Total Amount', f'₹{total_amount:,.0f}', '#06B6D4'),
        ('Collected', f'₹{total_collected:,.0f}', '#14B8A6'),
        ('Outstanding', f'₹{total_outstanding:,.0f}', '#F97316'),
        ('Advance', f'₹{advance_payment:,.0f}', '#6366F1'),
    ]
    
    total_width = (kpi_card_width + kpi_gap) * 9 - kpi_gap
    kpi_drawing = Drawing(total_width, kpi_card_height)
    
    for i, (label, value, color) in enumerate(kpi_data):
        x_pos = i * (kpi_card_width + kpi_gap)
        
        card_rect = Rect(x_pos, 0, kpi_card_width, kpi_card_height, 
                        fillColor=colors.HexColor(color),
                        strokeColor=None,
                        rx=3, ry=3)
        kpi_drawing.add(card_rect)
        
        label_text = String(x_pos + kpi_card_width/2, kpi_card_height - 12, label,
                           fontSize=8,
                           fillColor=colors.white,
                           textAnchor='middle',
                           fontName='Helvetica')
        kpi_drawing.add(label_text)
        
        value_text = String(x_pos + kpi_card_width/2, 12, value,
                           fontSize=12,
                           fillColor=colors.white,
                           textAnchor='middle',
                           fontName='Helvetica-Bold')
        kpi_drawing.add(value_text)
    
    elements.append(kpi_drawing)
    elements.append(Spacer(1, 15))
    
    # ============== CHARTS SECTION - 2x2 GRID ==============
    
    chart_width = 350
    chart_height = 180
    
    # ===== CHART 1: Revenue vs Collection Trends =====
    chart1 = Drawing(chart_width, chart_height)
    
    chart1.add(String(10, chart_height - 15, 'Revenue vs Collection Trends',
                     fontSize=9, fillColor=colors.HexColor('#1F2937'),
                     fontName='Helvetica-Bold'))
    
    if monthly_revenue and len(monthly_revenue) > 0:
        lc = HorizontalLineChart()
        lc.x = 30
        lc.y = 20
        lc.height = 120
        lc.width = 300
        
        revenue_data = [float(item['revenue']) for item in monthly_revenue]
        collection_dict = {item['month']: float(item['collected']) for item in monthly_collection}
        collection_data = [collection_dict.get(item['month'], 0) for item in monthly_revenue]
        
        lc.data = [revenue_data, collection_data]
        
        lc.lines[0].strokeColor = colors.HexColor('#3B82F6')
        lc.lines[0].strokeWidth = 2
        lc.lines[1].strokeColor = colors.HexColor('#10B981')
        lc.lines[1].strokeWidth = 2
        
        lc.categoryAxis.labels.fontSize = 6
        lc.categoryAxis.labels.fillColor = colors.HexColor('#6B7280')
        lc.categoryAxis.categoryNames = [item['month'].strftime('%b %y') for item in monthly_revenue]
        lc.categoryAxis.strokeColor = colors.HexColor('#E5E7EB')
        lc.categoryAxis.strokeWidth = 0.5
        
        lc.valueAxis.valueMin = 0
        max_val = max(max(revenue_data) if revenue_data else 0, 
                     max(collection_data) if collection_data else 0)
        lc.valueAxis.valueMax = max_val * 1.2 if max_val > 0 else 1000
        lc.valueAxis.labels.fontSize = 6
        lc.valueAxis.labels.fillColor = colors.HexColor('#6B7280')
        lc.valueAxis.strokeColor = colors.HexColor('#E5E7EB')
        lc.valueAxis.strokeWidth = 0.5
        lc.valueAxis.gridStrokeColor = colors.HexColor('#F3F4F6')
        lc.valueAxis.gridStrokeWidth = 0.5
        
        chart1.add(lc)
        
        # Legend
        chart1.add(Line(230, chart_height - 30, 240, chart_height - 30, 
                       strokeColor=colors.HexColor('#3B82F6'), strokeWidth=2))
        chart1.add(String(243, chart_height - 32, 'Revenue',
                         fontSize=6, fillColor=colors.HexColor('#6B7280')))
        
        chart1.add(Line(230, chart_height - 42, 240, chart_height - 42, 
                       strokeColor=colors.HexColor('#10B981'), strokeWidth=2))
        chart1.add(String(243, chart_height - 44, 'Collection',
                         fontSize=6, fillColor=colors.HexColor('#6B7280')))
    else:
        chart1.add(String(chart_width/2, chart_height/2, 'No data available',
                         fontSize=9, fillColor=colors.HexColor('#9CA3AF'),
                         textAnchor='middle'))
    
    # ===== CHART 2: Payment Status =====
    chart2 = Drawing(chart_width, chart_height)
    
    chart2.add(String(10, chart_height - 15, 'Payment Status',
                     fontSize=9, fillColor=colors.HexColor('#1F2937'),
                     fontName='Helvetica-Bold'))
    
    status_data = [paid_invoices, unpaid_invoices]
    status_labels = ['Paid', 'Unpaid']
    status_colors = [colors.HexColor('#10B981'), colors.HexColor('#EF4444')]
    
    if sum(status_data) > 0:
        pc = Pie()
        pc.x = 110
        pc.y = 35
        pc.width = 90
        pc.height = 90
        pc.data = status_data
        pc.labels = []
        pc.slices.strokeWidth = 0
        
        for i, color in enumerate(status_colors):
            if i < len(status_data):
                pc.slices[i].fillColor = color
        
        chart2.add(pc)
        
        for i, (label, count) in enumerate(zip(status_labels, status_data)):
            y = 115 - (i * 18)
            chart2.add(Line(230, y + 4, 240, y + 4, strokeColor=status_colors[i], strokeWidth=3))
            chart2.add(String(243, y + 2, f'{label}: {count}',
                            fontSize=7, fillColor=colors.HexColor('#6B7280')))
    else:
        chart2.add(String(chart_width/2, chart_height/2, 'No data available',
                         fontSize=9, fillColor=colors.HexColor('#9CA3AF'),
                         textAnchor='middle'))
    
    # First row of charts
    row1 = Table([[chart1, chart2]], colWidths=[chart_width + 5, chart_width + 5])
    row1.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(row1)
    elements.append(Spacer(1, 10))
    
    # ===== CHART 3: Subscription Plan Distribution =====
    chart3 = Drawing(chart_width, chart_height)
    
    chart3.add(String(10, chart_height - 15, 'Subscription Plan Distribution',
                     fontSize=9, fillColor=colors.HexColor('#1F2937'),
                     fontName='Helvetica-Bold'))
    
    if plan_growth:
        bc = VerticalBarChart()
        bc.x = 30
        bc.y = 20
        bc.height = 120
        bc.width = 280
        bc.data = [[p['count'] for p in plan_growth]]
        bc.bars[0].fillColor = colors.HexColor('#8B5CF6')
        bc.strokeColor = None
        
        bc.categoryAxis.labels.fontSize = 6
        bc.categoryAxis.labels.fillColor = colors.HexColor('#6B7280')
        bc.categoryAxis.labels.angle = 20
        bc.categoryAxis.labels.boxAnchor = 'e'
        bc.categoryAxis.categoryNames = [p['name'][:20] for p in plan_growth]
        bc.categoryAxis.strokeColor = colors.HexColor('#E5E7EB')
        
        bc.valueAxis.valueMin = 0
        bc.valueAxis.labels.fontSize = 6
        bc.valueAxis.labels.fillColor = colors.HexColor('#6B7280')
        bc.valueAxis.strokeColor = colors.HexColor('#E5E7EB')
        bc.valueAxis.gridStrokeColor = colors.HexColor('#F3F4F6')
        bc.valueAxis.gridStrokeWidth = 0.5
        
        chart3.add(bc)
    else:
        chart3.add(String(chart_width/2, chart_height/2, 'No Subscription Data',
                         fontSize=9, fillColor=colors.HexColor('#9CA3AF'),
                         textAnchor='middle'))
    
    # ===== CHART 4: Payment Methods =====
    chart4 = Drawing(chart_width, chart_height)
    
    chart4.add(String(10, chart_height - 15, 'Payment Methods',
                     fontSize=9, fillColor=colors.HexColor('#1F2937'),
                     fontName='Helvetica-Bold'))
    
    if payment_methods:
        method_data = [float(m['amount']) for m in payment_methods if m['payment_mode']]
        method_labels = [m['payment_mode'].upper() if m['payment_mode'] else 'Unknown' 
                        for m in payment_methods if m['payment_mode']]
        
        if method_data:
            pc2 = Pie()
            pc2.x = 110
            pc2.y = 35
            pc2.width = 90
            pc2.height = 90
            pc2.data = method_data[:5]
            pc2.labels = []
            pc2.slices.strokeWidth = 0
            
            method_colors = [
                colors.HexColor('#3B82F6'),
                colors.HexColor('#10B981'),
                colors.HexColor('#F59E0B'),
                colors.HexColor('#EF4444'),
                colors.HexColor('#8B5CF6'),
            ]
            
            for i, color in enumerate(method_colors[:len(method_data)]):
                pc2.slices[i].fillColor = color
            
            chart4.add(pc2)
            
            for i, (label, amount) in enumerate(zip(method_labels[:5], method_data[:5])):
                y = 135 - (i * 16)
                chart4.add(Line(230, y + 4, 240, y + 4, strokeColor=method_colors[i], strokeWidth=3))
                chart4.add(String(243, y + 2, f'{label}',
                                fontSize=6, fillColor=colors.HexColor('#6B7280')))
    else:
        chart4.add(String(chart_width/2, chart_height/2, 'No data available',
                         fontSize=9, fillColor=colors.HexColor('#9CA3AF'),
                         textAnchor='middle'))
    
    # Second row of charts
    row2 = Table([[chart3, chart4]], colWidths=[chart_width + 5, chart_width + 5])
    row2.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(row2)
    
    # ============== PAGE BREAK ==============
    elements.append(PageBreak())
    
    # ============== PAGE 2: DETAILED TABLES ==============
    elements.append(Paragraph('Detailed Breakdown', title_style))
    elements.append(Spacer(1, 10))
    
    left_column = []
    right_column = []
    
    # ===== TABLE 1: Monthly Revenue Breakdown =====
    if monthly_revenue:
        left_column.append(Paragraph('Monthly Revenue Breakdown', section_title_style))
        
        table_data = [['Month', 'Revenue', 'Invoices', 'Avg']]
        
        for item in monthly_revenue:
            revenue = float(item['revenue'])
            invoice_count = BillingRecord.objects.filter(
                billing_date__month=item['month'].month,
                billing_date__year=item['month'].year,
                billing_date__gte=start_date,
                billing_date__lte=end_date
            ).count()
            avg = revenue / invoice_count if invoice_count > 0 else 0
            
            table_data.append([
                item['month'].strftime('%b %Y'),
                f"₹{revenue:,.0f}",
                str(invoice_count),
                f"₹{avg:,.0f}"
            ])
        
        monthly_table = Table(table_data, colWidths=[0.9*inch, 0.9*inch, 0.75*inch, 0.85*inch])
        monthly_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F9FAFB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#374151')),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#E5E7EB')),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        left_column.append(monthly_table)
        left_column.append(Spacer(1, 15))
    
    # ===== TABLE 2: State-wise Distribution =====
    if state_distribution:
        left_column.append(Paragraph('State-wise Customer Distribution', section_title_style))
        
        state_table_data = [['State', 'Customers']]
        
        for state in state_distribution:
            state_table_data.append([
                state['state__name'] if state['state__name'] else 'Not Specified',
                str(state['customer_count'])
            ])
        
        state_table = Table(state_table_data, colWidths=[2.5*inch, 1*inch])
        state_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F9FAFB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#374151')),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#E5E7EB')),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        left_column.append(state_table)
    
    # ===== TABLE 3: Top Customers =====
    if top_customers:
        right_column.append(Paragraph('Top Customers by Revenue', section_title_style))
        customer_table_data = [['Customer', 'Revenue', 'Count', 'Avg']]
        
        for cust in top_customers:
            revenue = float(cust['revenue'])
            count = cust['invoice_count']
            avg = revenue / count if count > 0 else 0
            customer_table_data.append([
                cust['customer__name'][:20],
                f"₹{revenue:,.0f}",
                str(count),
                f"₹{avg:,.0f}"
            ])
        
        customer_table = Table(customer_table_data, colWidths=[1.3*inch, 0.85*inch, 0.6*inch, 0.65*inch])
        customer_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F9FAFB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#374151')),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#E5E7EB')),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        right_column.append(customer_table)
        right_column.append(Spacer(1, 15))
    
    # ===== TABLE 4: Recently Added (filtered by start_date) =====
    recent_customers = Customer.objects.filter(
        start_date__gte=start_date,
        start_date__lte=end_date
    ).select_related('state', 'subscription_plan').order_by('-start_date')[:10]
    
    recent_companies = Company.objects.filter(
        start_date__gte=start_date,
        start_date__lte=end_date
    ).select_related('state').order_by('-start_date')[:10]
    
    right_column.append(Paragraph('Recently Added', section_title_style))
    
    recent_table_data = [['Name', 'Type', 'Date']]
    
    for customer in recent_customers:
        recent_table_data.append([
            customer.name[:20],
            'Customer',
            customer.start_date.strftime('%d/%m/%Y') if customer.start_date else 'N/A'
        ])
    
    for company in recent_companies:
        recent_table_data.append([
            company.name[:20],
            'Company',
            company.start_date.strftime('%d/%m/%Y') if company.start_date else 'N/A'
        ])
    
    if len(recent_table_data) > 1:
        recent_table = Table(recent_table_data, colWidths=[1.8*inch, 0.8*inch, 0.8*inch])
        recent_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F9FAFB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#374151')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#E5E7EB')),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        right_column.append(recent_table)
    
    # Two-column layout
    available_width = landscape(A4)[0] - 60
    col_width = available_width / 2 - 15
    
    two_column_table = Table(
        [[left_column, right_column]], 
        colWidths=[col_width, col_width],
        hAlign='LEFT'
    )
    two_column_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (0, -1), 0),
        ('RIGHTPADDING', (0, 0), (0, -1), 15),
        ('LEFTPADDING', (1, 0), (1, -1), 15),
        ('RIGHTPADDING', (1, 0), (1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    
    elements.append(two_column_table)
    
    # Footer
    elements.append(Spacer(1, 15))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#9CA3AF'),
        alignment=TA_CENTER
    )
    elements.append(Paragraph(f'{org_name}', footer_style))
    elements.append(Paragraph('This report is confidential and intended for internal use only', footer_style))
    
    # Build PDF
    doc.build(elements)
    return response


def subscription_analytics_form(request):
    """Form to select date range and generate report"""
    
    if request.method == 'POST':
        # Generate report with selected dates
        return subscription_analytics_report(request)
    
    # Default dates for the form
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=365)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'analytics.html', context)

from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Count, Q, F, Avg, DecimalField, Case, When, Value
from django.db.models.functions import TruncMonth, TruncYear, TruncDate, TruncWeek
from datetime import datetime, timedelta
from decimal import Decimal

# ReportLab imports for PDF
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, 
    Spacer, PageBreak, Image as RLImage
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.graphics.shapes import Drawing, Rect, String, Line
from reportlab.graphics.charts.barcharts import VerticalBarChart, HorizontalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from io import BytesIO

# Import your models
from .models import (
    BillingRecord, Customer, CustomerLocation, SubscriptionPlan,
    State, City, Area, BillItem, data_logs, Company, Organization
)
from django.shortcuts import render, redirect
from django.db.models import Sum, Count, Q, F, Avg, DecimalField, Case, When
from django.db.models.functions import TruncMonth
from datetime import datetime, timedelta
from decimal import Decimal
from .models import Customer, Company, BillingRecord, State, City, SubscriptionPlan, BillItem
from .utils import get_filtered_companies, get_company_filter_context
def financial_sales_report_view(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    
    # Get filter parameters from URL
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    customer_id = request.GET.get('customer')
    location_id = request.GET.get('location')
    state_id = request.GET.get('state')
    city_id = request.GET.get('city')
    payment_status = request.GET.get('payment_status', 'all')
    subscription_plan_id = request.GET.get('subscription_plan')
    
    # ============= APPLY COMPANY FILTER FROM SESSION =============
    filtered_companies = get_filtered_companies(request)
    company_ids = list(filtered_companies.values_list('id', flat=True))
    
    # Check if filters are actively being used
    selected_gst_type = request.session.get('selected_gst_type', 'gst')
    selected_company_ids = request.session.get('selected_company_ids', [])
    filters_active = selected_gst_type or selected_company_ids
    
    # Base queryset - filter by company first
    if company_ids:
        billing_records = BillingRecord.objects.filter(
            Q(customer__company_id__in=company_ids) | 
            Q(customer_location__customer__company_id__in=company_ids)
        )
        customer_filter = Q(company_id__in=company_ids)
    elif filters_active:
        # ✅ Filters are active but returned no companies - show no data
        billing_records = BillingRecord.objects.none()
        customer_filter = Q(pk__in=[])
    else:
        # No filters applied - show all data
        billing_records = BillingRecord.objects.all()
        customer_filter = Q()
    
    # Date range filters (default to last 30 days)
    if not start_date:
        start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not end_date:
        end_date = datetime.now().strftime('%Y-%m-%d')
    
    billing_records = billing_records.filter(
        billing_date__gte=start_date,
        billing_date__lte=end_date
    )
    
    # Apply other filters
    if customer_id:
        billing_records = billing_records.filter(
            Q(customer_id=customer_id) | 
            Q(customer_location__customer_id=customer_id)
        )
        customer_filter = customer_filter & Q(id=customer_id)
    
    if location_id:
        billing_records = billing_records.filter(customer_location_id=location_id)
    
    if state_id:
        billing_records = billing_records.filter(
            Q(customer_location__state_id=state_id) |
            Q(customer__state_id=state_id)
        )
        customer_filter = customer_filter & Q(state_id=state_id)
    
    if city_id:
        billing_records = billing_records.filter(
            Q(customer_location__city_id=city_id) |
            Q(customer__city_id=city_id)
        )
    
    if payment_status == 'paid':
        billing_records = billing_records.filter(paid=True)
    elif payment_status == 'unpaid':
        billing_records = billing_records.filter(paid=False)
    
    if subscription_plan_id:
        billing_records = billing_records.filter(
            Q(customer_location__subscription_plan_id=subscription_plan_id) |
            Q(customer__subscription_plan_id=subscription_plan_id)
        )
    
    # ============= CALCULATE ADVANCE AMOUNT =============
    total_advance = Customer.objects.filter(
        customer_filter,
        advance_amount__isnull=False
    ).aggregate(
        total=Sum('advance_amount')
    )['total'] or Decimal('0')
    
    customers_with_advance = Customer.objects.filter(
        customer_filter,
        advance_amount__isnull=False,
        advance_amount__gt=0
    ).values('id', 'name', 'advance_amount').order_by('-advance_amount')[:10]
    
    total_customers_with_advance = Customer.objects.filter(
        customer_filter,
        advance_amount__isnull=False,
        advance_amount__gt=0
    ).count()
    
    avg_advance_per_customer = (total_advance / total_customers_with_advance) if total_customers_with_advance > 0 else Decimal('0')
    
    # ============= CALCULATE NEW CUSTOMERS/LOCATIONS ADDED =============
    
    # Count customers added in the date range
    new_customers = Customer.objects.filter(
        customer_filter,
        start_date__gte=start_date,
        start_date__lte=end_date
    ).count()
    
    # Count locations added in the date range
    if company_ids:
        new_locations = CustomerLocation.objects.filter(
            customer__company_id__in=company_ids,
            start_date__gte=start_date,
            start_date__lte=end_date
        ).count()
    elif filters_active:
        # ✅ Filters active but no companies matched
        new_locations = 0
    else:
        new_locations = CustomerLocation.objects.filter(
            start_date__gte=start_date,
            start_date__lte=end_date
        ).count()
    
    # Total new acquisitions
    total_new_acquisitions = new_customers + new_locations
    
    # Get breakdown details
    new_customers_list = Customer.objects.filter(
        customer_filter,
        start_date__gte=start_date,
        start_date__lte=end_date
    ).select_related('subscription_plan').values(
        'id', 'name', 'start_date', 'subscription_plan__name', 'email'
    ).order_by('-start_date')[:20]
    
    if company_ids:
        new_locations_list = CustomerLocation.objects.filter(
            customer__company_id__in=company_ids,
            start_date__gte=start_date,
            start_date__lte=end_date
        ).select_related('customer', 'subscription_plan').values(
            'id', 'location_name', 'customer__name', 'start_date', 'subscription_plan__name', 'location_email'
        ).order_by('-start_date')[:20]
    elif filters_active:
        # ✅ Filters active but no companies matched
        new_locations_list = []
    else:
        new_locations_list = CustomerLocation.objects.filter(
            start_date__gte=start_date,
            start_date__lte=end_date
        ).select_related('customer', 'subscription_plan').values(
            'id', 'location_name', 'customer__name', 'start_date', 'subscription_plan__name', 'location_email'
        ).order_by('-start_date')[:20]
    
    # ============= CALCULATE SUMMARY METRICS =============
    
    summary = billing_records.aggregate(
        total_invoices=Count('id'),
        total_amount=Sum('amount'),
        total_discount=Sum('discount_amount'),
        avg_invoice_value=Avg('amount')
    )
    
    paid_summary = billing_records.filter(paid=True).aggregate(
        total_paid_amount=Sum('amount')
    )
    
    total_invoices = summary['total_invoices'] or 0
    total_amount = summary['total_amount'] or Decimal('0')
    total_paid_amount = paid_summary['total_paid_amount'] or Decimal('0')
    total_discount = summary['total_discount'] or Decimal('0')
    avg_invoice_value = summary['avg_invoice_value'] or Decimal('0')
    
    total_outstanding = total_amount - total_paid_amount
    
    paid_count = billing_records.filter(paid=True).count()
    unpaid_count = billing_records.filter(paid=False).count()
    
    collection_rate = (total_paid_amount / total_amount * 100) if total_amount > 0 else 0
    
    # Calculate total cash flow
    total_cash_flow = total_paid_amount + total_advance
    
    total_tax = BillItem.objects.filter(
        billing_record__in=billing_records
    ).aggregate(total=Sum('tax_amount'))['total'] or Decimal('0')
    
    total_subtotal = BillItem.objects.filter(
        billing_record__in=billing_records
    ).aggregate(total=Sum('total_price'))['total'] or Decimal('0')
    
    # ============= MONTHLY BREAKDOWN =============
    
    monthly_data = billing_records.annotate(
        month=TruncMonth('billing_date')
    ).values('month').annotate(
        total_revenue=Sum('amount'),
        paid_amount=Sum(Case(
            When(paid=True, then=F('amount')),
            default=0,
            output_field=DecimalField()
        )),
        invoice_count=Count('id'),
        paid_count=Count(Case(When(paid=True, then=1))),
        unpaid_count=Count(Case(When(paid=False, then=1)))
    ).order_by('month')
    
    monthly_data_list = list(monthly_data)
    for item in monthly_data_list:
        total_rev = item.get('total_revenue') or Decimal('0')
        paid_amt = item.get('paid_amount') or Decimal('0')
        item['outstanding'] = total_rev - paid_amt
    
    # ============= PAYMENT MODE BREAKDOWN =============
    
    payment_mode_data = billing_records.filter(
        paid=True,
        payment_mode__isnull=False
    ).values('payment_mode').annotate(
        total_amount=Sum('amount'),
        count=Count('id')
    ).order_by('-total_amount')
    
    payment_mode_list = []
    for mode in payment_mode_data:
        mode_dict = dict(mode)
        mode_dict['average'] = mode['total_amount'] / mode['count'] if mode['count'] > 0 else Decimal('0')
        payment_mode_list.append(mode_dict)
    
    # ============= CUSTOMER ANALYSIS =============
    
    customer_revenue_data = {}
    
    for record in billing_records.select_related('customer', 'customer_location__customer'):
        if record.customer_location:
            actual_customer = record.customer_location.customer
        elif record.customer:
            actual_customer = record.customer
        else:
            continue
        
        if actual_customer.id not in customer_revenue_data:
            customer_revenue_data[actual_customer.id] = {
                'customer__id': actual_customer.id,
                'customer__name': actual_customer.name,
                'total_revenue': Decimal('0'),
                'paid_amount': Decimal('0'),
                'outstanding': Decimal('0'),
                'invoice_count': 0
            }
        
        customer_revenue_data[actual_customer.id]['total_revenue'] += record.amount
        customer_revenue_data[actual_customer.id]['invoice_count'] += 1
        
        if record.paid:
            customer_revenue_data[actual_customer.id]['paid_amount'] += record.amount
        else:
            customer_revenue_data[actual_customer.id]['outstanding'] += record.amount
    
    top_customers = sorted(
        customer_revenue_data.values(),
        key=lambda x: x['total_revenue'],
        reverse=True
    )[:10]
    
    top_customers_list = list(top_customers)
    for customer in top_customers_list:
        customer_obj = Customer.objects.filter(id=customer['customer__id']).first()
        customer['advance_amount'] = customer_obj.advance_amount if customer_obj and customer_obj.advance_amount else Decimal('0')
    
    # Customer outstanding
    customer_outstanding_data = {}
    
    for record in billing_records.filter(paid=False).select_related('customer', 'customer_location__customer'):
        if record.customer_location:
            actual_customer = record.customer_location.customer
        elif record.customer:
            actual_customer = record.customer
        else:
            continue
        
        if actual_customer.id not in customer_outstanding_data:
            customer_outstanding_data[actual_customer.id] = {
                'customer__id': actual_customer.id,
                'customer__name': actual_customer.name,
                'outstanding_amount': Decimal('0'),
                'overdue_count': 0
            }
        
        customer_outstanding_data[actual_customer.id]['outstanding_amount'] += record.amount
        customer_outstanding_data[actual_customer.id]['overdue_count'] += 1
    
    customer_outstanding = sorted(
        customer_outstanding_data.values(),
        key=lambda x: x['outstanding_amount'],
        reverse=True
    )[:10]
    
    # ============= LOCATION ANALYSIS =============
    
    location_revenue = billing_records.filter(
        customer_location__isnull=False
    ).values(
        'customer_location__id',
        'customer_location__location_name',
        'customer_location__customer__name'
    ).annotate(
        total_revenue=Sum('amount'),
        invoice_count=Count('id')
    ).order_by('-total_revenue')[:10]
    
    location_revenue_list = []
    for loc in location_revenue:
        location_revenue_list.append({
            'customer_location__id': loc['customer_location__id'],
            'customer_location__location_name': loc['customer_location__location_name'],
            'customer__name': loc['customer_location__customer__name'],
            'total_revenue': loc['total_revenue'],
            'invoice_count': loc['invoice_count']
        })
    
    # ============= STATE/CITY WISE BREAKDOWN =============
    
    state_revenue_data = {}
    
    for record in billing_records.select_related('customer__state', 'customer_location__state'):
        state = None
        if record.customer_location and record.customer_location.state:
            state = record.customer_location.state
        elif record.customer and record.customer.state:
            state = record.customer.state
        
        if state:
            state_name = state.name
            if state_name not in state_revenue_data:
                state_revenue_data[state_name] = {
                    'customer_location__state__name': state_name,
                    'total_revenue': Decimal('0'),
                    'customer_ids': set()
                }
            
            state_revenue_data[state_name]['total_revenue'] += record.amount
            if record.customer_location:
                state_revenue_data[state_name]['customer_ids'].add(record.customer_location.customer.id)
            elif record.customer:
                state_revenue_data[state_name]['customer_ids'].add(record.customer.id)
    
    state_revenue = []
    for state_name, data in state_revenue_data.items():
        state_revenue.append({
            'customer_location__state__name': state_name,
            'total_revenue': data['total_revenue'],
            'customer_count': len(data['customer_ids'])
        })
    
    state_revenue = sorted(state_revenue, key=lambda x: x['total_revenue'], reverse=True)
    
    city_revenue_data = {}
    
    for record in billing_records.select_related('customer__city', 'customer__state', 'customer_location__city', 'customer_location__state'):
        city = None
        state = None
        
        if record.customer_location:
            city = record.customer_location.city
            state = record.customer_location.state
        elif record.customer:
            city = record.customer.city
            state = record.customer.state
        
        if city:
            city_name = city.name
            state_name = state.name if state else 'Unknown'
            key = f"{city_name}_{state_name}"
            
            if key not in city_revenue_data:
                city_revenue_data[key] = {
                    'customer_location__city__name': city_name,
                    'customer_location__state__name': state_name,
                    'total_revenue': Decimal('0')
                }
            
            city_revenue_data[key]['total_revenue'] += record.amount
    
    city_revenue = sorted(
        city_revenue_data.values(),
        key=lambda x: x['total_revenue'],
        reverse=True
    )[:10]
    
    # ============= SUBSCRIPTION PLAN ANALYSIS =============
    
    plan_revenue_data = {}
    
    for record in billing_records.select_related('customer__subscription_plan', 'customer_location__subscription_plan', 'customer_location__customer'):
        plan = None
        location_id = None
        
        if record.customer_location and record.customer_location.subscription_plan:
            plan = record.customer_location.subscription_plan
            location_id = record.customer_location.id
        elif record.customer and record.customer.subscription_plan:
            plan = record.customer.subscription_plan
            location_id = None
        
        if plan:
            if plan.id not in plan_revenue_data:
                plan_revenue_data[plan.id] = {
                    'customer_location__subscription_plan__id': plan.id,
                    'customer_location__subscription_plan__name': plan.name,
                    'customer_location__subscription_plan__price': plan.price,
                    'total_revenue': Decimal('0'),
                    'location_ids': set()
                }
            
            plan_revenue_data[plan.id]['total_revenue'] += record.amount
            if location_id:
                plan_revenue_data[plan.id]['location_ids'].add(location_id)
    
    plan_revenue = []
    for plan_id, data in plan_revenue_data.items():
        plan_revenue.append({
            'customer_location__subscription_plan__id': data['customer_location__subscription_plan__id'],
            'customer_location__subscription_plan__name': data['customer_location__subscription_plan__name'],
            'customer_location__subscription_plan__price': data['customer_location__subscription_plan__price'],
            'total_revenue': data['total_revenue'],
            'location_count': len(data['location_ids'])
        })
    
    plan_revenue = sorted(plan_revenue, key=lambda x: x['total_revenue'], reverse=True)
    
    # ============= AGING ANALYSIS =============
    
    today = datetime.now().date()
    
    aging_0_30 = billing_records.filter(
        paid=False,
        due_date__lt=today,
        due_date__gte=today - timedelta(days=30)
    ).aggregate(amount=Sum('amount'), count=Count('id'))
    
    aging_31_60 = billing_records.filter(
        paid=False,
        due_date__lt=today - timedelta(days=30),
        due_date__gte=today - timedelta(days=60)
    ).aggregate(amount=Sum('amount'), count=Count('id'))
    
    aging_61_90 = billing_records.filter(
        paid=False,
        due_date__lt=today - timedelta(days=60),
        due_date__gte=today - timedelta(days=90)
    ).aggregate(amount=Sum('amount'), count=Count('id'))
    
    aging_90_plus = billing_records.filter(
        paid=False,
        due_date__lt=today - timedelta(days=90)
    ).aggregate(amount=Sum('amount'), count=Count('id'))
    
    not_due = billing_records.filter(
        paid=False,
        due_date__gte=today
    ).aggregate(amount=Sum('amount'), count=Count('id'))
    
    # ============= ITEM ANALYSIS =============
    
    top_items = BillItem.objects.filter(
        billing_record__in=billing_records
    ).values('item_name').annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum(F('quantity') * F('unit_price')),
        times_billed=Count('id')
    ).order_by('-total_revenue')[:10]
    
    # ============= DETAILED RECORDS =============
    
    detailed_records = billing_records.select_related(
        'customer',
        'customer_location',
        'customer_location__customer',
        'customer_location__state',
        'customer_location__city',
        'customer__state',
        'customer__city'
    ).order_by('-billing_date')[:100]
    
    # ============= FILTER OPTIONS =============
    
    if company_ids:
        customers_for_filter = Customer.objects.filter(
            company_id__in=company_ids
        ).order_by('name')
    elif filters_active:
        # ✅ Filters active but no companies matched
        customers_for_filter = Customer.objects.none()
    else:
        customers_for_filter = Customer.objects.all().order_by('name')
    
    # ============= CONTEXT =============
    
    context = {
        # Filters
        'start_date': start_date,
        'end_date': end_date,
        'selected_customer': customer_id,
        'selected_location': location_id,
        'selected_state': state_id,
        'selected_city': city_id,
        'payment_status': payment_status,
        'selected_plan': subscription_plan_id,
        
        # Filter options
        'customers': customers_for_filter,
        'states': State.objects.all().order_by('name'),
        'cities': City.objects.all().order_by('name'),
        'subscription_plans': SubscriptionPlan.objects.all().order_by('name'),
        
        # Summary metrics
        'total_invoices': total_invoices,
        'total_amount': total_amount,
        'total_paid_amount': total_paid_amount,
        'total_outstanding': total_outstanding,
        'total_discount': total_discount,
        'total_tax': total_tax,
        'total_subtotal': total_subtotal,
        'avg_invoice_value': avg_invoice_value,
        'paid_count': paid_count,
        'unpaid_count': unpaid_count,
        'collection_rate': round(collection_rate, 2),
        
        # Advance amount metrics
        'total_advance': total_advance,
        'total_cash_flow': total_cash_flow,
        'customers_with_advance': list(customers_with_advance),
        'total_customers_with_advance': total_customers_with_advance,
        'avg_advance_per_customer': avg_advance_per_customer,
        
        # New acquisitions metrics
        'new_customers': new_customers,
        'new_locations': new_locations,
        'total_new_acquisitions': total_new_acquisitions,
        'new_customers_list': list(new_customers_list),
        'new_locations_list': list(new_locations_list),
        
        # Breakdowns
        'monthly_data': monthly_data_list,
        'payment_mode_data': payment_mode_list,
        'top_customers': top_customers_list,
        'customer_outstanding': customer_outstanding,
        'location_revenue': location_revenue_list,
        'state_revenue': state_revenue,
        'city_revenue': city_revenue,
        'plan_revenue': plan_revenue,
        'top_items': list(top_items),
        
        # Aging
        'aging_0_30': aging_0_30,
        'aging_31_60': aging_31_60,
        'aging_61_90': aging_61_90,
        'aging_90_plus': aging_90_plus,
        'not_due': not_due,
        
        # Details
        'detailed_records': detailed_records,
        
        # Debug info
        'filtered_company_ids': company_ids,
        'filtered_companies_count': len(company_ids),
        'user': user,
        'role_permissions': role_permissions,
    }
    
    context.update(get_company_filter_context(request))
    
    return render(request, 'salesreport.html', context)


def export_financial_sales_pdf(request):
    """
    Export Financial Sales Report as PDF with professional formatting
    """
    # Get the same filters as the main view
    start_date = request.GET.get('start_date', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    customer_id = request.GET.get('customer')
    payment_status = request.GET.get('payment_status', 'all')
    
    # Get filtered billing records (same logic as main view)
    billing_records = BillingRecord.objects.filter(
        billing_date__gte=start_date,
        billing_date__lte=end_date
    )
    
    if customer_id:
        billing_records = billing_records.filter(customer_id=customer_id)
    
    if payment_status == 'paid':
        billing_records = billing_records.filter(paid=True)
    elif payment_status == 'unpaid':
        billing_records = billing_records.filter(paid=False)
    
    # Calculate metrics
    summary = billing_records.aggregate(
        total_invoices=Count('id'),
        total_amount=Sum('amount'),
        total_discount=Sum('discount_amount')
    )
    
    # Calculate paid amount separately
    paid_summary = billing_records.filter(paid=True).aggregate(
        total_paid_amount=Sum('amount')
    )
    
    total_amount = summary['total_amount'] or Decimal('0')
    total_paid = paid_summary['total_paid_amount'] or Decimal('0')
    total_outstanding = total_amount - total_paid
    
    # Monthly data
    monthly_data = billing_records.annotate(
        month=TruncMonth('billing_date')
    ).values('month').annotate(
        revenue=Sum('amount'),
        count=Count('id')
    ).order_by('month')
    
    # Top customers
    top_customers = billing_records.values(
        'customer__name'
    ).annotate(
        revenue=Sum('amount'),
        count=Count('id')
    ).order_by('-revenue')[:10]
    
    # Create PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="financial_sales_report_{start_date}_to_{end_date}.pdf"'
    
    # Use A4 portrait
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=22,
        textColor=colors.HexColor('#1F2937'),
        spaceAfter=10,
        spaceBefore=0,
        fontName='Helvetica-Bold',
        alignment=TA_CENTER
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#6B7280'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    section_title_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#374151'),
        spaceAfter=10,
        spaceBefore=15,
        fontName='Helvetica-Bold'
    )
    
    # Get organization details
    try:
        organization = Organization.objects.first()
        org_name = organization.name if organization else "Your Organization"
    except:
        org_name = "Your Organization"
    
    # Title
    elements.append(Paragraph(org_name, title_style))
    elements.append(Paragraph('Financial Sales Report', title_style))
    date_range = f'Period: {datetime.strptime(start_date, "%Y-%m-%d").strftime("%d %b %Y")} - {datetime.strptime(end_date, "%Y-%m-%d").strftime("%d %b %Y")}'
    elements.append(Paragraph(date_range, subtitle_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # ===== SUMMARY KPI CARDS =====
    kpi_data = [
        ['Total Invoices', f"{summary['total_invoices']}", colors.HexColor('#3B82F6')],
        ['Total Amount', f"₹{total_amount:,.2f}", colors.HexColor('#10B981')],
        ['Collected', f"₹{total_paid:,.2f}", colors.HexColor('#8B5CF6')],
        ['Outstanding', f"₹{total_outstanding:,.2f}", colors.HexColor('#EF4444')]
    ]
    
    kpi_table_data = []
    kpi_row = []
    
    for label, value, color in kpi_data:
        cell_content = [
            Paragraph(f'<font name="Helvetica-Bold" size="10" color="#{color.hexval()[2:]}">{label}</font>', styles['Normal']),
            Paragraph(f'<font name="Helvetica-Bold" size="14">{value}</font>', styles['Normal'])
        ]
        kpi_row.append(cell_content)
    
    kpi_table_data.append(kpi_row)
    
    kpi_table = Table(kpi_table_data, colWidths=[1.2*inch] * 4)
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F9FAFB')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
        ('INNERGRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    
    elements.append(kpi_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # ===== MONTHLY REVENUE BREAKDOWN =====
    if monthly_data:
        elements.append(Paragraph('Monthly Revenue Breakdown', section_title_style))
        
        monthly_table_data = [['Month', 'Revenue', 'Invoices', 'Average']]
        
        for item in monthly_data:
            avg = float(item['revenue']) / item['count'] if item['count'] > 0 else 0
            monthly_table_data.append([
                item['month'].strftime('%B %Y'),
                f"₹{float(item['revenue']):,.2f}",
                str(item['count']),
                f"₹{avg:,.2f}"
            ])
        
        monthly_table = Table(monthly_table_data, colWidths=[1.5*inch, 1.3*inch, 1*inch, 1.3*inch])
        monthly_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor('#E5E7EB')),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        elements.append(monthly_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # ===== TOP CUSTOMERS =====
    if top_customers:
        elements.append(Paragraph('Top 10 Customers by Revenue', section_title_style))
        
        customer_table_data = [['Customer Name', 'Revenue', 'Invoices', 'Average']]
        
        for cust in top_customers:
            avg = float(cust['revenue']) / cust['count'] if cust['count'] > 0 else 0
            customer_table_data.append([
                cust['customer__name'][:30],
                f"₹{float(cust['revenue']):,.2f}",
                str(cust['count']),
                f"₹{avg:,.2f}"
            ])
        
        customer_table = Table(customer_table_data, colWidths=[2*inch, 1.3*inch, 0.8*inch, 1*inch])
        customer_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor('#E5E7EB')),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        elements.append(customer_table)
    
    # Footer
    elements.append(Spacer(1, 0.5*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#9CA3AF'),
        alignment=TA_CENTER
    )
    elements.append(Paragraph(f'Generated on {datetime.now().strftime("%d %B %Y at %I:%M %p")}', footer_style))
    elements.append(Paragraph('This report is confidential and for internal use only', footer_style))
    
    # Build PDF
    doc.build(elements)
    return response


def export_financial_sales_excel(request):
    """
    Export Financial Sales Report as Excel with multiple sheets
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Get filters
    start_date = request.GET.get('start_date', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    
    # Get data
    billing_records = BillingRecord.objects.filter(
        billing_date__gte=start_date,
        billing_date__lte=end_date
    )
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== SHEET 1: SUMMARY =====
    ws_summary = wb.create_sheet("Summary")
    
    # Title
    ws_summary['A1'] = "Financial Sales Report - Summary"
    ws_summary['A1'].font = title_font
    ws_summary['A2'] = f"Period: {start_date} to {end_date}"
    ws_summary.merge_cells('A1:D1')
    ws_summary.merge_cells('A2:D2')
    
    # Summary metrics
    summary = billing_records.aggregate(
        total_invoices=Count('id'),
        total_amount=Sum('amount'),
        total_discount=Sum('discount_amount')
    )
    
    # Calculate paid separately
    paid_summary = billing_records.filter(paid=True).aggregate(
        total_paid=Sum('amount')
    )
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Invoices', summary['total_invoices'] or 0],
        ['Total Amount', float(summary['total_amount'] or 0)],
        ['Total Collected', float(paid_summary['total_paid'] or 0)],
        ['Total Outstanding', float((summary['total_amount'] or 0) - (paid_summary['total_paid'] or 0))],
        ['Total Discount', float(summary['total_discount'] or 0)],
    ]
    
    start_row = 4
    for row_idx, row_data in enumerate(summary_data, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == start_row:
                cell.fill = header_fill
                cell.font = header_font
            cell.border = border
    
    # Auto-size columns
    for col in ws_summary.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_summary.column_dimensions[column].width = adjusted_width
    
    # ===== SHEET 2: MONTHLY BREAKDOWN =====
    ws_monthly = wb.create_sheet("Monthly Breakdown")
    
    monthly_data = list(billing_records.annotate(
        month=TruncMonth('billing_date')
    ).values('month').annotate(
        revenue=Sum('amount'),
        paid=Sum(Case(
            When(paid=True, then=F('amount')),
            default=0,
            output_field=DecimalField()
        )),
        count=Count('id')
    ).order_by('month'))
    
    ws_monthly['A1'] = "Monthly Revenue Breakdown"
    ws_monthly['A1'].font = title_font
    ws_monthly.merge_cells('A1:E1')
    
    headers = ['Month', 'Total Revenue', 'Collected', 'Outstanding', 'Invoice Count']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_monthly.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for row_idx, item in enumerate(monthly_data, start=4):
        ws_monthly.cell(row=row_idx, column=1, value=item['month'].strftime('%B %Y')).border = border
        ws_monthly.cell(row=row_idx, column=2, value=float(item['revenue'])).border = border
        ws_monthly.cell(row=row_idx, column=3, value=float(item['paid'])).border = border
        ws_monthly.cell(row=row_idx, column=4, value=float(item['revenue'] - item['paid'])).border = border
        ws_monthly.cell(row=row_idx, column=5, value=item['count']).border = border
    
    # ===== SHEET 3: DETAILED RECORDS =====
    ws_details = wb.create_sheet("Billing Details")
    
    ws_details['A1'] = "Detailed Billing Records"
    ws_details['A1'].font = title_font
    ws_details.merge_cells('A1:J1')
    
    detail_headers = [
        'Invoice Number', 'Customer', 'Location', 'Billing Date',
        'Due Date', 'Amount', 'Discount', 'Status', 'Payment Mode', 'Payment Date'
    ]
    
    for col_idx, header in enumerate(detail_headers, start=1):
        cell = ws_details.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    detailed_records = billing_records.select_related('customer', 'customer_location').order_by('-billing_date')
    
    for row_idx, record in enumerate(detailed_records, start=4):
        ws_details.cell(row=row_idx, column=1, value=record.invoice_number or '').border = border
        ws_details.cell(row=row_idx, column=2, value=record.customer.name if record.customer else '').border = border
        ws_details.cell(row=row_idx, column=3, value=record.customer_location.location_name if record.customer_location else '').border = border
        ws_details.cell(row=row_idx, column=4, value=record.billing_date.strftime('%Y-%m-%d')).border = border
        ws_details.cell(row=row_idx, column=5, value=record.due_date.strftime('%Y-%m-%d')).border = border
        ws_details.cell(row=row_idx, column=6, value=float(record.amount)).border = border
        ws_details.cell(row=row_idx, column=7, value=float(record.discount_amount)).border = border
        ws_details.cell(row=row_idx, column=8, value='Paid' if record.paid else 'Unpaid').border = border
        ws_details.cell(row=row_idx, column=9, value=record.payment_mode or '').border = border
        ws_details.cell(row=row_idx, column=10, value=record.payment_date.strftime('%Y-%m-%d') if record.payment_date else '').border = border
    
    # Auto-size all columns in all sheets
    for ws in wb.worksheets:
        for column_cells in ws.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="financial_sales_report_{start_date}_to_{end_date}.xlsx"'
    
    return response

import json
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from .models import Company

@require_http_methods(["GET"])
def get_companies(request):
    """
    API endpoint to get all companies with their GST status.
    """
    # Debug: Check session state when API is called
    gst_type = request.session.get('gst_type', '')
    selected_companies = request.session.get('selected_companies', [])
    
    companies = Company.objects.all().values('id', 'name', 'gst_registered')
    companies_list = list(companies)
    return JsonResponse({'companies': companies_list})
def create_billing(request):
    """Create a new billing record with items"""
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')

    try:
        user = Custom_User.objects.get(username=user)
    except Custom_User.DoesNotExist:
        return redirect('login')

    if request.method == 'POST':
        try:
            with transaction.atomic():

                # Basic Data
                customer_id          = request.POST.get('customer')
                customer_location_id = request.POST.get('customer_location')

                if not customer_id:
                    messages.error(request, 'Please select a customer.')
                    return redirect('create_billing')

                customer_location_id = customer_location_id if customer_location_id else None

                location = None
                if customer_id and customer_location_id:
                    try:
                        location = CustomerLocation.objects.get(
                            id=customer_location_id,
                            customer_id=customer_id
                        )
                    except CustomerLocation.DoesNotExist:
                        messages.error(
                            request,
                            'Selected location does not belong to the selected customer.'
                        )
                        return redirect('create_billing')

                # Items Data
                item_names   = request.POST.getlist('item_name[]')
                descriptions = request.POST.getlist('description[]')
                quantities   = request.POST.getlist('quantity[]')
                unit_prices  = request.POST.getlist('unit_price[]')

                if not item_names or not any(item_names):
                    messages.error(request, 'Please add at least one item to the bill.')
                    return redirect('create_billing')

                # Subtotal
                subtotal = Decimal('0.00')
                for i in range(len(item_names)):
                    if item_names[i]:
                        qty   = Decimal(quantities[i])  if (i < len(quantities)  and quantities[i])  else Decimal('1')
                        price = Decimal(unit_prices[i]) if (i < len(unit_prices) and unit_prices[i]) else Decimal('0')
                        subtotal += qty * price

                # Fetch Customer
                customer = Customer.objects.select_related(
                    'company', 'state', 'company__state'
                ).get(id=customer_id)

                # -------------------------------------------------------
                # GST Calculation — based on COMPANY's gst_registered flag
                # -------------------------------------------------------
                total_tax = Decimal('0.00')
                gst_type  = 'NONE'

                if not customer.company or not customer.company.gst_registered:
                    # Company is not GST registered → No GST
                    gst_type  = 'NON-GST'
                    total_tax = Decimal('0.00')
                else:
                    if customer.state and customer.company.state:
                        if customer.state.id == customer.company.state.id:
                            total_tax = subtotal * Decimal('0.18')
                            gst_type  = 'INTRA-STATE'
                        else:
                            total_tax = subtotal * Decimal('0.18')
                            gst_type  = 'INTER-STATE'
                    else:
                        gst_type  = 'NONE'
                        total_tax = Decimal('0.00')
                        messages.warning(
                            request,
                            'GST was not calculated because customer or company state '
                            'information is missing. Please update state information to '
                            'calculate GST correctly.'
                        )

                # Discount and Total
                discount_str = request.POST.get('discount_amount', '0')
                discount     = Decimal(discount_str) if discount_str else Decimal('0')
                total_amount = subtotal + total_tax - discount

                if total_amount < 0:
                    messages.error(
                        request,
                        'Total amount cannot be negative. Please check the discount amount.'
                    )
                    return redirect('create_billing')

                # Generate invoice number inside atomic block to prevent race conditions
                with transaction.atomic():
                    last_invoice = BillingRecord.objects.select_for_update().filter(
                        invoice_number__startswith=f'INV-{date.today().year}-'
                    ).order_by('-invoice_number').first()

                    if last_invoice and last_invoice.invoice_number:
                        try:
                            last_sequence = int(last_invoice.invoice_number.split('-')[-1])
                            new_sequence  = last_sequence + 1
                        except (ValueError, IndexError):
                            new_sequence = 1
                    else:
                        new_sequence = 1

                    invoice_number = f'INV-{date.today().year}-{new_sequence:03d}'

                # Create Billing Record
                billing_record = BillingRecord.objects.create(
                    customer_id          = customer_id,
                    customer_location_id = customer_location_id,
                    billing_date         = date.today(),
                    due_date             = request.POST.get('due_date'),
                    amount               = total_amount,
                    gst_amount           = total_tax,
                    gst_type             = gst_type,
                    discount_amount      = discount,
                    balance_amount       = total_amount,
                    notes                = request.POST.get('notes', ''),
                    invoice_number       = invoice_number,
                )

                billing_record.refresh_from_db()

                # Create Bill Items
                for i in range(len(item_names)):
                    if item_names[i]:
                        qty   = Decimal(quantities[i])  if (i < len(quantities)  and quantities[i])  else Decimal('1')
                        price = Decimal(unit_prices[i]) if (i < len(unit_prices) and unit_prices[i]) else Decimal('0')

                        BillItem.objects.create(
                            billing_record = billing_record,
                            item_name      = item_names[i],
                            description    = descriptions[i] if i < len(descriptions) else '',
                            quantity       = qty,
                            unit_price     = price,
                        )

                # Build GST Info String for Logs
                if not customer.company or not customer.company.gst_registered:
                    gst_log_info = 'GST: Not applicable (Company not GST registered)'
                elif total_tax > 0:
                    if gst_type == 'INTRA-STATE':
                        cgst = total_tax / 2
                        sgst = total_tax / 2
                        gst_log_info = (
                            'GST Type: Intra-State (CGST 9% + SGST 9%)\n'
                            'CGST: Rs.' + str(cgst.quantize(Decimal('0.01'))) + '\n'
                            'SGST: Rs.' + str(sgst.quantize(Decimal('0.01')))
                        )
                    else:
                        gst_log_info = (
                            'GST Type: Inter-State\n'
                            'IGST: Rs.' + str(total_tax.quantize(Decimal('0.01')))
                        )
                else:
                    gst_log_info = 'GST: Not applied (Missing state information)'

                # Location Label for Logs
                location_info = ''
                if customer_location_id and billing_record.customer_location:
                    location_info = ' - Location: ' + billing_record.customer_location.location_name

                # LOG 1: Bill creation entry (no payment, no approval needed)
                log_details = (
                    'Invoice Number: ' + str(billing_record.invoice_number) + '\n'
                    'Entity: ' + customer.name + location_info + '\n'
                    'Billing Date: ' + str(billing_record.billing_date) + '\n'
                    'Due Date: ' + str(billing_record.due_date) + '\n'
                    'Subtotal: Rs.' + str(subtotal.quantize(Decimal('0.01'))) + '\n'
                    + gst_log_info + '\n'
                    'Total Tax: Rs.' + str(total_tax.quantize(Decimal('0.01'))) + '\n'
                    'Discount: Rs.' + str(discount.quantize(Decimal('0.01'))) + '\n'
                    'Total Amount: Rs.' + str(total_amount.quantize(Decimal('0.01'))) + '\n'
                    'Number of Items: ' + str(len([n for n in item_names if n])) + '\n'
                    'Notes: ' + (billing_record.notes or 'None')
                )

                data_logs.objects.create(
                    user           = user,
                    customer       = customer,
                    location       = location if customer_location_id else None,
                    billing_record = billing_record,
                    payment_amount = Decimal('0.00'),
                    status         = 'Pending',
                    balance_amount = total_amount,
                    is_payment     = False,
                    is_approved    = False,
                    total_paid     = Decimal('0.00'),
                    action         = 'Bill Created',
                    details        = log_details,
                )

                # LOG 2: Advance adjustment entry (only if customer has advance)
                customer.refresh_from_db()
                advance_available = Decimal(str(customer.advance_amount or 0))

                if advance_available > 0:
                    advance_used      = min(advance_available, total_amount)
                    new_balance       = total_amount - advance_used
                    remaining_advance = advance_available - advance_used

                    billing_record.paid_amount    = advance_used
                    billing_record.balance_amount = new_balance
                    billing_record.paid           = new_balance <= 0
                    billing_record.discount_amount = advance_used
                    billing_record.payment_mode   = 'Advance Adjustment'
                    billing_record.payment_date   = date.today()   # ← payment date set to today
                    billing_record.save()

                    customer.advance_amount = remaining_advance
                    customer.save(update_fields=['advance_amount'])

                    advance_log_details = (
                        'Advance Adjustment for Invoice: ' + str(billing_record.invoice_number) + '\n'
                        'Customer: ' + customer.name + location_info + '\n'
                        'Bill Total: Rs.' + str(total_amount.quantize(Decimal('0.01'))) + '\n'
                        'Advance Used: Rs.' + str(advance_used.quantize(Decimal('0.01'))) + '\n'
                        'Remaining Advance: Rs.' + str(remaining_advance.quantize(Decimal('0.01'))) + '\n'
                        'Balance After Adjustment: Rs.' + str(new_balance.quantize(Decimal('0.01')))
                    )

                    data_logs.objects.create(
                        user           = user,
                        customer       = customer,
                        location       = location if customer_location_id else None,
                        billing_record = billing_record,
                        payment_amount = advance_used,
                        status         = 'Paid' if new_balance <= 0 else 'Partially Paid',
                        balance_amount = new_balance,
                        is_payment     = True,
                        is_approved    = True,           # ← FIXED: advance is auto-approved
                        total_paid     = advance_used,
                        payment_mode   = 'Advance Adjustment',
                        payment_date   = date.today(),   # ← FIXED: payment date is today
                        action         = 'Advance Adjusted on Bill Creation',
                        details        = advance_log_details,
                    )

                # User Log
                location_log = ''
                if customer_location_id and billing_record.customer_location:
                    location_log = ' for Location: ' + billing_record.customer_location.location_name

                user_logs.objects.create(
                    user    = user.username,
                    action  = 'Created Bill',
                    details = (
                        'Invoice: ' + str(billing_record.invoice_number) + ', '
                        'Customer: ' + customer.name + location_log + ', '
                        'Amount: Rs.' + str(total_amount.quantize(Decimal('0.01')))
                    ),
                )

                # Success Message
                location_msg = ''
                if customer_location_id and billing_record.customer_location:
                    location_msg = ' for location: ' + billing_record.customer_location.location_name

                if not customer.company or not customer.company.gst_registered:
                    gst_msg = ' (No GST - company not GST registered)'
                elif total_tax > 0:
                    gst_msg = (
                        ' (Subtotal: Rs.' + str(subtotal.quantize(Decimal('0.01')))
                        + ' + GST: Rs.' + str(total_tax.quantize(Decimal('0.01')))
                        + ' - Discount: Rs.' + str(discount.quantize(Decimal('0.01'))) + ')'
                    )
                else:
                    gst_msg = ' (No GST applied - missing state information)'

                messages.success(
                    request,
                    'Bill ' + str(billing_record.invoice_number) + ' created successfully'
                    + location_msg + '! '
                    + 'Total: Rs.' + str(total_amount.quantize(Decimal('0.01')))
                    + gst_msg
                )
                return redirect('billing_management')

        except ValueError as e:
            messages.error(request, 'Invalid number format: ' + str(e))
            import traceback; traceback.print_exc()
            return redirect('create_billing')

        except Exception as e:
            messages.error(request, 'Error creating bill: ' + str(e))
            import traceback; traceback.print_exc()
            return redirect('create_billing')

    # ── GET ───────────────────────────────────────────────────────────────────
    filtered_companies = get_filtered_companies(request)
    company_ids        = list(filtered_companies.values_list('id', flat=True))

    if company_ids:
        customers = Customer.objects.filter(
            company_id__in=company_ids
        ).select_related('company', 'state', 'company__state', 'subscription_plan').order_by('name')
    else:
        customers = Customer.objects.select_related(
            'company', 'state', 'company__state', 'subscription_plan'
        ).order_by('name')

    companies = filtered_companies.order_by('name')
    current_year = date.today().year
    last_invoice = BillingRecord.objects.filter(
        invoice_number__startswith=f'INV-{current_year}-'
    ).order_by('-invoice_number').first()

    if last_invoice and last_invoice.invoice_number:
        try:
            last_seq = int(last_invoice.invoice_number.split('-')[-1])
            next_seq = last_seq + 1
        except (ValueError, IndexError):
            next_seq = 1
    else:
        next_seq = 1

    next_invoice_number = f'INV-{current_year}-{next_seq:03d}'

    context = {
        'customers'          : customers,
        'companies'          : companies,
        'next_invoice_number': next_invoice_number,
        **get_company_filter_context(request),
        'user'               : user,
        'role_permissions'   : role_permissions,
    }

    return render(request, 'billing.html', context)
    
from django.http import JsonResponse
from django.views.decorators.http import require_GET

@require_GET
def get_billing_locations(request):
    """API endpoint to get locations for a customer or company"""
    customer_id = request.GET.get('customer_id')
    company_id = request.GET.get('company_id')
    locations = []
    
    if customer_id:
        try:
            customer_locations = CustomerLocation.objects.filter(
                customer_id=customer_id,
                is_active=True
            ).select_related('state')
            
            locations = [
                {
                    'id': loc.id,
                    'location_name': loc.location_name,
                    'state': str(loc.state) if loc.state else '',
                    'display_text': f"{loc.location_name}{' - ' + str(loc.state) if loc.state else ''}"
                }
                for loc in customer_locations
            ]
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    elif company_id:
        try:
            company_locations = CompanyLocation.objects.filter(
                company_id=company_id,
                is_active=True
            ).select_related('state')
            
            locations = [
                {
                    'id': loc.id,
                    'location_name': loc.location_name,
                    'state': str(loc.state) if loc.state else '',
                    'display_text': f"{loc.location_name}{' - ' + str(loc.state) if loc.state else ''}"
                }
                for loc in company_locations
            ]
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'locations': locations})




def billing_detail(request, pk):
    """View billing record details"""
    billing_record = get_object_or_404(
        BillingRecord.objects.select_related(
            'customer', 'company', 'customer_location', 'company_location'
        ).prefetch_related('bill_items'),
        pk=pk
    )
    
    # Calculate totals from items
    subtotal = sum(item.total_price for item in billing_record.bill_items.all())
    total_tax = sum(item.tax_amount for item in billing_record.bill_items.all())
    
    context = {
        'billing_record': billing_record,
        'subtotal': subtotal,
        'total_tax': total_tax,
    }
    return render(request, 'billing/billing_detail.html', context)


def edit_billing(request, pk):
    """Edit existing billing record"""
    billing_record = get_object_or_404(BillingRecord, pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Update billing record
                customer_id = request.POST.get('customer')
                company_id = request.POST.get('company')
                
                # Validate
                if not customer_id and not company_id:
                    messages.error(request, 'Please select either a customer or company.')
                    return redirect('edit_billing', pk=pk)
                
                if customer_id and company_id:
                    messages.error(request, 'Cannot select both customer and company.')
                    return redirect('edit_billing', pk=pk)
                
                # Get items data
                item_names = request.POST.getlist('item_name[]')
                descriptions = request.POST.getlist('description[]')
                quantities = request.POST.getlist('quantity[]')
                unit_prices = request.POST.getlist('unit_price[]')
                tax_percentages = request.POST.getlist('tax_percentage[]')
                
                # Calculate new totals
                subtotal = Decimal('0.00')
                total_tax = Decimal('0.00')
                
                for i in range(len(item_names)):
                    if item_names[i]:
                        qty = Decimal(quantities[i]) if i < len(quantities) else Decimal('1')
                        price = Decimal(unit_prices[i]) if i < len(unit_prices) else Decimal('0')
                        tax_pct = Decimal(tax_percentages[i]) if i < len(tax_percentages) else Decimal('0')
                        
                        item_total = qty * price
                        item_tax = (item_total * tax_pct) / 100
                        
                        subtotal += item_total
                        total_tax += item_tax
                
                discount = Decimal(request.POST.get('discount_amount', '0'))
                total_amount = subtotal + total_tax - discount
                
                # Update billing record fields
                billing_record.customer_id = customer_id if customer_id else None
                billing_record.company_id = company_id if company_id else None
                billing_record.customer_location_id = request.POST.get('customer_location') or None
                billing_record.company_location_id = request.POST.get('company_location') or None
                billing_record.invoice_number = request.POST.get('invoice_number')
                billing_record.billing_date = request.POST.get('billing_date')
                billing_record.due_date = request.POST.get('due_date')
                billing_record.amount = total_amount
                billing_record.notes = request.POST.get('notes', '')
                billing_record.save()
                
                # Delete existing items and create new ones
                billing_record.bill_items.all().delete()
                
                # Create new items
                for i in range(len(item_names)):
                    if item_names[i]:
                        BillItem.objects.create(
                            billing_record=billing_record,
                            item_name=item_names[i],
                            description=descriptions[i] if i < len(descriptions) else '',
                            quantity=quantities[i] if i < len(quantities) else 1,
                            unit_price=unit_prices[i] if i < len(unit_prices) else 0,
                            tax_percentage=tax_percentages[i] if i < len(tax_percentages) else 0,
                        )
                
                messages.success(request, 'Bill updated successfully!')
                return redirect('billing_detail', pk=billing_record.pk)
                
        except Exception as e:
            messages.error(request, f'Error updating bill: {str(e)}')
            return redirect('edit_billing', pk=pk)
    
    # GET request
    customers = Customer.objects.all().order_by('name')
    companies = Company.objects.all().order_by('name')
    
    # Get locations for the selected customer/company
    customer_locations = []
    company_locations = []
    
    if billing_record.customer:
        customer_locations = CustomerLocation.objects.filter(
            customer=billing_record.customer, is_active=True
        )
    
    if billing_record.company:
        company_locations = CompanyLocation.objects.filter(
            company=billing_record.company, is_active=True
        )
    
    context = {
        'billing_record': billing_record,
        'customers': customers,
        'companies': companies,
        'customer_locations': customer_locations,
        'company_locations': company_locations,
    }
    return render(request, 'billing/edit_billing.html', context)


@require_http_methods(["GET"])
def get_locations(request):
    """AJAX endpoint to get locations for a customer or company"""
    customer_id = request.GET.get('customer_id')
    company_id = request.GET.get('company_id')
    
    locations = []
    
    if customer_id:
        locations = list(CustomerLocation.objects.filter(
            customer_id=customer_id, is_active=True
        ).values('id', 'location_name', 'address'))
    elif company_id:
        locations = list(CompanyLocation.objects.filter(
            company_id=company_id, is_active=True
        ).values('id', 'location_name', 'address'))
    
    return JsonResponse({'locations': locations})



from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
import json
from .models import Company

FILTER_PASSWORD = "kinggate@123"  # Change this to your desired password


@require_http_methods(["POST"])
def validate_filter_password(request):
    """
    Validate the password for accessing company filters
    """
    try:
        data = json.loads(request.body)
        password = data.get('password', '')
        
        if password == FILTER_PASSWORD:
            # Store unlock status in session
            request.session['filter_unlocked'] = True
            return JsonResponse({
                'success': True,
                'message': 'Password validated successfully'
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Incorrect password'
            })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@require_http_methods(["POST"])
def lock_filters(request):
    """
    Lock filters and reset to default state
    """
    try:
        # Reset all filter-related session data to defaults
        request.session['filter_unlocked'] = False
        request.session['selected_gst_type'] = 'gst'  # Reset to default GST type
        request.session['selected_company_ids'] = []  # Clear all company selections
        request.session.modified = True
        
        return JsonResponse({
            'success': True,
            'message': 'Filters locked and reset to default'
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


def get_companies(request):
    """
    Get all companies for the filter dropdown
    """
    try:
        from .models import Company  # Adjust import path as needed
        companies = Company.objects.all().values('id', 'name', 'gst_registered')
        return JsonResponse({
            'success': True,
            'companies': list(companies)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@require_http_methods(["POST"])
def update_company_filter(request):
    try:
        data = json.loads(request.body)
        gst_type = data.get('gst_type', 'gst')  # Default to 'gst'
        companies = data.get('companies', [])
        
        # Convert company IDs to integers
        company_ids = []
        for comp_id in companies:
            try:
                company_ids.append(int(comp_id))
            except (ValueError, TypeError):
                pass
        
        # Store in session
        request.session['selected_gst_type'] = gst_type if gst_type else 'gst'
        request.session['selected_company_ids'] = company_ids
        request.session.modified = True
        
        return JsonResponse({
            'success': True,
            'gst_type': request.session['selected_gst_type'],
            'companies': request.session['selected_company_ids'],
            'message': 'Filters updated successfully'
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


def get_company_filter_context(request):
    """
    Helper function to get filter context for templates
    Call this in your views to pass filter data to templates
    """
    # Get GST type from session, default to 'gst'
    selected_gst_type = request.session.get('selected_gst_type', 'gst')
    
    # Get selected company IDs from session
    selected_company_ids = request.session.get('selected_company_ids', [])
    
    # Check if filters are unlocked
    filter_unlocked = request.session.get('filter_unlocked', False)
    
    return {
        'selected_gst_type': selected_gst_type,
        'selected_company_ids': selected_company_ids,
        'filter_unlocked': filter_unlocked
    }


from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from decimal import Decimal
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

from .models import Customer, CustomerLocation, data_logs, BillingRecord
# Adjust these imports to your project structure
# from .utils import get_logged_in_user, get_company_filter_context, get_filtered_companies


# ─────────────────────────────────────────────────────────────────────────────
# HELPER
# ─────────────────────────────────────────────────────────────────────────────

def get_filtered_companies(request):
    """
    Returns queryset of companies based on session GST / company filters.
    """
    from django.db.models import Q
    from .models import Company

    selected_gst_type    = request.session.get('selected_gst_type', 'gst')
    selected_company_ids = request.session.get('selected_company_ids', [])

    companies = Company.objects.all()

    if selected_gst_type == 'gst':
        companies = companies.filter(gst_registered=True)
    elif selected_gst_type == 'non-gst':
        companies = companies.filter(gst_registered=False)

    if selected_company_ids:
        companies = companies.filter(id__in=selected_company_ids)

    return companies


def _build_approval_qs(request, selected_customer, start_date, end_date,
                       company_ids, filters_active):
    """
    Shared queryset builder used by the list view, approve views, and export.
    Returns a data_logs queryset (or .none()).
    """
    base = data_logs.objects.filter(
        is_payment=True,
        is_approved=False,
        payment_date__isnull=False,
        payment_mode__isnull=False,
    ).select_related(
        'customer', 'location',
        'location__city', 'location__state',
        'approved_by', 'billing_record',
    ).order_by('-timestamp')

    if start_date:
        base = base.filter(payment_date__gte=start_date)
    if end_date:
        base = base.filter(payment_date__lte=end_date)

    if selected_customer == 'all':
        if company_ids:
            return base.filter(customer__company_id__in=company_ids)
        if not filters_active:
            return base
        return data_logs.objects.none()

    if selected_customer:
        try:
            customer_obj = Customer.objects.get(id=selected_customer)
            if company_ids and customer_obj.company_id not in company_ids:
                return data_logs.objects.none()
            if filters_active and not company_ids:
                return data_logs.objects.none()
            return base.filter(customer_id=selected_customer)
        except Customer.DoesNotExist:
            return data_logs.objects.none()

    return data_logs.objects.none()


# ─────────────────────────────────────────────────────────────────────────────
# MAIN LIST VIEW
# ─────────────────────────────────────────────────────────────────────────────

def payment_approval_list(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')

    selected_customer = request.GET.get('customer_id', '')   # '' | 'all' | '<id>'
    start_date        = request.GET.get('start_date', '')
    end_date          = request.GET.get('end_date', '')

    filtered_companies   = get_filtered_companies(request)
    company_ids          = list(filtered_companies.values_list('id', flat=True))
    selected_gst_type    = request.session.get('selected_gst_type', 'gst')
    selected_company_ids = request.session.get('selected_company_ids', [])
    filters_active       = selected_gst_type or selected_company_ids

    if company_ids:
        customers = Customer.objects.filter(company_id__in=company_ids).order_by('name')
    elif filters_active:
        customers = Customer.objects.none()
    else:
        customers = Customer.objects.all().order_by('name')

    payment_logs = _build_approval_qs(
        request, selected_customer, start_date, end_date,
        company_ids, filters_active
    )
    filtered_count = payment_logs.count()

    context = {
        'payment_logs':         payment_logs,
        'customers':            customers,
        'selected_customer':    selected_customer,
        'start_date':           start_date,
        'end_date':             end_date,
        'filtered_count':       filtered_count,
        'filtered_company_ids': company_ids,
        'user':                 user,
        'role_permissions':     role_permissions,
    }
    context.update(get_company_filter_context(request))
    return render(request, 'paymentapproval.html', context)


# ─────────────────────────────────────────────────────────────────────────────
# SINGLE APPROVE  (AJAX POST)
# ─────────────────────────────────────────────────────────────────────────────

@require_POST
def approve_payment(request, payment_id):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return JsonResponse({'success': False, 'message': 'Not authenticated'}, status=401)

    if not getattr(role_permissions, 'payment_approval_e', False):
        return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)

    try:
        log = data_logs.objects.select_related('billing_record', 'customer').get(
            id=payment_id, is_payment=True, is_approved=False
        )
    except data_logs.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Payment log not found or already approved'})

    try:
        # Approve the log entry
        log.is_approved  = True
        log.approved_by  = user
        log.approved_at  = timezone.now()
        log.payment_date = log.payment_date or timezone.now().date()
        log.status       = 'Paid' if (log.balance_amount or 0) <= 0 else 'Pending'
        log.save()

        # Sync the linked BillingRecord
        br = log.billing_record
        if br:
            br.payment_date  = log.payment_date
            br.payment_mode  = log.payment_mode
            br.paid_amount   = log.total_paid or log.payment_amount or Decimal('0')
            br.balance_amount = log.balance_amount or Decimal('0')
            br.paid          = (br.balance_amount <= 0)
            br.save(update_fields=[
                'payment_date', 'payment_mode',
                'paid_amount', 'balance_amount', 'paid'
            ])

        return JsonResponse({
            'success': True,
            'message': f'Payment approved for {log.customer.name if log.customer else "customer"}.'
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# ─────────────────────────────────────────────────────────────────────────────
# BULK APPROVE  (AJAX POST)
# ─────────────────────────────────────────────────────────────────────────────

@require_POST
def bulk_approve_payments(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return JsonResponse({'success': False, 'message': 'Not authenticated'}, status=401)

    if not getattr(role_permissions, 'payment_approval_e', False):
        return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)

    payment_ids = request.POST.getlist('payment_ids')
    if not payment_ids:
        return JsonResponse({'success': False, 'message': 'No payment IDs provided'})

    approved_count = 0
    errors         = []

    for pid in payment_ids:
        try:
            log = data_logs.objects.select_related('billing_record').get(
                id=pid, is_payment=True, is_approved=False
            )
            log.is_approved  = True
            log.approved_by  = user
            log.approved_at  = timezone.now()
            log.payment_date = log.payment_date or timezone.now().date()
            log.status       = 'Paid' if (log.balance_amount or 0) <= 0 else 'Pending'
            log.save()

            br = log.billing_record
            if br:
                br.payment_date   = log.payment_date
                br.payment_mode   = log.payment_mode
                br.paid_amount    = log.total_paid or log.payment_amount or Decimal('0')
                br.balance_amount = log.balance_amount or Decimal('0')
                br.paid           = (br.balance_amount <= 0)
                br.save(update_fields=[
                    'payment_date', 'payment_mode',
                    'paid_amount', 'balance_amount', 'paid'
                ])

            approved_count += 1

        except data_logs.DoesNotExist:
            errors.append(f'ID {pid} not found or already approved')
        except Exception as e:
            errors.append(f'ID {pid}: {str(e)}')

    if approved_count > 0:
        msg = f'{approved_count} payment(s) approved successfully.'
        if errors:
            msg += f' {len(errors)} failed.'
        return JsonResponse({'success': True, 'message': msg, 'approved_count': approved_count})

    return JsonResponse({
        'success': False,
        'message': 'No payments were approved. ' + '; '.join(errors)
    })


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT TO EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def export_payment_approval_excel(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')

    selected_customer    = request.GET.get('customer_id', '')
    start_date           = request.GET.get('start_date', '')
    end_date             = request.GET.get('end_date', '')

    filtered_companies   = get_filtered_companies(request)
    company_ids          = list(filtered_companies.values_list('id', flat=True))
    selected_gst_type    = request.session.get('selected_gst_type', 'gst')
    selected_company_ids = request.session.get('selected_company_ids', [])
    filters_active       = selected_gst_type or selected_company_ids

    logs = _build_approval_qs(
        request, selected_customer, start_date, end_date,
        company_ids, filters_active
    )

    # ── Workbook setup ────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pending Approvals'

    # Style helpers
    def hfont(bold=False, color='000000', size=11, italic=False):
        return Font(name='Calibri', bold=bold, color=color, size=size, italic=italic)

    def fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def border():
        s = Side(style='thin', color='D1D5DB')
        return Border(left=s, right=s, top=s, bottom=s)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right',  vertical='center')

    # ── Row 1: Title ─────────────────────────────────────────────
    ws.merge_cells('A1:L1')
    c = ws['A1']
    c.value     = 'Pending Payment Approvals'
    c.font      = hfont(bold=True, size=15, color='0F172A')
    c.fill      = fill('F8FAFC')
    c.alignment = center
    ws.row_dimensions[1].height = 32

    # ── Row 2: Filter summary ─────────────────────────────────────
    filter_parts = []
    if selected_customer == 'all' or not selected_customer:
        filter_parts.append('Customer: All')
    else:
        try:
            co = Customer.objects.get(id=selected_customer)
            filter_parts.append(f'Customer: {co.name}')
        except Customer.DoesNotExist:
            filter_parts.append('Customer: Unknown')
    if start_date:
        filter_parts.append(f'From: {start_date}')
    if end_date:
        filter_parts.append(f'To: {end_date}')
    filter_parts.append(f'Exported: {datetime.now().strftime("%d-%m-%Y %H:%M")}')

    ws.merge_cells('A2:L2')
    c2 = ws['A2']
    c2.value     = '   |   '.join(filter_parts)
    c2.font      = hfont(italic=True, color='64748B', size=9)
    c2.alignment = left
    ws.row_dimensions[2].height = 16

    # ── Row 3: spacer ─────────────────────────────────────────────
    ws.row_dimensions[3].height = 8

    # ── Row 4: Column headers ─────────────────────────────────────
    headers = [
        '#', 'Customer Name', 'Location',
        'Billing Start', 'Billing End',
        'Amount (₹)', 'Payment Date', 'Payment Mode',
        'Transaction ID', 'Receipt Number',
        'Notes', 'Submitted At',
    ]
    col_widths = [5, 26, 22, 15, 15, 14, 15, 15, 24, 18, 32, 20]

    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=hdr)
        cell.font      = hfont(bold=True, color='FFFFFF', size=10)
        cell.fill      = fill('1A56DB')
        cell.alignment = center
        cell.border    = border()
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[4].height = 22

    # ── Data rows ─────────────────────────────────────────────────
    log_list = list(logs)  # evaluate once

    for row_idx, log in enumerate(log_list, 1):
        excel_row   = row_idx + 4
        row_fill    = fill('FFFFFF') if row_idx % 2 == 1 else fill('F8FAFC')

        p_start   = log.billing_period_start.strftime('%d-%m-%Y') if log.billing_period_start else '—'
        p_end     = log.billing_period_end.strftime('%d-%m-%Y')   if log.billing_period_end   else '—'
        pay_date  = log.payment_date.strftime('%d-%m-%Y')          if log.payment_date          else '—'
        loc_name  = log.location.location_name                     if log.location              else 'Main Address'
        ts        = log.timestamp.strftime('%d-%m-%Y %H:%M')       if log.timestamp             else '—'

        row_data = [
            row_idx,
            log.customer.name if log.customer else '—',
            loc_name,
            p_start, p_end,
            float(log.payment_amount) if log.payment_amount else 0.0,
            pay_date,
            (log.payment_mode or '—').upper(),
            log.transaction_id or '—',
            log.receipt_number or '—',
            log.payment_notes  or '—',
            ts,
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell            = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.fill       = row_fill
            cell.border     = border()
            cell.alignment  = center if col_idx in (1, 6) else left

            # Amount column styling
            if col_idx == 6:
                cell.number_format = '#,##0.00'
                cell.font = hfont(bold=True, color='1A56DB', size=11)
            else:
                cell.font = hfont(size=10)

        ws.row_dimensions[excel_row].height = 18

    # ── Totals row ────────────────────────────────────────────────
    total_row = len(log_list) + 5
    label_cell = ws.cell(row=total_row, column=5, value='TOTAL')
    label_cell.font      = hfont(bold=True, color='0F172A', size=10)
    label_cell.alignment = right
    label_cell.fill      = fill('DEF7EC')
    label_cell.border    = border()

    total_cell = ws.cell(
        row=total_row, column=6,
        value=f'=SUM(F5:F{total_row - 1})'
    )
    total_cell.number_format = '#,##0.00'
    total_cell.font          = hfont(bold=True, color='057A55', size=11)
    total_cell.fill          = fill('DEF7EC')
    total_cell.border        = border()
    total_cell.alignment     = center

    ws.row_dimensions[total_row].height = 22

    # Freeze panes below header
    ws.freeze_panes = 'A5'

    # ── Serve ─────────────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"pending_approvals_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def bulk_approve_payments(request):
    
    
    # Get logged in user
    username = request.session.get('username')
    logged_in_user = Custom_User.objects.get(username=username) if username else None
    
    if request.method == 'POST':
        payment_ids = request.POST.getlist('payment_ids')
        
        if not payment_ids:
            return JsonResponse({'success': False, 'message': 'No payments selected'})
        
        try:
            payments = data_logs.objects.filter(
                id__in=payment_ids,
                is_payment=True,
                is_approved=False
            )
            
            approved_count = 0
            total_amount = Decimal('0.00')
            customer_names = []
            
            for payment in payments:
                payment.is_approved = True
                payment.approved_at = timezone.now()
                payment.status = 'approved'
                payment.action = 'Payment Received (Approved)'
                payment.save()
                
                # Track for logging
                total_amount += Decimal(str(payment.payment_amount or 0))
                if payment.customer and payment.customer.name not in customer_names:
                    customer_names.append(payment.customer.name)
                
                # Update billing record if linked
                if payment.billing_record:
                    billing_record = payment.billing_record
                    if billing_record.paid_amount:
                        billing_record.paid_amount += payment.payment_amount
                    else:
                        billing_record.paid_amount = payment.payment_amount
                    
                    if billing_record.paid_amount >= billing_record.amount:
                        billing_record.paid = True
                        billing_record.payment_date = payment.payment_date
                        billing_record.payment_mode = payment.payment_mode
                    
                    billing_record.save()
                
                approved_count += 1
            
            # Log bulk approval
            if logged_in_user:
                customer_list = ", ".join(customer_names[:3])  # First 3 customers
                if len(customer_names) > 3:
                    customer_list += f" and {len(customer_names) - 3} more"
                
                user_logs.objects.create(
                    user=logged_in_user.username,
                    action='Bulk Approved Payments',
                    details=f'Approved {approved_count} payment(s), Total Amount: {total_amount:.2f}, Customers: {customer_list}'
                )
            return JsonResponse({
                'success': True,
                'message': f'{approved_count} payment(s) approved successfully',
                'approved_count': approved_count
            })
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


def approve_single_payment(request, payment_id):
    
    
    # Get logged in user
    username = request.session.get('username')
    logged_in_user = Custom_User.objects.get(username=username) if username else None
    
    if request.method == 'POST':
        try:
            payment = data_logs.objects.get(
                id=payment_id,
                is_payment=True,
                is_approved=False
            )
            
            payment.is_approved = True
            payment.approved_at = timezone.now()
            payment.status = 'approved'
            payment.action = 'Payment Received (Approved)'
            payment.save()
            
            # Update billing record if linked
            if payment.billing_record:
                billing_record = payment.billing_record
                if billing_record.paid_amount:
                    billing_record.paid_amount += payment.payment_amount
                else:
                    billing_record.paid_amount = payment.payment_amount
                
                if billing_record.paid_amount >= billing_record.amount:
                    billing_record.paid = True
                    billing_record.payment_date = payment.payment_date
                    billing_record.payment_mode = payment.payment_mode
                
                billing_record.save()
            
            # Log single approval
            if logged_in_user:
                customer_name = payment.customer.name if payment.customer else "Unknown"
                invoice_num = payment.billing_record.invoice_number if payment.billing_record else "N/A"
                
                user_logs.objects.create(
                    user=logged_in_user.username,
                    action='Approved Payment',
                    details=f'Customer: {customer_name}, Invoice: {invoice_num}, Amount: {payment.payment_amount:.2f}, Mode: {payment.payment_mode or "Not specified"}'
                )
            
            return JsonResponse({
                'success': True,
                'message': f'Payment of ₹{payment.payment_amount} approved successfully'
            })
        except data_logs.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Payment not found or already approved'})
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models import user_logs, Custom_User


def user_activity_logs_report(request):
    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    selected_user = request.GET.get('user', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    search_action = request.GET.get('search_action', '')
    
    # Check for export request
    export_excel = request.GET.get('export_excel', '')
    
    # Get all Custom_User objects for dropdown
    all_users = Custom_User.objects.all().order_by('username')
    
    # Initialize logs as None (will show "no filters applied" message)
    logs = None
    user_stats = None
    
    # Build query if any filter is applied
    if selected_user or start_date or end_date or search_action:
        logs = user_logs.objects.all()
        
        # Apply filters
        if selected_user:
            logs = logs.filter(user=selected_user)
        
        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                logs = logs.filter(timestamp__date__gte=start_date_obj)
            except ValueError:
                pass
        
        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                logs = logs.filter(timestamp__date__lte=end_date_obj)
            except ValueError:
                pass
        
        if search_action:
            logs = logs.filter(action__icontains=search_action)
        
        # Order by most recent first
        logs = logs.order_by('-timestamp')
        
        # Calculate statistics for the filtered user
        if logs:
            # Get action breakdown
            action_breakdown = logs.values('action').annotate(
                count=Count('action')
            ).order_by('-count')
            
            user_stats = {
                'total_actions': logs.count(),
                'action_breakdown': {item['action']: item['count'] for item in action_breakdown if item['action']},
            }
    
    # Handle Excel export
    if export_excel and logs is not None:
        return export_user_logs_to_excel(logs, selected_user, start_date, end_date, search_action)
    
    context = {
        'logs': logs,
        'all_users': all_users,
        'selected_user': selected_user,
        'start_date': start_date,
        'end_date': end_date,
        'search_action': search_action,
        'user_stats': user_stats,
        'user': user,
        'role_permissions': role_permissions,
    }
    
    return render(request, 'user_logs.html', context)


def export_user_logs_to_excel(logs, selected_user, start_date, end_date, search_action):
    """
    Export user activity logs to Excel
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "User Activity Logs"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Report Header
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "USER ACTIVITY LOGS REPORT"
    title_cell.font = Font(bold=True, size=16, color="1F4E78")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Filter Information
    row = 3
    ws[f'A{row}'] = "Report Generated:"
    ws[f'B{row}'] = datetime.now().strftime("%d %B, %Y %I:%M %p")
    ws[f'B{row}'].font = Font(bold=True)
    row += 1
    
    if selected_user:
        ws[f'A{row}'] = "User:"
        ws[f'B{row}'] = selected_user
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
    
    if start_date:
        ws[f'A{row}'] = "Start Date:"
        ws[f'B{row}'] = datetime.strptime(start_date, '%Y-%m-%d').strftime("%d %B, %Y")
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
    
    if end_date:
        ws[f'A{row}'] = "End Date:"
        ws[f'B{row}'] = datetime.strptime(end_date, '%Y-%m-%d').strftime("%d %B, %Y")
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
    
    if search_action:
        ws[f'A{row}'] = "Action Filter:"
        ws[f'B{row}'] = search_action
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
    
    row += 1
    
    # Table Headers
    headers = ["#", "Date & Time", "User", "Action", "Details"]
    
    header_row = row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Set column widths
    column_widths = [5, 22, 25, 30, 60]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Data rows
    row = header_row + 1
    for idx, log in enumerate(logs, 1):
        ws.cell(row=row, column=1, value=idx).alignment = center_alignment
        ws.cell(row=row, column=1).border = border
        
        ws.cell(row=row, column=2, value=log.timestamp.strftime("%d/%m/%Y %I:%M %p")).alignment = cell_alignment
        ws.cell(row=row, column=2).border = border
        
        ws.cell(row=row, column=3, value=log.user or "System").alignment = cell_alignment
        ws.cell(row=row, column=3).border = border
        
        ws.cell(row=row, column=4, value=log.action or "—").alignment = cell_alignment
        ws.cell(row=row, column=4).border = border
        
        ws.cell(row=row, column=5, value=log.details or "—").alignment = cell_alignment
        ws.cell(row=row, column=5).border = border
        
        row += 1
    
    # Freeze panes (freeze header row)
    ws.freeze_panes = f'A{header_row + 1}'
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"User_Activity_Logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


# Helper function to log user activities
def log_user_activity(user, action, details=None):
   
    return user_logs.objects.create(
        user=user,
        action=action,
        details=details
    )
import io
import threading
import unicodedata
from datetime import date
from decimal import Decimal

import pandas as pd
from django.contrib import messages
from django.shortcuts import redirect
from django.http import HttpResponse

# ─────────────────────────────────────────────────────────────────────────────
# Column constants
# ─────────────────────────────────────────────────────────────────────────────
COL_CUSTOMER    = "Customer"
COL_PHONE       = "Phone"
COL_EMAIL       = "Email"
COL_STATE       = "State"
COL_CITY        = "City"
COL_AREA        = "Area"
COL_COMPANY     = "Company"
COL_ADDRESS     = "Address"
COL_PINCODE     = "Pincode"
COL_LOC_LINK    = "Location Link"
COL_SVC_START   = "Service Start"
COL_ACCOUNTING  = "Accounting"
COL_ADVANCE     = "Advance Amt"
COL_OUTSTANDING = "Outstanding Amount"
COL_CUSTOM_AMT  = "Custom Amount"
COL_GST_REG     = "GST Registered"
COL_GST_NUM     = "GST Number"
COL_TAX_PCT     = "Tax Percentage"
COL_REMARKS     = "Remarks"
COL_CAMERAS     = "Camera Serials"
COL_STATUS      = "Status"


# ─────────────────────────────────────────────────────────────────────────────
# Phone normaliser
# ─────────────────────────────────────────────────────────────────────────────

def _normalise_phone(raw):
    """
    Cleans any phone value from Excel into a plain digit string.
    Handles: '9350181360.0', '9.35E+09', '935 018 1360', '935-018-1360'
    """
    if not raw:
        return ''
    phone = str(raw).strip()
    phone = phone.replace(' ', '').replace('-', '').replace('nan', '')
    try:
        if 'e' in phone.lower():
            phone = str(int(float(phone)))
    except Exception:
        pass
    if phone.endswith('.0'):
        phone = phone[:-2]
    return phone.strip()


# ─────────────────────────────────────────────────────────────────────────────
# Unicode sanitizer
# ─────────────────────────────────────────────────────────────────────────────

def _sanitize(value):
    if not value:
        return value
    value = value.replace('\x00', '')
    value = ''.join(
        ch for ch in value
        if unicodedata.category(ch) != 'Cc' or ch in ('\n', '\r', '\t')
    )
    value = unicodedata.normalize('NFC', value)
    return value.strip()


def _ascii_email_safe(name):
    normalized = unicodedata.normalize('NFKD', name)
    ascii_str  = normalized.encode('ascii', 'ignore').decode('ascii')
    safe       = ascii_str.replace(' ', '_').lower()
    return safe or 'customer'


# ─────────────────────────────────────────────────────────────────────────────
# DB lookup: find customer only if BOTH name AND phone match same record
# ─────────────────────────────────────────────────────────────────────────────

def _resolve_customer_from_db(customer_name, phone):
    if not phone:
        existing = Customer.objects.filter(name__iexact=customer_name).first()
        return (existing, True) if existing else (None, False)

    customer_by_name = Customer.objects.filter(
        name__iexact=customer_name
    ).first()

    if not customer_by_name:
        return (None, False)

    phone_belongs = CustomerContact.objects.filter(
        phone_number=phone,
        customer=customer_by_name
    ).exists()

    if phone_belongs:
        return (customer_by_name, True)

    # Name same but phone different → treat as NEW customer
    return (None, False)


# ─────────────────────────────────────────────────────────────────────────────
# Duplicate location check — checks address AND location_name in DB
# Used ONLY when a customer already existed before this upload session,
# to prevent re-uploading the same file from duplicating locations.
# NOT used for customers created fresh in this upload (their locations
# are always new by definition).
# ─────────────────────────────────────────────────────────────────────────────

def _location_already_exists(customer, location_name):
    if not location_name:
        return False
    name_clean = location_name.strip()
    return CustomerLocation.objects.filter(
        customer=customer
    ).filter(
        models.Q(address__iexact=name_clean) |
        models.Q(location_name__iexact=name_clean)
    ).exists()


# ─────────────────────────────────────────────────────────────────────────────
# Helper: parse camera serials
# ─────────────────────────────────────────────────────────────────────────────

def _parse_camera_serials(raw):
    if pd.isna(raw) or str(raw).strip() in ('', 'nan'):
        return []
    return [_sanitize(s.strip()) for s in str(raw).split(',') if s.strip()]


# ─────────────────────────────────────────────────────────────────────────────
# Helper: save cameras
# ─────────────────────────────────────────────────────────────────────────────

def _save_cameras(serials, customer, location):
    if not serials:
        return 0
    existing_serials = set(
        customer_cameras.objects.filter(
            customer=customer
        ).values_list('seriak_number', flat=True)
    )
    to_create = []
    for serial in serials:
        if serial not in existing_serials:
            to_create.append(customer_cameras(
                customer=customer,
                seriak_number=serial,
                customer_location=location,
            ))
            existing_serials.add(serial)
    if to_create:
        customer_cameras.objects.bulk_create(to_create)
    return len(to_create)


# ─────────────────────────────────────────────────────────────────────────────
# Helper: create Customer
# ─────────────────────────────────────────────────────────────────────────────

def _create_customer(row, customer_name, phone, db_user):

    def _str(col, default=''):
        val = row.get(col, default)
        if val == '' or (not isinstance(val, str) and pd.isna(val)):
            return default
        s = str(val).strip()
        return default if s.lower() == 'nan' else _sanitize(s)

    def _decimal(col, default=None):
        val = row.get(col)
        if val == '' or (not isinstance(val, str) and pd.isna(val)):
            return default
        s = str(val).strip()
        if s in ['', 'nan']:
            return default
        try:
            return Decimal(s)
        except Exception:
            return default

    def _date(col):
        val = row.get(col)
        if val == '' or (not isinstance(val, str) and pd.isna(val)):
            return None
        s = str(val).strip()
        if s in ['', 'nan']:
            return None
        try:
            return pd.to_datetime(s, dayfirst=True).date()
        except Exception:
            return None

    # Company
    company      = None
    company_name = _str(COL_COMPANY)
    if company_name:
        company = Company.objects.filter(name__iexact=company_name).first()
        if not company:
            raise ValueError(f"Company '{company_name}' not found. Please create it first.")

    # State
    state_name = _str(COL_STATE)
    if not state_name:
        raise ValueError("State is required")
    state = State.objects.filter(name__iexact=state_name).first()
    if not state:
        state = State.objects.create(name=state_name.upper())

    # City
    city_name = _str(COL_CITY)
    if not city_name:
        raise ValueError("City is required")
    city = City.objects.filter(name__iexact=city_name, state=state).first()
    if not city:
        city = City.objects.create(name=city_name, state=state)

    # Area
    area      = None
    area_name = _str(COL_AREA)
    if area_name:
        area = Area.objects.filter(name__iexact=area_name, city=city).first()
        if not area:
            area = Area.objects.create(name=area_name, city=city)

    # GST
    gst_raw        = _str(COL_GST_REG, 'Non-GST').lower()
    gst_registered = gst_raw in ['gst', 'yes', 'y', 'true', '1', 'registered']
    gst_number     = None
    raw_gst        = _str(COL_GST_NUM).upper()
    if raw_gst:
        if len(raw_gst) != 15:
            raise ValueError(f"GST number '{raw_gst}' must be exactly 15 characters")
        if Customer.objects.filter(gst_number=raw_gst).exclude(name__iexact=customer_name).exists():
            raise ValueError(f"GST number '{raw_gst}' already exists for a different customer")
        gst_registered = True
        gst_number     = raw_gst

    # Plan
    plan      = None
    plan_name = _str(COL_ACCOUNTING)
    if plan_name:
        plan = SubscriptionPlan.objects.filter(name__icontains=plan_name).first()
        if not plan:
            plan = SubscriptionPlan.objects.filter(
                name__icontains=plan_name.split()[0]
            ).first()

    # Phone — soft check, never block customer creation
    save_phone = True
    if phone and CustomerContact.objects.filter(phone_number=phone).exists():
        save_phone = False
        print(f"DEBUG >>> Phone '{phone}' already used — creating '{customer_name}' without phone contact")

    # Email
    email = _str(COL_EMAIL)
    if not email:
        email = f"{_ascii_email_safe(customer_name)}@upload.local"

    customer = Customer.objects.create(
        name                       = customer_name,
        email                      = email,
        company                    = company,
        state                      = state,
        city                       = city,
        area                       = area,
        address                    = _str(COL_ADDRESS),
        pincode                    = _str(COL_PINCODE),
        location_link              = _str(COL_LOC_LINK) or None,
        gst_registered             = gst_registered,
        gst_number                 = gst_number,
        tax_percentage             = _decimal(COL_TAX_PCT),
        start_date                 = _date(COL_SVC_START),
        subscription_plan          = plan,
        custom_subscription_amount = _decimal(COL_CUSTOM_AMT),
        outstanding_amount         = _decimal(COL_OUTSTANDING) or Decimal('0.00'),
        advance_amount             = _decimal(COL_ADVANCE),
        status                     = _str(COL_STATUS, 'Active') or 'Active',
        remarks                    = _str(COL_REMARKS),
    )

    if phone and save_phone:
        CustomerContact.objects.create(customer=customer, phone_number=phone)

    user_logs.objects.create(
        user    = db_user.username if db_user else 'system',
        action  = 'Bulk Upload - Customer Created',
        details = _sanitize(f'{customer_name} created via flat bulk upload'),
    )
    return customer


# ─────────────────────────────────────────────────────────────────────────────
# Helper: create Location
# ─────────────────────────────────────────────────────────────────────────────

def _create_location(row, customer, phone, db_user):
    """
    phone here is already normalised by _normalise_phone() in the main loop.
    We use it directly for location_contact — no .0 issue.
    """

    def _str(col, default=''):
        val = row.get(col, default)
        if val == '' or (not isinstance(val, str) and pd.isna(val)):
            return default
        s = str(val).strip()
        return default if s.lower() == 'nan' else _sanitize(s)

    def _decimal(col, default=None):
        val = row.get(col)
        if val == '' or (not isinstance(val, str) and pd.isna(val)):
            return default
        s = str(val).strip()
        if s in ['', 'nan']:
            return default
        try:
            return Decimal(s)
        except Exception:
            return default

    def _date(col):
        val = row.get(col)
        if val == '' or (not isinstance(val, str) and pd.isna(val)):
            return None
        s = str(val).strip()
        if s in ['', 'nan']:
            return None
        try:
            return pd.to_datetime(s, dayfirst=True).date()
        except Exception:
            return None

    location_name = _str(COL_ADDRESS)
    if not location_name:
        raise ValueError("Address (Location Name) is required for a location row")

    state = city = area = None

    state_name = _str(COL_STATE)
    if state_name:
        state = State.objects.filter(name__iexact=state_name).first()
        if not state:
            state = State.objects.create(name=state_name.upper())

    city_name = _str(COL_CITY)
    if city_name and state:
        city = City.objects.filter(name__iexact=city_name, state=state).first()
        if not city:
            city = City.objects.create(name=city_name, state=state)

    area_name = _str(COL_AREA)
    if area_name and city:
        area = Area.objects.filter(name__iexact=area_name, city=city).first()
        if not area:
            area = Area.objects.create(name=area_name, city=city)

    gst_raw        = _str(COL_GST_REG, 'Non-GST').lower()
    gst_registered = gst_raw in ['gst', 'yes', 'y', 'true', '1', 'registered']
    gst_number     = None
    raw_gst        = _str(COL_GST_NUM).upper()
    if raw_gst:
        if len(raw_gst) != 15:
            raise ValueError(f"GST number '{raw_gst}' must be exactly 15 characters")
        gst_registered = True
        gst_number     = raw_gst

    plan      = None
    plan_name = _str(COL_ACCOUNTING)
    if plan_name:
        plan = SubscriptionPlan.objects.filter(name__icontains=plan_name).first()
        if not plan:
            plan = SubscriptionPlan.objects.filter(
                name__icontains=plan_name.split()[0]
            ).first()

    location_link = _str(COL_LOC_LINK) or None
    if location_link and not location_link.startswith(('http://', 'https://')):
        location_link = 'https://' + location_link

    location_email = f"{_ascii_email_safe(location_name)}@upload.local"
    start_date     = _date(COL_SVC_START)

    loc = CustomerLocation.objects.create(
        customer                   = customer,
        location_name              = location_name,
        state                      = state,
        city                       = city,
        area                       = area,
        address                    = location_name,
        pincode                    = _str(COL_PINCODE),
        location_contact           = phone,
        location_email             = location_email,
        location_link              = location_link,
        subscription_plan          = plan,
        custom_subscription_amount = _decimal(COL_CUSTOM_AMT),
        outstanding_amount         = _decimal(COL_OUTSTANDING) or Decimal('0.00'),
        gst_registered             = gst_registered,
        gst_number                 = gst_number,
        start_date                 = start_date,
        is_active                  = _str(COL_STATUS, 'Active').strip().lower() not in (
                                         'inactive', 'no', 'false', '0'),
        remarks                    = _str(COL_REMARKS),
    )

    user_logs.objects.create(
        user    = db_user.username if db_user else 'system',
        action  = 'Bulk Upload - Location Created',
        details = _sanitize(f'{location_name} under {customer.name} via flat bulk upload'),
    )

    first_day = date.today().replace(day=1)
    queued    = bool(start_date and plan and start_date < first_day)
    return loc, queued


# ─────────────────────────────────────────────────────────────────────────────
# Main view
# ─────────────────────────────────────────────────────────────────────────────

def bulk_upload_customers_and_locations(request):
    """
    ROW DECISION LOGIC:
    ════════════════════
    key = (customer_name.lower(), normalised_phone)

    First time key appears in file:
      → _resolve_customer_from_db(name, phone)
        NEW customer  → create it, mark seen_groups[key] = (customer, is_new=True)
        EXISTING customer → reuse it, mark seen_groups[key] = (customer, is_new=False)

    Same key appears again (rows 2, 3, 4, 5 ... with same name+phone):
      → Always treated as a LOCATION for that customer.
      → If customer was NEW this session (is_new=True):
            Always create the location — it cannot exist yet.
      → If customer was EXISTING before this upload (is_new=False):
            Check DB first to avoid duplicating locations on re-upload.
            If location address already exists → skip it.
            If location address is new → create it.

    This means:
      - Fresh upload of 1 customer + 4 locations → creates 1 customer + 4 locations ✅
      - Re-upload of same file → skips customer + skips all 4 existing locations ✅
      - Upload with genuinely new locations for existing customer → creates only the new ones ✅

    ROW FILTERING:
    ══════════════
    Only rows with BOTH a non-empty Customer name AND Phone are processed.
    """
    if request.method != 'POST' or not request.FILES.get('excel_file'):
        messages.error(request, 'No file uploaded.')
        return redirect('customer_management')

    user, _ = get_logged_in_user(request)
    db_user = None
    if user:
        try:
            db_user = Custom_User.objects.get(username=user)
        except Custom_User.DoesNotExist:
            pass
    bg_user_id = db_user.id if db_user else None

    try:
        df = pd.read_excel(
            request.FILES['excel_file'],
            header=0,
            dtype=str,
            engine='openpyxl',
        )
        df.columns = df.columns.str.strip()

        # ── Normalise all cells ───────────────────────────────────────────────
        df = df.fillna('')
        df = df.replace('nan',  '', regex=False)
        df = df.replace('NaN',  '', regex=False)
        df = df.replace('None', '', regex=False)

        # ── Strip whitespace from every cell ─────────────────────────────────
        df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)

        raw_total = len(df)

        if COL_CUSTOMER not in df.columns:
            messages.error(
                request,
                f'Invalid template — missing "{COL_CUSTOMER}" column. '
                'Please download the latest template.'
            )
            return redirect('customer_management')

        # ── STRICT FILTER: keep only rows with BOTH Customer AND Phone ────────
        df = df[
            df[COL_CUSTOMER].ne('') &
            df[COL_PHONE].ne('')
        ]
        df = df.reset_index(drop=True)
        filtered_total = len(df)

        print(f"DEBUG >>> Raw rows from Excel          : {raw_total}")
        print(f"DEBUG >>> Valid rows (name+phone)       : {filtered_total}")
        print(f"DEBUG >>> Dropped phantom/empty rows    : {raw_total - filtered_total}")
        print(f"DEBUG >>> Columns                       : {list(df.columns)}")

        customer_success  = 0
        location_success  = 0
        camera_success    = 0
        customer_skipped  = 0
        location_skipped  = 0
        errors            = []
        failed_rows       = []
        new_customer_ids  = []
        new_location_ids  = []

        # seen_groups[key] = (customer_obj, is_new_this_session: bool)
        # is_new_this_session = True  → customer was CREATED in this upload
        # is_new_this_session = False → customer already existed before this upload
        seen_groups = {}

        for idx, row in df.iterrows():
            row_num = idx + 2

            # ── Customer name ─────────────────────────────────────────────────
            raw_name = row.get(COL_CUSTOMER, '')
            if not raw_name or str(raw_name).strip() in ('', 'nan'):
                continue

            customer_name = _sanitize(str(raw_name).strip())

            # ── Phone — normalised ONCE here ──────────────────────────────────
            phone = _normalise_phone(row.get(COL_PHONE, ''))

            if not phone:
                print(f"DEBUG >>> Row {row_num}: Phone empty after normalisation — skipping")
                continue

            key      = (customer_name.lower(), phone)
            row_dict = {k: str(v) for k, v in row.to_dict().items()}

            print(f"DEBUG >>> Row {row_num} | name='{customer_name}' | phone='{phone}' | in_cache={key in seen_groups}")

            camera_serials = _parse_camera_serials(row.get(COL_CAMERAS, ''))

            # ══════════════════════════════════════════════════════════════════
            # FIRST occurrence of this name+phone → Customer
            # ══════════════════════════════════════════════════════════════════
            if key not in seen_groups:

                customer, is_existing = _resolve_customer_from_db(customer_name, phone)

                if is_existing and customer:
                    # Customer existed BEFORE this upload session
                    print(f"DEBUG >>> Row {row_num}: '{customer_name}' found in DB (id={customer.id}) — skipping create")
                    seen_groups[key]  = (customer, False)   # False = pre-existing
                    customer_skipped += 1
                    if camera_serials:
                        camera_success += _save_cameras(camera_serials, customer, None)

                else:
                    # Brand new customer — create it
                    try:
                        customer = _create_customer(row, customer_name, phone, db_user)
                        seen_groups[key]  = (customer, True)   # True = new this session
                        customer_success += 1
                        print(f"DEBUG >>> Row {row_num}: ✅ Customer CREATED id={customer.id}")

                        if camera_serials:
                            camera_success += _save_cameras(camera_serials, customer, None)

                        first_day = date.today().replace(day=1)
                        if (customer.start_date and
                                customer.subscription_plan and
                                customer.start_date < first_day):
                            new_customer_ids.append(customer.id)

                    except Exception as exc:
                        import traceback; traceback.print_exc()
                        err_str  = str(exc)
                        friendly = (
                            f"Encoding error — run ALTER TABLE utf8mb4 on DB. Detail: {exc}"
                            if ('Incorrect string value' in err_str or '1366' in err_str)
                            else err_str
                        )
                        print(f"DEBUG >>> Row {row_num}: ❌ Customer FAILED — {friendly}")
                        errors.append(f"Row {row_num} ('{customer_name}'): {friendly}")
                        failed_rows.append({
                            'row_num': row_num,
                            'data'   : row_dict,
                            'error'  : f"[Customer] {friendly}",
                        })
                        seen_groups[key] = (None, False)

            # ══════════════════════════════════════════════════════════════════
            # REPEATED name+phone → Location
            # ══════════════════════════════════════════════════════════════════
            else:
                customer, is_new_this_session = seen_groups[key]

                if customer is None:
                    err_msg = f"Parent customer '{customer_name}' (phone: {phone}) failed to create."
                    errors.append(f"Row {row_num}: Skipped — {err_msg}")
                    failed_rows.append({
                        'row_num': row_num,
                        'data'   : row_dict,
                        'error'  : f"[Location skipped] {err_msg}",
                    })
                    continue

                raw_addr      = row.get(COL_ADDRESS, '')
                location_name = _sanitize(str(raw_addr).strip()) if raw_addr else ''

                if not location_name or location_name == 'nan':
                    err_msg = "Address is empty — location skipped."
                    errors.append(f"Row {row_num}: {err_msg}")
                    failed_rows.append({
                        'row_num': row_num,
                        'data'   : row_dict,
                        'error'  : f"[Location] {err_msg}",
                    })
                    continue

                # ── DUPLICATE GUARD ───────────────────────────────────────────
                # Only check the DB for duplicates when the customer already
                # existed before this upload. If the customer was created fresh
                # in this session (is_new_this_session=True), its locations
                # cannot possibly exist yet — skip the DB check entirely.
                if not is_new_this_session:
                    if _location_already_exists(customer, location_name):
                        print(f"DEBUG >>> Row {row_num}: Location '{location_name}' already exists "
                              f"under '{customer.name}' — skipping duplicate")
                        location_skipped += 1
                        if camera_serials:
                            existing_loc = CustomerLocation.objects.filter(
                                customer=customer
                            ).filter(
                                models.Q(address__iexact=location_name.strip()) |
                                models.Q(location_name__iexact=location_name.strip())
                            ).first()
                            camera_success += _save_cameras(camera_serials, customer, existing_loc)
                        continue

                print(f"DEBUG >>> Row {row_num}: Same name+phone → LOCATION "
                      f"'{location_name}' under '{customer.name}' (id={customer.id}) "
                      f"[new_session={is_new_this_session}]")

                try:
                    loc, queued   = _create_location(row, customer, phone, db_user)
                    location_success += 1
                    print(f"DEBUG >>> Row {row_num}: ✅ Location CREATED id={loc.id}")

                    if camera_serials:
                        camera_success += _save_cameras(camera_serials, customer, loc)

                    if queued:
                        new_location_ids.append(loc.id)

                except Exception as exc:
                    import traceback; traceback.print_exc()
                    err_str  = str(exc)
                    friendly = (
                        f"Encoding error — run ALTER TABLE utf8mb4 on DB. Detail: {exc}"
                        if ('Incorrect string value' in err_str or '1366' in err_str)
                        else err_str
                    )
                    print(f"DEBUG >>> Row {row_num}: ❌ Location FAILED — {friendly}")
                    errors.append(f"Row {row_num} (Location for '{customer_name}'): {friendly}")
                    failed_rows.append({
                        'row_num': row_num,
                        'data'   : row_dict,
                        'error'  : f"[Location] {friendly}",
                    })

        # ── Save failed rows to session ────────────────────────────────────────
        if failed_rows:
            request.session['bulk_upload_failed_rows'] = failed_rows
            request.session.modified = True

        # ── Summary ───────────────────────────────────────────────────────────
        print(f"DEBUG >>> DONE | new_customers={customer_success} skipped={customer_skipped} | "
              f"new_locations={location_success} skipped={location_skipped} | "
              f"cameras={camera_success} | errors={len(errors)}")

        if customer_success:
            messages.success(request, f'✅ {customer_success} customer(s) created.')
        if location_success:
            messages.success(request, f'✅ {location_success} location(s) created.')
        if camera_success:
            messages.success(request, f'✅ {camera_success} camera serial(s) saved.')
        if customer_skipped or location_skipped:
            messages.info(
                request,
                f'ℹ️ Skipped {customer_skipped} existing customer(s) and '
                f'{location_skipped} existing location(s) — no duplicates created.'
            )
        if failed_rows:
            messages.warning(
                request,
                f'⚠️ {len(failed_rows)} row(s) failed. '
                f'<a href="/download-failed-upload/" class="btn btn-sm btn-warning ms-2">'
                f'⬇ Download Failed Records</a>'
            )
        if not customer_success and not location_success and not failed_rows:
            messages.info(request, 'ℹ️ All records already exist — nothing new to create.')

        if new_customer_ids:
            threading.Thread(
                target=_generate_bills_for_customers,
                args=(new_customer_ids, bg_user_id),
                daemon=True,
            ).start()
            messages.info(request, f'🔄 Billing started for {len(new_customer_ids)} customer(s).')

        if new_location_ids:
            threading.Thread(
                target=_generate_location_bills_bg,
                args=(new_location_ids, bg_user_id),
                daemon=True,
            ).start()
            messages.info(request, f'🔄 Billing started for {len(new_location_ids)} location(s).')

    except Exception as exc:
        import traceback; traceback.print_exc()
        messages.error(request, f'❌ Error processing file: {exc}')

    return redirect('customer_management')


# ─────────────────────────────────────────────────────────────────────────────
# Download failed records as Excel
# ─────────────────────────────────────────────────────────────────────────────

def download_failed_upload_records(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    failed_rows = request.session.pop('bulk_upload_failed_rows', [])
    if not failed_rows:
        messages.info(request, 'ℹ️ No failed records found.')
        return redirect('customer_management')

    HDR_FILL    = PatternFill("solid", fgColor="C0392B")
    ERR_FILL    = PatternFill("solid", fgColor="FADBD8")
    ALT_FILL    = PatternFill("solid", fgColor="FEF9F9")
    HDR_FONT    = Font(bold=True,   color="FFFFFF", size=10, name="Arial")
    ERR_FONT    = Font(italic=True, color="C0392B", size=9,  name="Arial")
    BODY_FONT   = Font(size=9, name="Arial")
    ROWNUM_FONT = Font(bold=True, color="C0392B", size=9, name="Arial")
    CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    THIN        = Side(style="thin", color="CCCCCC")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    all_data_cols = list(dict.fromkeys(
        col for r in failed_rows for col in r.get('data', {}).keys()
    ))
    headers = ["Row #"] + all_data_cols + ["❌ Error Reason"]

    WIDTH_MAP = {
        "Row #": 7, "❌ Error Reason": 60,
        COL_CUSTOMER: 25, COL_PHONE: 16, COL_EMAIL: 28,
        COL_STATE: 18, COL_CITY: 20, COL_AREA: 18,
        COL_ADDRESS: 28, COL_PINCODE: 10, COL_REMARKS: 24,
        COL_CAMERAS: 35, COL_STATUS: 14,
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Failed Records"

    for ci, header in enumerate(headers, 1):
        c = ws.cell(1, ci, header)
        c.fill = ERR_FILL if header == "❌ Error Reason" else HDR_FILL
        c.font = HDR_FONT; c.alignment = CENTER; c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = WIDTH_MAP.get(
            header, max(14, len(str(header)) + 4))

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "B2"

    for ri, failed in enumerate(failed_rows, 2):
        row_data  = failed.get('data', {})
        error_msg = failed.get('error', 'Unknown error')
        row_num   = failed.get('row_num', '?')
        is_alt    = (ri % 2 == 0)

        rn = ws.cell(ri, 1, row_num)
        rn.font = ROWNUM_FONT; rn.alignment = CENTER; rn.border = BORDER
        if is_alt: rn.fill = ALT_FILL

        for ci, col in enumerate(all_data_cols, 2):
            cell = ws.cell(ri, ci, row_data.get(col, ''))
            cell.font = BODY_FONT; cell.alignment = LEFT; cell.border = BORDER
            if is_alt: cell.fill = ALT_FILL

        err_cell = ws.cell(ri, len(headers), error_msg)
        err_cell.font = ERR_FONT; err_cell.fill = ERR_FILL
        err_cell.alignment = LEFT; err_cell.border = BORDER
        ws.row_dimensions[ri].height = 20

    sr = len(failed_rows) + 3
    ws.cell(sr, 1, f"Total Failed: {len(failed_rows)} record(s)").font = Font(
        bold=True, size=11, color="C0392B", name="Arial")
    ws.cell(sr, 2, "Fix errors above and re-upload only this file.").font = Font(
        italic=True, size=9, color="888888", name="Arial")

    output = io.BytesIO()
    wb.save(output); output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="failed_upload_records.xlsx"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# Template download
# ─────────────────────────────────────────────────────────────────────────────

def download_combined_customer_template(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    COLUMNS = [
        ("Customer", 22), ("Phone", 15), ("Email", 26),
        ("State", 18), ("City", 20), ("Area", 18),
        ("Company", 18), ("Address", 22), ("Pincode", 10),
        ("Location Link", 28), ("Service Start", 14), ("Accounting", 30),
        ("Previous Plan", 16), ("Upgrade Date", 14), ("Service", 16),
        ("Advance Amt", 14), ("Outstanding Amount", 18), ("Custom Amount", 14),
        ("GST Registered", 14), ("GST Number", 20), ("Tax Percentage", 14),
        ("Rate", 12), ("Created", 14), ("Remarks", 22),
        ("Camera Serials", 35), ("Status", 14),
    ]

    HDR_FILL    = PatternFill("solid", fgColor="1F4E79")
    CAM_FILL    = PatternFill("solid", fgColor="145A32")
    STATUS_FILL = PatternFill("solid", fgColor="6E2FA0")
    HDR_FONT    = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    THIN        = Side(style="thin", color="CCCCCC")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Upload"

    col_index = {}
    for ci, (name, width) in enumerate(COLUMNS, 1):
        fill = CAM_FILL if name == "Camera Serials" else STATUS_FILL if name == "Status" else HDR_FILL
        c = ws.cell(1, ci, name)
        c.fill = fill; c.font = HDR_FONT; c.alignment = CENTER; c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
        col_index[name] = ci

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    HINT_FONT  = Font(italic=True, color="888888", size=9, name="Arial")
    HINT_ALIGN = Alignment(horizontal="left", vertical="center")
    HINTS = {
        "Customer"       : "Any name — repeat same name+phone for extra locations",
        "Phone"          : "10-digit, no spaces — repeat same for extra locations",
        "Email"          : "Optional — auto-generated if blank",
        "State"          : "e.g. Maharashtra",
        "City"           : "e.g. Pune",
        "Area"           : "Optional locality",
        "Company"        : "Must already exist in system",
        "Address"        : "Location address (fill on every row)",
        "Pincode"        : "6-digit PIN",
        "Location Link"  : "Google Maps URL",
        "Service Start"  : "DD/MM/YYYY",
        "Accounting"     : "Subscription plan name",
        "GST Registered" : "GST / Non-GST / Yes / No",
        "GST Number"     : "Exactly 15 characters",
        "Camera Serials" : "e.g. SN001, SN002, SN003",
        "Status"         : "Active / Inactive",
    }
    for col_name, hint_text in HINTS.items():
        if col_name in col_index:
            cell = ws.cell(2, col_index[col_name], hint_text)
            cell.font = HINT_FONT; cell.alignment = HINT_ALIGN

    ws.row_dimensions[2].height = 16

    output = io.BytesIO()
    wb.save(output); output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="bulk_upload_template.xlsx"'
    return response


from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.db import transaction
from django.db.models import F
from decimal import Decimal
import json
from datetime import datetime

from .models import data_logs, BillingRecord, Customer, CustomerLocation

PAYMENT_MODES = ['cash', 'upi', 'card', 'netbanking', 'cheque']


@require_http_methods(['GET', 'POST'])
def edit_payment_approval(request, log_id):

    user, role_permissions = get_logged_in_user(request)
    if not user:
        return redirect('login')
    if not role_permissions.payment_approval_e:
        return JsonResponse({'success': False, 'message': 'Permission denied.'}, status=403)

    log = get_object_or_404(data_logs, id=log_id, is_payment=True)

    # GET - feed the modal
    if request.method == 'GET':
        billing = BillingRecord.objects.filter(pk=log.billing_record_id).first()
        return JsonResponse({
            'success': True,
            'log': {
                'id':             log.id,
                'payment_amount': str(log.payment_amount or ''),
                'payment_date':   log.payment_date.strftime('%Y-%m-%d') if log.payment_date else '',
                'payment_mode':   log.payment_mode or '',
                'transaction_id': log.transaction_id or '',
                'receipt_number': log.receipt_number or '',
                'payment_notes':  log.payment_notes or '',
                'is_approved':    log.is_approved,
                'bill_total':     str(billing.amount)                           if billing else '0.00',
                'bill_paid':      str(billing.paid_amount   or Decimal('0.00')) if billing else '0.00',
                'bill_balance':   str(billing.balance_amount or Decimal('0.00')) if billing else '0.00',
            },
        })

    # POST - parse
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        body = request.POST

    new_amount_str = str(body.get('payment_amount', '')).strip()
    new_date_str   = str(body.get('payment_date',   '')).strip()
    new_mode       = str(body.get('payment_mode',   '')).strip()
    new_txn        = str(body.get('transaction_id', '')).strip()
    new_receipt    = str(body.get('receipt_number', '')).strip()
    new_notes      = str(body.get('payment_notes',  '')).strip()

    if not new_amount_str:
        return JsonResponse({'success': False, 'message': 'Payment amount is required.'})
    if not new_date_str:
        return JsonResponse({'success': False, 'message': 'Payment date is required.'})
    if new_mode and new_mode not in PAYMENT_MODES:
        return JsonResponse({'success': False, 'message': 'Invalid payment mode.'})

    try:
        new_amount = Decimal(new_amount_str)
        if new_amount <= 0:
            raise ValueError
    except Exception:
        return JsonResponse({'success': False, 'message': 'Invalid payment amount.'})

    try:
        new_date = datetime.strptime(new_date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'success': False, 'message': 'Invalid date format.'})

    with transaction.atomic():

        # Fresh DB fetch with row locks
        log     = data_logs.objects.select_for_update().get(pk=log_id)
        billing = BillingRecord.objects.select_for_update().get(pk=log.billing_record_id) if log.billing_record_id else None

        old_amount = log.payment_amount or Decimal('0.00')

      

        # STEP 1 - update the log
        log.payment_amount = new_amount
        log.payment_date   = new_date
        log.payment_mode   = new_mode if new_mode else log.payment_mode
        log.transaction_id = new_txn
        log.receipt_number = new_receipt
        log.payment_notes  = new_notes
        log.save()


        if billing:
            new_paid    = new_amount
            new_balance = max(Decimal('0.00'), billing.amount - new_paid)

           

            billing.paid_amount    = new_paid
            billing.balance_amount = new_balance
            billing.paid           = (new_balance == Decimal('0.00'))
            billing.payment_date   = new_date
            billing.payment_mode   = new_mode if new_mode else billing.payment_mode
            billing.transaction_id = new_txn if new_txn else billing.transaction_id
            billing.save(update_fields=[
                'paid_amount', 'balance_amount', 'paid',
                'payment_date', 'payment_mode', 'transaction_id',
            ])

            billing.refresh_from_db()
           
            log.total_paid     = new_paid
            log.balance_amount = new_balance
            log.save(update_fields=['total_paid', 'balance_amount'])

          
            diff = new_amount - old_amount
            print(f"[STEP 3] outstanding diff={diff}  (new {new_amount} - old {old_amount})")

            if log.location_id:
                before = CustomerLocation.objects.get(pk=log.location_id).outstanding_amount
                CustomerLocation.objects.filter(pk=log.location_id).update(
                    outstanding_amount=F('outstanding_amount') - diff
                )
                CustomerLocation.objects.filter(pk=log.location_id, outstanding_amount__lt=0).update(
                    outstanding_amount=Decimal('0.00')
                )
                after = CustomerLocation.objects.get(pk=log.location_id).outstanding_amount
                print(f"[STEP 3] location outstanding {before} -> {after}")

            elif log.customer_id:
                before = Customer.objects.get(pk=log.customer_id).outstanding_amount
                Customer.objects.filter(pk=log.customer_id).update(
                    outstanding_amount=F('outstanding_amount') - diff
                )
                Customer.objects.filter(pk=log.customer_id, outstanding_amount__lt=0).update(
                    outstanding_amount=Decimal('0.00')
                )
                after = Customer.objects.get(pk=log.customer_id).outstanding_amount
                print(f"[STEP 3] customer outstanding {before} -> {after}")
        else:
            print("[STEP 2] SKIPPED - no billing record")

    if billing:
        billing.refresh_from_db()
       

    return JsonResponse({
        'success': True,
        'message': 'Payment log updated successfully.',
        'updated': {
            'bill_total':      str(billing.amount)         if billing else None,
            'bill_paid':       str(billing.paid_amount)    if billing else None,
            'bill_balance':    str(billing.balance_amount) if billing else None,
            'bill_fully_paid': billing.paid                if billing else None,
        },
    })


@require_POST
def delete_payment_approval(request, log_id):

    user, role_permissions = get_logged_in_user(request)
    if not user:
        return JsonResponse({'success': False, 'message': 'Not authenticated.'}, status=401)
    if not role_permissions.payment_approval_e:
        return JsonResponse({'success': False, 'message': 'Permission denied.'}, status=403)

    log = get_object_or_404(data_logs, id=log_id, is_payment=True)

    with transaction.atomic():

        log     = data_logs.objects.select_for_update().get(pk=log_id)
        billing = BillingRecord.objects.select_for_update().get(pk=log.billing_record_id) if log.billing_record_id else None
        amount  = log.payment_amount or Decimal('0.00')

        print(f"\n[DELETE] log #{log_id}  amount={amount}  billing={billing}")

        if billing:
            # Remove this payment from received total
            new_paid    = max(Decimal('0.00'), (billing.paid_amount or Decimal('0.00')) - amount)
            new_balance = max(Decimal('0.00'), billing.amount - new_paid)

            print(f"[DELETE] new_paid={new_paid}  new_balance={new_balance}")

            billing.paid_amount    = new_paid
            billing.balance_amount = new_balance
            billing.paid           = (new_balance == Decimal('0.00'))

            if new_paid == Decimal('0.00'):
                billing.payment_date   = None
                billing.payment_mode   = None
                billing.transaction_id = None

            billing.save(update_fields=[
                'paid_amount', 'balance_amount', 'paid',
                'payment_date', 'payment_mode', 'transaction_id',
            ])

            if log.location_id:
                CustomerLocation.objects.filter(pk=log.location_id).update(
                    outstanding_amount=F('outstanding_amount') + amount
                )
            elif log.customer_id:
                Customer.objects.filter(pk=log.customer_id).update(
                    outstanding_amount=F('outstanding_amount') + amount
                )

        log.delete()
        print(f"[DELETE] done")

    return JsonResponse({'success': True, 'message': 'Payment log deleted successfully.'})